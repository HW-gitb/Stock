#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EGS v7.11 量化选股框架 — 周频增强版 + as-of backtest mode (2026.07.16)

v7.10 changelog (从 v7.9):
- ESP 极端低基数增长 cap：esp_raw_w 上限 200，降低低基数同比暴涨对排序的污染
- backtest Tier2 filler 排除未知行业，避免为凑满 watch_n 混入无行业参照系样本
- 输出 downgrade_reasons / score_penalty_reasons，便于回测和复盘追踪过滤原因

v7.9 changelog (从 v7.8):
- data_quality.completeness_score 改为按候选股实际字段缺失动态计算，不再硬编码 60

v7.8 changelog (从 v7.7):
- 修复 SW L1 行业映射全为“未知”的问题：Tushare L2 parent_code 对应 L1 industry_code，不是 index_code
- SW 行业映射缓存升级为 v6，隔离 v4 薄覆盖坏缓存和 v5 L1 缺失缓存

v7.7 changelog (从 v7.6):
- 修复 get_sw_industry_map() 在 Tushare 返回不完整时静默产出 ~75 条破损映射的 bug
- 加 SW_INDUSTRY_MIN_ACTIVE=3000 + index_member_all/index_member/L2 分批 三段回退
- cache key bump v4→v5 invalidate 旧坏缓存
- L3 PIT 三模式 (--l3-mode pit/today/neutralize) + state/l3_snapshots/ snapshot 累积
- analysis_input.source 加版本化 HiThink provider/complete-catalog/main-board lineage（schema 1.2.0）
- 影响：所有 v7.6 候选池逻辑上失效，需重生成才能下统计结论
============================================================
v7.5 新增（周频优化 6 项）：
  ✅ pct_60d：60日趋势，区分短脉冲与中期趋势
  ✅ drawdown_20d：20日高点回撤，识别过热后滑坡
  ✅ overheat_flag：5日涨幅>阈值 或 20日涨幅过大且仍高位 → Tier1降级+标注
  ✅ entry_flag：周一确认提示（可直接观察/需周一确认/高开过多暂停/资金流背离/题材过热）
  ✅ 双层输出：Top15候选池（含风险标注）+ Top5最终推荐
  ✅ 周度追踪扩展：记录Top15、周内高低点、最大收益/回撤、是否仍在池

修复清单（沿用 v7.4 全部 24 项）：

  ── 崩溃修复（4项）
  ✅ get_daily_basic：开盘前可回退至前3个官方交易日，并保留真实 source date
  ✅ get_unlock_future：分母缺失时 unknown/blocked，不猜测解禁比例
  ✅ get_suspend_info：无日线数据时阻断，不把未知源当成空停牌集
  ✅ filter_l0：pct_20d 全 NaN 时跳过过滤，不输出空表

  ── 逻辑修复（10项）
  ✅ score_l1 ITF-ADJ：用原始 pe 列（pe_n 在 l2 才创建）
  ✅ score_l1 groupby.apply：兼容 pandas 2.2+ include_groups
  ✅ score_l1 reset_index：.rename().reset_index() 兼容 pandas 2.x
  ✅ score_l2 DATA-INC 误标：行业中位数≤0 时不打 DATA-INC
  ✅ score_l4 TIER2_FORCED：移至所有加分完成后判定
  ✅ score_l4 ind_mom_cnt：groupby().count()+map 修复 pandas 2.x
  ✅ score_l2 esp_raw 允许负值（v7.3 双侧分布修复）
  ✅ get_financial_data df_inc：降序保留最新公告
  ✅ score_l5 同分次要排序：l4_score → pct_20d_n 确定性排序
  ✅ score_l5 行业未知→Tier2：l2_name=未知 股票降级

  ── 数据修复（4项）
  ✅ circ_share 单位：去掉错误的 /10000（万元/元=万股）
  ✅ get_moneyflow：trade_amount 本地从各档位买卖额推算
  ✅ get_financial_data：income 不支持批量，改用
     fina_indicator 的 dtprofit_to_profit + profit_dedt
  ✅ get_sw_industry_map：parent_code 三重格式匹配

  ── v7.3 框架对齐（6项）
  ✅ z2s 精确映射：-2.5/-0.5/0/1/2/3 六档（消除 81% 集中问题）
  ✅ ALPHA 加分 +10（原 +5）
  ✅ esp_raw > 0 硬条件 + 财务覆盖率 ≥ 70% 保护门槛
  ✅ top_n = 50（Phase 2 未接入保持 50 只）
  ✅ L2 行业截断 _cap_l2（>20 只截至 15）
  ✅ safe_api 异常信息打印

  ── 已知限制（永久冻结）
  ✅ L3 催化剂：B档虚拟概念指数合成，已接入
  ⬜ L6 盘中动态止损：需实时分钟数据，Tushare 不提供
  ⬜ 调研次数扣分：无可用数据源
  ⬜ L1 行业毛利率趋势：固定免检 0.5 分
"""
import argparse, os, sys, time, pickle, logging, warnings, shutil, uuid, re, hashlib
from contextlib import contextmanager, nullcontext
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
try:
    import tushare as ts
except ModuleNotFoundError:
    if any(arg in ("-h", "--help") for arg in sys.argv[1:]):
        ts = None
    else:
        raise
try:
    from tqdm import tqdm
except ModuleNotFoundError:
    if any(arg in ("-h", "--help") for arg in sys.argv[1:]):
        def tqdm(iterable=None, *args, **kwargs):
            return iterable if iterable is not None else []
    else:
        raise

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════
# §0 配置
# ═══════════════════════════════════════════════════
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
RESULT_DIR = os.path.join(SCRIPT_DIR, "Result")
LOG_DIR = os.path.join(PROJECT_ROOT, "logs")
L3_SNAPSHOT_DIR = os.path.join(PROJECT_ROOT, "state", "l3_snapshots")
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from engine.data.analysis_input_contract import (
    build_a_short_run_identity,
    validate_analysis_input_contract,
    validate_json_schema,
)
from engine.a_short_observability import safe_exception_summary
from engine.data.a_share_board_scope import is_a_share_main_board
from engine.egs_industry_heat import (
    compute_industry_heat_score, get_active_weights, final_score_and_tier,
    select_profile_watch_pool, write_weight_comparison, load_governance,
)
from engine.a_short_industry_theme import (
    classify_industry_trend, taxonomy_by_code, unavailable_theme_taxonomy,
)
from engine.a_short_legacy_llm_tasks import build_task_configs
from engine.a_short_hithink_l3 import (
    SOURCE_ID as HITHINK_L3_SOURCE_ID,
    HiThinkL3SourceError,
    MIN_CONCEPT_CATALOG_BOARD_COUNT,
    catalog_digest,
    fetch_complete_concept_graph,
)
from engine.a_short_run_paths import weight_comparison_path
from engine.a_short_runtime_config import load_runtime_configuration, runtime_configuration_lineage
from engine.a_share_market_clock import a_share_market_date, a_share_market_wall_time
from engine.a_short_tushare_client import init_tushare_pro
from engine.a_short_rule6_contract import (
    RULE6_CONDITIONAL_NA_REASONS,
    RULE6_D_TIER_REASONS,
    validate_rule6_check_contract,
)
from engine.a_short_rule6_evaluation import (
    evaluate_ar_growth_gt_revenue_growth,
    evaluate_block_trade_discount,
    evaluate_cash_debt_double_high,
    evaluate_holder_below_5pct,
    evaluate_margin_extreme_accumulation,
    evaluate_short_selling_surge,
    evaluate_volume_stall,
)
from engine.a_short_delisting import derive_delisting_flags

TOKEN = os.environ.get("TUSHARE_TOKEN")
if not TOKEN and not any(arg in ("-h", "--help") for arg in sys.argv[1:]):
    raise RuntimeError("请先设置环境变量 TUSHARE_TOKEN：$env:TUSHARE_TOKEN = \"your_token\"")
# 建议通过环境变量设置：$env:TUSHARE_TOKEN = "your_token"

_RUNTIME_CONFIGURATION = load_runtime_configuration()
_SCREENING_THRESHOLDS = _RUNTIME_CONFIGURATION["screening"]
_PORTFOLIO_RISK_THRESHOLDS = _RUNTIME_CONFIGURATION["m67"]["portfolio_risk"]

CONF = {
    "request_delay":    0.42,
    "chunk_size":       150,
    "financial_chunk_size": 80,
    "financial_min_chunk":  10,
    "result_dir":       RESULT_DIR,
    "cache_dir":        os.path.join(RESULT_DIR, "egs_cache"),
    "cache_ttl":        20 * 3600,
    "cache_policy":     "enabled",  # runtest capsules force disabled: no read or write of EGS cache
    "l3_cache_mode":    "refresh",  # formal runs refresh L3; tests may set reuse
    "l3_allow_stale_cache": False,   # test-only override; stale reuse otherwise fails closed
    **_SCREENING_THRESHOLDS,
}

if TOKEN and ts is not None:
    # 直接把 token 传给 pro_api,**不调 ts.set_token**:set_token 会在 import 期写 ~/tk.csv(import 文件副作用 +
    # 沙箱/受限环境 PermissionError → 卡住只读单测),且 egs_main 全程用本地 `pro` 客户端、不依赖全局 token
    # (weekly pipeline 等都 api=pro)。共享初始化器同时 pin 已验证 endpoint 并拒绝版本/私有结构漂移。
    pro = init_tushare_pro(TOKEN, ts_module=ts)
else:
    pro = None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("EGS")

EGS_VERSION = "v7.12"
ANALYSIS_INPUT_SCHEMA_VERSION = "1.2.0"
SUSPEND_DAILY_COVERAGE_LOG_SCHEMA_VERSION = "1.0.0"
_LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION = None
_LAST_HARD_VETO_SOURCE_HEALTH = {
    "suspension": {"status": "unknown", "observed_at": None},
    "unlock": {"status": "unknown", "observed_at": None},
    "holder_reduction": {"status": "unknown", "observed_at": None},
}
_LAST_UNLOCK_DETAILS = {}


def _record_hard_veto_source(name, status, observed_at=None, **details):
    if status not in {"known_clear", "known_hit", "unknown"}:
        raise ValueError(f"invalid hard-veto source status: {status}")
    payload = {"status": status, "observed_at": observed_at}
    payload.update({k: _json_value(v) for k, v in details.items()})
    _LAST_HARD_VETO_SOURCE_HEALTH[name] = payload
    return payload
# Canonical list of Tushare API endpoints egs_main consumes during a
# screening run. Sourced here so downstream consumers (data_health,
# backtest_rank's data_lineage) read one truth via _current_egs_api_families()
# regex parsing. When adding a new Tushare call, update this list.
EGS_API_FAMILIES = [
    "daily", "adj_factor", "daily_basic", "fina_indicator", "index_daily",
    "moneyflow", "moneyflow_hsgt", "margin_detail",
    "share_float", "stk_holdertrade", "balancesheet", "block_trade", "stock_basic", "namechange", "trade_cal",
    "index_member_all", "index_member", "index_classify",
]
REALTIME_CACHE_TTL = CONF["cache_ttl"]
BACKTEST_CACHE_TTL = 10 * 365 * 24 * 3600
TODAY = a_share_market_date()
TODAY_DT = a_share_market_wall_time()


# ═══════════════════════════════════════════════════
# §1 工具函数
# ═══════════════════════════════════════════════════
def _cp(key):
    os.makedirs(CONF["cache_dir"], exist_ok=True)
    return os.path.join(CONF["cache_dir"], f"{key}.pkl")

def _rp(filename):
    os.makedirs(CONF["result_dir"], exist_ok=True)
    return os.path.join(CONF["result_dir"], filename)

def load_cache(key):
    if CONF.get("cache_policy") == "disabled":
        return None
    p = _cp(key)
    if not os.path.exists(p): return None
    if time.time() - os.path.getmtime(p) > CONF["cache_ttl"]: return None
    with open(p, "rb") as f: return pickle.load(f)

def save_cache(key, data):
    if CONF.get("cache_policy") == "disabled":
        return
    p = _cp(key)
    tmp = p + ".tmp"
    if os.path.exists(p):
        try:
            os.chmod(p, 0o666)
        except OSError:
            pass
    with open(tmp, "wb") as f:
        pickle.dump(data, f)
    os.replace(tmp, p)

def ensure_writable(path):
    if os.path.exists(path):
        try:
            os.chmod(path, 0o666)
        except OSError:
            pass

def write_csv_atomic(df, path, **kwargs):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    ensure_writable(path)
    tmp = path + ".tmp"
    df.to_csv(tmp, **kwargs)
    os.replace(tmp, path)

def write_json_atomic(path, data):
    import json as _json
    os.makedirs(os.path.dirname(path), exist_ok=True)
    ensure_writable(path)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        _json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


@contextmanager
def official_output_transaction(paths):
    """Restore every prior official surface if any publish step fails."""
    ordered = list(dict.fromkeys(os.path.abspath(str(path)) for path in paths))
    backups = {}
    absent = set()
    try:
        for path in ordered:
            if os.path.exists(path):
                backup = f"{path}.rollback-{uuid.uuid4().hex}"
                shutil.copy2(path, backup)
                backups[path] = backup
            else:
                absent.add(path)
        yield
    except BaseException:
        rollback_errors = []
        for path in ordered:
            try:
                backup = backups.get(path)
                if backup and os.path.exists(backup):
                    ensure_writable(path)
                    os.replace(backup, path)
                elif path in absent and os.path.exists(path):
                    os.remove(path)
            except OSError as exc:
                rollback_errors.append(f"{path}: {exc}")
        if rollback_errors:
            raise RuntimeError(
                "official EGS publish failed and rollback was incomplete: "
                + "; ".join(rollback_errors)
            )
        raise
    finally:
        for backup in backups.values():
            try:
                if os.path.exists(backup):
                    os.remove(backup)
            except OSError:
                pass

def _suspend_daily_coverage_log_path(as_of):
    return os.path.join(LOG_DIR, f"suspend_daily_coverage_{as_of}.json")

def _record_suspend_daily_coverage_observation(
    *,
    as_of,
    trade_date,
    status,
    stock_universe_count=None,
    daily_payload_row_count=None,
    traded_in_universe_count=None,
    suspended_count=None,
    coverage_ratio=None,
    min_coverage=None,
    attempted_trade_dates=None,
    source="tushare.pro.daily",
    message=None,
):
    global _LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION
    payload = {
        "schema_name": "suspend_daily_coverage_log",
        "schema_version": SUSPEND_DAILY_COVERAGE_LOG_SCHEMA_VERSION,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "as_of": str(as_of),
        "trade_date": str(trade_date) if trade_date is not None else None,
        "status": status,
        "source": source,
        "stock_universe_count": (
            int(stock_universe_count) if stock_universe_count is not None else None
        ),
        "daily_payload_row_count": (
            int(daily_payload_row_count) if daily_payload_row_count is not None else None
        ),
        "traded_in_universe_count": (
            int(traded_in_universe_count) if traded_in_universe_count is not None else None
        ),
        "suspended_count": int(suspended_count) if suspended_count is not None else None,
        "coverage_ratio": float(coverage_ratio) if coverage_ratio is not None else None,
        "min_coverage": float(min_coverage) if min_coverage is not None else None,
        "attempted_trade_dates": list(attempted_trade_dates or []),
        "message": message,
    }
    _LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION = payload
    try:
        write_json_atomic(_suspend_daily_coverage_log_path(as_of), payload)
    except Exception as exc:
        log.warning(f"suspend daily coverage log write failed: {type(exc).__name__}: {exc}")
    return payload

def _current_suspend_daily_coverage_observation():
    if _LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION is None:
        return {
            "schema_name": "suspend_daily_coverage_log",
            "schema_version": SUSPEND_DAILY_COVERAGE_LOG_SCHEMA_VERSION,
            "status": "not_observed",
            "message": "get_suspend_info has not produced a coverage observation in this process",
        }
    return dict(_LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION)

def save_ranked_xlsx(df, path, group_size=5):
    """
    另存带颜色分组的 Excel 观察池。
    CSV 不支持单元格颜色，因此保留 CSV 的同时额外输出 xlsx。
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl 未安装，跳过 egs_tier1.xlsx 彩色输出")
        return

    fills = [
        PatternFill("solid", fgColor="FFFF00"),  # 1-5 黄
        PatternFill("solid", fgColor="DAF2D0"),  # 6-10 浅绿
        PatternFill("solid", fgColor="CAEDFB"),  # 11-15 浅蓝
        PatternFill("solid", fgColor="FCE4D6"),  # 16-20 备用浅橙
    ]

    ensure_writable(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tier1")
        ws = writer.sheets["Tier1"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(2, ws.max_row + 1):
            fill = fills[((row_idx - 2) // group_size) % len(fills)]
            for cell in ws[row_idx]:
                cell.fill = fill

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 18)

def _json_value(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        return float(value)
    return value

def _json_float(value):
    value = _json_value(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def _json_score(value):
    value = _json_float(value)
    if value is None:
        return None
    return max(0.0, min(100.0, value))

def _json_int(value):
    value = _json_value(value)
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

def _json_bool(value):
    value = _json_value(value)
    if value is None:
        return None
    return bool(value)

def _json_str(value):
    value = _json_value(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)

def _eod_price_time(trade_date):
    if not trade_date or len(str(trade_date)) != 8:
        return None
    s = str(trade_date)
    return f"{s[:4]}-{s[4:6]}-{s[6:]}T15:00:00+08:00"

def _row_get(row, key, default=None):
    if key not in row.index:
        return default
    value = _json_value(row.get(key))
    return default if value is None else value


def _historical_replay_mode():
    """Whether the current EGS invocation is replaying an as-of date."""
    return TODAY < a_share_market_date()

def _board_from_code(ts_code):
    symbol = str(ts_code).split(".")[0]
    if symbol.startswith(("300", "301")):
        return "chinext"
    if symbol.startswith(("688", "689")):
        return "star"
    if symbol.startswith(("8", "4", "920")):
        return "bj"
    if is_a_share_main_board(ts_code):
        return "main"
    return "unknown"   # 防御纵深(Slice 1 P2-1):B 股(900/200)/畸形码不再默认 'main';filter_l0 strict 已在上游剔,此处兜手构/未来旁路

def _rule_check(check_id, name, group, status="unknown", severity="watch", metrics=None, notes=None):
    return {
        "id": check_id,
        "name": name,
        "group": group,
        "status": status,
        "severity": severity,
        "metrics": metrics or {},
        "evidence": [],
        "notes": notes,
    }

def _rule6_evaluated_check(check_id, name, group, evaluations):
    """Turn a pure evaluator result into the analysis-input Rule6 record."""
    evaluation = (evaluations or {}).get(check_id)
    if not isinstance(evaluation, dict):
        return _rule_check(check_id, name, group, "unknown", "watch",
                           notes="machine-checkable Rule6 source/evaluation unavailable")
    status = evaluation.get("status")
    severity = evaluation.get("severity")
    allowed_statuses = {"pass", "fail", "unknown"}
    if check_id in RULE6_CONDITIONAL_NA_REASONS:
        allowed_statuses = allowed_statuses | {"not_applicable"}
    if status not in allowed_statuses or severity not in {"none", "watch", "hard_veto"}:
        return _rule_check(check_id, name, group, "unknown", "watch",
                           notes="machine-checkable Rule6 evaluation malformed")
    return _rule_check(check_id, name, group, status, severity,
                       metrics=evaluation.get("metrics"), notes=evaluation.get("notes"))


def _candidate_from_row(row, rank, final_codes, latest_td, unlock_set, suspended_set,
                        industry_heat_governance=None, rule6_evaluations=None):
    ts_code = str(_row_get(row, "ts_code", ""))
    exchange = ts_code.split(".")[-1] if "." in ts_code else "SZ"
    close = _json_float(_row_get(row, "close"))
    high_20d = _json_float(_row_get(row, "high_20d"))
    low_20d = _json_float(_row_get(row, "low_20d"))
    avg_amount_5d = _json_float(_row_get(row, "avg_amount_5d"))
    avg_amount_20d = _json_float(_row_get(row, "avg_amount_20d"))
    turnover_rate = _json_float(_row_get(row, "turnover_rate"))
    reduce_deduct = _json_int(_row_get(row, "reduce_deduct", 0)) or 0
    has_crash_veto = _json_bool(_row_get(row, "has_crash_veto"))
    overheat_flag = _json_bool(_row_get(row, "overheat_flag"))
    chasing_high = _json_bool(_row_get(row, "chasing_high"))
    is_breakout = _json_bool(_row_get(row, "is_breakout"))
    is_lock = _json_bool(_row_get(row, "is_lock"))
    quote_source_date = str(_row_get(row, "source_trade_date", latest_td))
    unlock_detail = dict(_LAST_UNLOCK_DETAILS.get(ts_code) or {})
    suspension_health = dict(_LAST_HARD_VETO_SOURCE_HEALTH.get("suspension") or {})
    unlock_health = dict(_LAST_HARD_VETO_SOURCE_HEALTH.get("unlock") or {})
    reduction_health = dict(_LAST_HARD_VETO_SOURCE_HEALTH.get("holder_reduction") or {})
    delisting_flags = derive_delisting_flags(row, historical=_historical_replay_mode())

    rule6_checks = [
        _rule_check(
            "rule6_holder_reduction",
            "Major shareholder/controller/executive reduction",
            "pre_veto",
            "fail" if reduce_deduct else "pass",
            "hard_veto" if reduce_deduct else "none",
            {"reduce_deduct": reduce_deduct},
        ),
        _rule_check(
            "rule6_crash_veto",
            "Single-day crash without repair",
            "pre_veto",
            "fail" if has_crash_veto else "pass",
            "hard_veto" if has_crash_veto else "none",
            {"has_crash_veto": has_crash_veto},
        ),
        _rule6_evaluated_check("rule6_holder_below_5pct", "Holder stake reduced below 5%", "pre_veto", rule6_evaluations),
        # IV is materialized by the weekly IV feed immediately before the M6.7 gate.
        _rule6_evaluated_check("rule6_50etf_iv", "50ETF IV hard veto", "pre_veto", rule6_evaluations),
        _rule6_evaluated_check("rule6_cash_debt_double_high", "Cash and debt both abnormally high", "pre_veto", rule6_evaluations),
        _rule_check("rule6_regulatory_48h", "Regulatory inquiry or concern within 48h", "pre_veto",
                    "not_applicable", "review", notes=RULE6_D_TIER_REASONS["rule6_regulatory_48h"]),
        _rule_check("rule6_good_data_bad_reaction", "Good data bad reaction", "post_veto",
                    "not_applicable", "review", notes=RULE6_D_TIER_REASONS["rule6_good_data_bad_reaction"]),
        _rule6_evaluated_check("rule6_volume_stall", "Volume stall distribution", "post_veto", rule6_evaluations),
        _rule6_evaluated_check("rule6_margin_extreme_accumulation", "Margin extreme accumulation", "post_veto", rule6_evaluations),
        _rule6_evaluated_check("rule6_block_trade_discount", "Consecutive block-trade discount", "post_veto", rule6_evaluations),
        _rule_check("rule6_northbound_selloff", "Northbound consecutive selloff", "post_veto",
                    "not_applicable", "review", notes=RULE6_D_TIER_REASONS["rule6_northbound_selloff"]),
        _rule6_evaluated_check("rule6_short_selling_surge", "Short selling balance abnormal surge", "post_veto", rule6_evaluations),
        _rule6_evaluated_check("rule6_ar_growth_gt_revenue_growth", "AR growth faster than revenue growth", "post_veto", rule6_evaluations),
    ]
    validate_rule6_check_contract(rule6_checks)

    industry_trend_signal = classify_industry_trend(
        score=_row_get(row, "industry_heat_score"),
        sw_l2_code=_row_get(row, "l2_code"),
        sw_l2_name=_row_get(row, "l2_name"),
        source_as_of=quote_source_date,
        expected_as_of=latest_td,
        governance=industry_heat_governance or load_governance(),
    )
    theme_taxonomy = _row_get(row, "theme_taxonomy")
    if not isinstance(theme_taxonomy, dict):
        theme_taxonomy = unavailable_theme_taxonomy(
            str(latest_td),
            "l3_taxonomy_not_available_for_this_run",
            l3_provider=CONF.get("l3_provider"),
            l3_snapshot_date=CONF.get("l3_snapshot_date"),
            l3_coverage=CONF.get("l3_coverage"),
        )

    def _is_missing(value):
        return value is None or value == "" or value == "未知"

    core_quality_fields = [
        ("quote.close", close),
        ("quote.high_20d", high_20d),
        ("quote.low_20d", low_20d),
        ("quote.avg_amount_5d", avg_amount_5d),
        ("quote.avg_amount_20d", avg_amount_20d),
        ("quote.turnover_rate", turnover_rate),
        ("industry.sw_l1_name", _row_get(row, "l1_name")),
        ("industry.sw_l2_name", _row_get(row, "l2_name")),
        ("scores.final_score", _row_get(row, "final_score")),
        ("scores.esp_score", _row_get(row, "esp_score")),
        ("scores.cat_score", _row_get(row, "cat_score")),
        ("scores.l4_score", _row_get(row, "l4_score")),
        ("technical.pct_5d", _row_get(row, "pct_5d_n")),
        ("technical.pct_20d", _row_get(row, "pct_20d_n")),
        ("technical.pct_60d", _row_get(row, "pct_60d")),
        ("technical.drawdown_20d", _row_get(row, "drawdown_20d")),
        ("fundamental.q0_dt_yoy", _row_get(row, "q0_dt_yoy")),
        ("fundamental.q1_dt_yoy", _row_get(row, "q1_dt_yoy")),
        ("fundamental.pe_ttm", _row_get(row, "pe_ttm")),
        ("fundamental.pb", _row_get(row, "pb")),
        ("fundamental.roe", _row_get(row, "roe")),
        ("fundamental.total_mv", _row_get(row, "total_mv")),
        ("capital_flow.big_order_ratio", _row_get(row, "big_ratio")),
    ]
    actual_missing_fields = [name for name, value in core_quality_fields if _is_missing(value)]
    planned_unavailable_fields = [
        "market_context.volatility.iv_percentile_252d",
        "technical.atr.atr_14",
        "technical.moving_averages",
        "technical.rsi_14",
        "technical.macd",
        "capital_flow.northbound",
        "capital_flow.block_trade",
        "analyst.target_price_mean",
    ]
    missing_fields = actual_missing_fields + planned_unavailable_fields
    present_count = len(core_quality_fields) - len(actual_missing_fields)
    completeness_score = round((present_count / len(core_quality_fields)) * 100, 2)
    pending_fields = [
        "industry.industry_fundamental_trend",
        "event_risk.regulatory",
        "catalyst.policy_news",
    ]

    candidate = {
        "ts_code": ts_code,
        "name": str(_row_get(row, "name", "")),
        "exchange": exchange if exchange in ("SH", "SZ") else "SZ",
        "board": _board_from_code(ts_code),
        "analysis_role": "final" if ts_code in final_codes else "watch",
        "selection": {
            "rank": rank,
            "tier": str(_row_get(row, "tier", "Unknown")),
            "entry_flag": _json_value(_row_get(row, "entry_flag")),
            "cninfo_flag": _json_value(_row_get(row, "cninfo_flag")),
            "still_in_pool": True,
        },
        "quote": {
            "close": close,
            "current_price": close,
            "open": None,
            "high": None,
            "low": None,
            "pre_close": None,
            "pct_change": None,
            "price_source": "tushare_eod",
            "price_time": _eod_price_time(quote_source_date),
            "source_trade_date": quote_source_date,
            "adjustment": "qfq",
            "ex_rights_30d": None,
        },
        "industry": {
            "sw_l1_code": _json_str(_row_get(row, "l1_code")),
            "sw_l1_name": _json_value(_row_get(row, "l1_name")),
            "sw_l2_code": _json_str(_row_get(row, "l2_code")),
            "sw_l2_name": _json_value(_row_get(row, "l2_name")),
            "industry_trend": industry_trend_signal["industry_trend"],
            "industry_trend_signal": industry_trend_signal,
            "industry_fundamental_trend": "pending_llm",
            "industry_fundamental_trend_evidence": [],
        },
        "scores": {
            "final_score": _json_score(_row_get(row, "final_score")),
            "egs_base": _json_score(_row_get(row, "egs_base")),
            "esp_score": _json_score(_row_get(row, "esp_score")),
            "cat_score": _json_score(_row_get(row, "cat_score")),
            "l4_score": _json_score(_row_get(row, "l4_score")),
            "industry_heat_score": _json_score(_row_get(row, "industry_heat_score")),
            "l1_score": _json_float(_row_get(row, "l1_score")),
            "l2_flags": _json_value(_row_get(row, "l2_flags")),
            "l4_flag": _json_value(_row_get(row, "l4_flag")),
            "cat_flag": _json_value(_row_get(row, "cat_flag")),
            "deduct": _json_float(_row_get(row, "deduct")),
            "multiplier": _json_float(_row_get(row, "mult")),
        },
        "technical": {
            "pct_5d": _json_float(_row_get(row, "pct_5d")),
            "pct_5d_n": _json_float(_row_get(row, "pct_5d_n", _row_get(row, "pct_5d"))),
            "pct_20d": _json_float(_row_get(row, "pct_20d")),
            "pct_20d_n": _json_float(_row_get(row, "pct_20d_n", _row_get(row, "pct_20d"))),
            "pct_60d": _json_float(_row_get(row, "pct_60d")),
            "drawdown_20d": _json_float(_row_get(row, "drawdown_20d")),
            "high_20d": high_20d,
            "low_20d": low_20d,
            "avg_amount_5d": avg_amount_5d,
            "avg_amount_20d": avg_amount_20d,
            "yesterday_amount": None,
            "volume_ratio_5d": None,
            "amplitude": None,
            "close_position_in_range": None,
            "support": {"price": low_20d, "method": "20d_close_range", "confidence": "medium" if low_20d else "unknown"},
            "resistance": {"price": high_20d, "method": "20d_close_range", "confidence": "medium" if high_20d else "unknown"},
            "atr": {"atr_14": None, "atr_window": None, "ex_rights_adjusted": None},
            "moving_averages": {"ma5": None, "ma10": None, "ma20": None, "ma60": None},
            "rsi_14": None,
            "macd": {"dif": None, "dea": None, "hist": None},
            "bollinger": {"upper": None, "middle": None, "lower": None},
            "limit_up_count_10d": _json_int(_row_get(row, "limit_10d")),
            "limit_up_count_20d": _json_int(_row_get(row, "limit_20d")),
            "direction_lock": "breakout" if is_breakout else "unknown",
            "coarse_reward_risk": None,
            "precise_reward_risk": None,
        },
        "fundamental": {
            "valuation": {
                "pe": _json_float(_row_get(row, "pe")),
                "pe_ttm": _json_float(_row_get(row, "pe_ttm")),
                "pb": _json_float(_row_get(row, "pb")),
                "peg": _json_float(_row_get(row, "peg_n")),
                "total_mv": _json_float(_row_get(row, "total_mv")),
                "circ_mv": _json_float(_row_get(row, "circ_mv")),
                "val_bonus": _json_float(_row_get(row, "val_bonus")),
                "val_penalty": _json_float(_row_get(row, "val_penalty")),
            },
            "profitability": {
                "roe": _json_float(_row_get(row, "roe")),
                "q0_dt_yoy": _json_float(_row_get(row, "q0_dt_yoy")),
                "q1_dt_yoy": _json_float(_row_get(row, "q1_dt_yoy")),
                "q0_profit_dedt": _json_float(_row_get(row, "q0_profit_dedt")),
                "q0_net_income": _json_float(_row_get(row, "q0_net_income")),
                "ttm_profit_dedt": _json_float(_row_get(row, "ttm_profit_dedt")),
            },
            "quality": {
                "ttm_ocf_ratio": _json_float(_row_get(row, "ttm_ocf_ratio")),
                "q0_dt_profit_ratio": _json_float(_row_get(row, "q0_dt_profit_ratio")),
                "cash_debt_high_risk": None,
                "ar_growth_gt_revenue_growth_2q": None,
                "contract_liability_exemption": None,
            },
            "expectation": {
                "ind_median_profit_growth": _json_float(_row_get(row, "ind_med")),
                "esp_raw": _json_float(_row_get(row, "esp_raw")),
                "consensus_profit_growth": None,
                "earnings_report_date": None,
                "earnings_less_than_24h": None,
                "good_data_bad_reaction": None,
                "guidance_credibility_damaged": None,
            },
        },
        "capital_flow": {
            "moneyflow": {
                "big_order_ratio": _json_float(_row_get(row, "big_ratio")),
                "net_inflow_5d": None,
                "divergence_flag": True if (_json_float(_row_get(row, "big_ratio")) or 0) < -0.05 else None,
            },
            "margin": {
                "balance": None,
                "balance_change_5d_pct": None,
                "balance_change_10d_pct": None,
                "balance_to_float_mv_pct": None,
                "extreme_accumulation": None,
            },
            "northbound": {
                "holding_ratio": None,
                "consecutive_net_sell_days": None,
                "net_sell_to_total_share_pct": None,
            },
            "block_trade": {
                "discount_trade_count_10d": None,
                "avg_discount_pct": None,
                "amount_10d": None,
            },
        },
        "event_risk": {
            "rule6_checks": rule6_checks,
            "regulatory": {
                "has_inquiry_or_concern_48h": None,
                "negative_depth": "pending_llm",
                "evidence": [],
            },
            "holder_reduction": {
                "active_plan": True if reduce_deduct else False,
                "completed_3m_pct_share": None,
                "completed_3m_amount": None,
                "reduce_penalty": _json_float(_row_get(row, "reduce_penalty")),
                "source_status": reduction_health.get("status", "unknown"),
                "observed_at": reduction_health.get("observed_at"),
            },
            "unlock": {
                "unlock_pct": unlock_detail.get("unlock_pct"),
                "unlock_date": unlock_detail.get("unlock_date"),
                "large_unlock_flag": ts_code in unlock_set,
                "source_status": unlock_detail.get("status", unlock_health.get("status", "unknown")),
                "observed_at": unlock_detail.get("observed_at", unlock_health.get("observed_at")),
                "denominator": unlock_detail.get("denominator"),
            },
            "suspension": {
                "is_suspended": ts_code in suspended_set,
                "recent_suspension_5d": None,
                "source_status": suspension_health.get("status", "unknown"),
                "observed_at": suspension_health.get("observed_at"),
            },
            "delisting": {
                "st_flag": delisting_flags["st_flag"],
                "delisting_warning": delisting_flags["delisting_warning"],
                "non_standard_audit": None,
                "negative_net_asset": None,
            },
        },
        "catalyst": {
            "concepts": [],
            "concept_strength_score": _json_score(_row_get(row, "cat_score")),
            "theme_taxonomy": theme_taxonomy,
            "policy_news": [],
            "earnings": {"has_recent_report": None, "is_primary_catalyst": None},
            "time_window": "unknown",
        },
        "liquidity": {
            "avg_amount_5d": avg_amount_5d,
            "avg_amount_20d": avg_amount_20d,
            "turnover_rate": turnover_rate,
            "yesterday_amount": None,
            "spread_pct": None,
            "one_minute_capacity": None,
            "position_amount_cap": None,
            "split_order_required": None,
        },
        "volatility": {
            "hv_252d": None,
            "iv_hv_ratio": None,
            "iv_hv_position_cut_pct": None,
        },
        "analyst": {
            "coverage_count": None,
            "target_price_mean": None,
            "downgrade_count_1m": None,
            "target_below_current": None,
        },
        "portfolio_impact": {
            "same_sw_l2_exposure_after_buy_pct": None,
            "factor_exposures": [
                {"factor": "sw_l2_industry", "value": None,
                 "threshold": _PORTFOLIO_RISK_THRESHOLDS["same_sw_l2_threshold_pct"], "status": "unknown"},
                {"factor": "northbound_holding_ratio", "value": None,
                 "threshold": _PORTFOLIO_RISK_THRESHOLDS["northbound_threshold_pct"], "status": "unknown"},
                {"factor": "margin_balance_to_float_mv", "value": None,
                 "threshold": _PORTFOLIO_RISK_THRESHOLDS["margin_threshold_pct"], "status": "unknown"},
                {"factor": "index_component", "value": None,
                 "threshold": _PORTFOLIO_RISK_THRESHOLDS["large_index_threshold_pct"], "status": "unknown"},
                {"factor": "small_float_mv", "value": None,
                 "threshold": _PORTFOLIO_RISK_THRESHOLDS["small_float_mv_threshold_pct"], "status": "unknown"},
            ],
            "correlation_action": "unknown",
        },
        "derived_flags": {
            "chasing_high": chasing_high,
            "overheat_flag": overheat_flag,
            "has_crash_veto": has_crash_veto,
            "is_lock": is_lock,
            "is_breakout": is_breakout,
            "vol_confirm": _json_bool(_row_get(row, "vol_confirm")),
            "m4_review_required": None,
            "hard_veto": bool(reduce_deduct or has_crash_veto),
        },
        "llm_tasks": [],
        "data_quality": {
            "completeness_score": _json_score(completeness_score),
            "missing_fields": missing_fields,
            "pending_fields": pending_fields,
            "rule11_required": False,
        },
    }
    candidate["llm_tasks"] = build_task_configs(candidate, latest_td)
    return candidate

def _code_set(value):
    if value is None:
        return set()
    if isinstance(value, pd.DataFrame):
        if "ts_code" not in value.columns:
            return set()
        series = value["ts_code"]
    elif isinstance(value, pd.Series):
        series = value
    else:
        try:
            return {str(item) for item in value if item is not None and str(item).strip()}
        except TypeError:
            return set()
    return {
        str(item) for item in series.dropna().tolist()
        if str(item).strip()
    }


def build_rank_universe_reconciliation(df_l0, stages, sources):
    """Account for every post-L0 symbol and expose source truncation.

    ``stages`` entries are ``(name, dataframe, expected_exclusion, reason)``.
    Scoring-only joins must set ``expected_exclusion=False`` so any row loss
    becomes a publish-blocking error instead of disappearing silently.
    ``sources`` entries are ``(requested, available, min_coverage)``.
    """
    l0_codes = _code_set(df_l0)
    l0_duplicate_count = (
        int(df_l0["ts_code"].duplicated().sum())
        if isinstance(df_l0, pd.DataFrame) and "ts_code" in df_l0.columns else 0
    )
    active_codes = set(l0_codes)
    terminal = {}
    stage_counts = []
    expected_excluded_count = 0
    unexpected_excluded_count = 0
    unexpected_added_count = 0
    duplicate_code_count = l0_duplicate_count

    for stage_name, stage_df, expected_exclusion, reason in stages:
        stage_codes = _code_set(stage_df)
        stage_duplicate_count = (
            int(stage_df["ts_code"].duplicated().sum())
            if isinstance(stage_df, pd.DataFrame) and "ts_code" in stage_df.columns else 0
        )
        duplicate_code_count += stage_duplicate_count
        added = stage_codes - active_codes
        excluded = active_codes - stage_codes
        if expected_exclusion and isinstance(reason, dict):
            classified_excluded = {code for code in excluded if code in reason}
            expected_excluded_count += len(classified_excluded)
            unexpected_excluded_count += len(excluded - classified_excluded)
        elif expected_exclusion:
            expected_excluded_count += len(excluded)
        else:
            unexpected_excluded_count += len(excluded)
        unexpected_added_count += len(added)
        for ts_code in excluded:
            terminal_reason = (
                reason.get(ts_code, "stage_exclusion_unclassified")
                if isinstance(reason, dict) else reason
            )
            terminal[ts_code] = {
                "outcome": "excluded",
                "terminal_stage": stage_name,
                "reason": terminal_reason,
            }
        stage_counts.append({
            "stage": stage_name,
            "input_count": len(active_codes),
            "output_count": len(stage_codes & l0_codes),
            "excluded_count": len(excluded),
            "added_count": len(added),
            "expected_exclusion": bool(expected_exclusion),
        })
        active_codes = stage_codes & l0_codes

    final_stage = stage_counts[-1]["stage"] if stage_counts else "l0"
    for ts_code in active_codes:
        terminal[ts_code] = {
            "outcome": "ranked",
            "terminal_stage": final_stage,
            "reason": "ranked",
        }
    unaccounted_codes = l0_codes - set(terminal)
    detail = pd.DataFrame([
        {"ts_code": ts_code, **terminal[ts_code]}
        for ts_code in sorted(terminal)
    ], columns=["ts_code", "outcome", "terminal_stage", "reason"])
    # Preserve the pre-filter feature surface for comparison-only consumers.
    # The ranked CSV intentionally excludes l2_crash_veto members, so it cannot
    # be the sole source for later matched-control evidence.
    feature_columns = [
        "name", "l1_name", "l2_name", "total_mv", "pct_20d", "avg_amount_20d",
    ]
    if isinstance(df_l0, pd.DataFrame) and "ts_code" in df_l0.columns:
        available = [column for column in feature_columns if column in df_l0.columns]
        if available:
            l0_features = df_l0[["ts_code", *available]].copy()
            l0_features["ts_code"] = l0_features["ts_code"].astype(str)
            detail = detail.merge(l0_features.drop_duplicates("ts_code"), on="ts_code", how="left")

    source_coverage = {}
    source_coverage_failure_count = 0
    for source_name, source_spec in (sources or {}).items():
        requested, available, min_coverage = source_spec
        requested_codes = _code_set(requested)
        available_codes = _code_set(available)
        covered_count = len(requested_codes & available_codes)
        requested_count = len(requested_codes)
        missing_count = requested_count - covered_count
        coverage_ratio = covered_count / requested_count if requested_count else 1.0
        source_status = "pass" if coverage_ratio >= float(min_coverage) else "fail"
        if source_status == "fail":
            source_coverage_failure_count += 1
        source_coverage[source_name] = {
            "requested_count": requested_count,
            "covered_count": covered_count,
            "missing_count": missing_count,
            "coverage_ratio": float(coverage_ratio),
            "min_coverage": float(min_coverage),
            "status": source_status,
        }

    accounted_count = len(terminal)
    accounting_balanced = (
        accounted_count == len(l0_codes)
        and len(active_codes) + expected_excluded_count + unexpected_excluded_count == len(l0_codes)
        and not unaccounted_codes
    )
    unexpected_stage_change_count = unexpected_excluded_count + unexpected_added_count
    status = "pass"
    if (
        not accounting_balanced
        or unexpected_stage_change_count
        or duplicate_code_count
        or source_coverage_failure_count
    ):
        status = "fail"
    summary = {
        "status": status,
        "l0_count": len(l0_codes),
        "ranked_count": len(active_codes),
        "expected_excluded_count": expected_excluded_count,
        "unexpected_excluded_count": unexpected_excluded_count,
        "unexpected_added_count": unexpected_added_count,
        "unexpected_stage_change_count": unexpected_stage_change_count,
        "accounted_count": accounted_count,
        "unaccounted_count": len(unaccounted_codes),
        "duplicate_code_count": duplicate_code_count,
        "accounting_balanced": accounting_balanced,
        "detail_row_count": int(len(detail)),
        "source_coverage_failure_count": source_coverage_failure_count,
        "stage_counts": stage_counts,
        "source_coverage": source_coverage,
    }
    return summary, detail


def _rank_stage_excluded_count(rank_reconciliation, stage_name):
    for item in (rank_reconciliation or {}).get("stage_counts", []):
        if item.get("stage") == stage_name:
            return int(item.get("excluded_count", 0))
    return 0


def export_analysis_input(df_full, watch_df, tier1_final, latest_td, trade_dates,
                          unlock_set, suspended_set, relisted_set, red_dict,
                          tier1_csv_path, full_csv_path, output_root=None,
                          rank_reconciliation=None, rule6_evaluations_by_code=None):
    import json as _json

    project_root = os.path.dirname(SCRIPT_DIR)
    if output_root:
        base_root = output_root if os.path.isabs(output_root) else os.path.join(project_root, output_root)
    else:
        base_root = os.path.join(project_root, "result", "a_short")
    out_dir = os.path.join(base_root, latest_td)
    os.makedirs(out_dir, exist_ok=True)

    final_codes = set(tier1_final.head(CONF["final_n"]).get("ts_code", pd.Series(dtype=str)).tolist()) \
        if tier1_final is not None and not tier1_final.empty else set()

    industry_heat_governance = load_governance()
    candidates = [
        _candidate_from_row(
            row, rank, final_codes, latest_td, unlock_set, suspended_set,
            industry_heat_governance,
            rule6_evaluations=(rule6_evaluations_by_code or {}).get(str(_row_get(row, "ts_code", ""))),
        )
        for rank, (_, row) in enumerate(watch_df.iterrows(), start=1)
    ]
    run_identity = build_a_short_run_identity(latest_td, candidates)

    analysis_input = {
        "schema_name": "analysis_input",
        "schema_version": ANALYSIS_INPUT_SCHEMA_VERSION,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "trade_date": latest_td,
        "preset": "a_short",
        "market": "A",
        "horizon": "short",
        "source": {
            "screening_engine": "egs_main.py",
            "screening_engine_version": EGS_VERSION,
            "data_provider": (
                "mixed" if CONF.get("l3_provider") == HITHINK_L3_SOURCE_ID else "tushare"
            ),
            "runtime_configuration": runtime_configuration_lineage(_RUNTIME_CONFIGURATION),
            "l3_mode": CONF.get("l3_mode", "today"),
            "l3_pit_strict": bool(CONF.get("l3_pit_strict", False)),
            "l3_snapshot_date": CONF.get("l3_snapshot_date"),
            "l3_provider": CONF.get("l3_provider"),
            "l3_coverage": CONF.get("l3_coverage"),
            "hard_veto_source_health": {
                name: dict(_LAST_HARD_VETO_SOURCE_HEALTH.get(name) or {})
                for name in ("suspension", "unlock", "holder_reduction")
            },
            "run_identity": run_identity,
            "input_files": [
                {"role": "watch_pool", "path": os.path.relpath(tier1_csv_path, project_root), "sha256": None},
                {"role": "full_rank", "path": os.path.relpath(full_csv_path, project_root), "sha256": None},
            ],
            "notes": [
                f"Generated by EGS {EGS_VERSION} Phase 1b exporter.",
                "Fields marked pending_data or pending_llm are reserved for analyzer/state/Skill phases.",
            ],
        },
        "universe_summary": {
            "listed_count": None,
            "after_l0_count": (
                int(rank_reconciliation["l0_count"])
                if rank_reconciliation is not None else None
            ),
            "full_count": int(len(df_full)),
            "watch_count": int(len(watch_df)),
            "final_count": int(min(len(tier1_final), CONF["final_n"])) if tier1_final is not None else 0,
            "excluded_counts": {
                "unlock": int(len(unlock_set or [])),
                "suspended": int(len(suspended_set or [])),
                "relisted": int(len(relisted_set or [])),
                "holder_reduction_veto_10d": int(len((red_dict or {}).get("veto_10d", set()))),
            },
            # 排名层淘汰不是 L0 硬否决；独立存放，避免 weekly exclusion_summary 把它们误当
            # 上游硬过滤原因。旧 v1.4 混装产物由 weekly 对明确键做兼容，不放松其他未知键 fail-closed。
            "rank_exclusion_counts": {
                "l1_industry_leader": _rank_stage_excluded_count(
                    rank_reconciliation, "l1_industry_leader"
                ),
                "l2_quality_risk": _rank_stage_excluded_count(
                    rank_reconciliation, "l2_quality_risk"
                ),
                "rank_unexpected": int(
                    (rank_reconciliation or {}).get("unexpected_stage_change_count", 0)
                ),
            },
        },
        "market_context": {
            "trade_calendar": {
                "latest_trade_date": latest_td,
                "next_trade_date": None,
                "calendar_source": "tushare.trade_cal",
                "recent_trade_dates": list(trade_dates),
                "is_pre_holiday_window": False,
                "holiday_days_ahead": None,
            },
            "market_regime": {
                "status": "unknown",
                "confidence": "unknown",
                "position_cap_single_pct": None,
                "position_cap_total_pct": None,
                "min_reward_risk": None,
                "triggers": [],
            },
            "volatility": {
                "iv_symbol": "50ETF",
                "iv_value": None,
                "iv_percentile_252d": None,
                "iv_change_abs_1d_pctpt": None,
                "rule3_status": "unknown",
                "awakening_status": "unknown",
                "cash_reclaim_pct": None,
            },
            "breadth": {
                "limit_up_count": None,
                "limit_down_count": None,
                "limit_up_index_pct_change": None,
                "consecutive_board_height": None,
                "csi300_pct_change_window": None,
            },
            "liquidity": {
                "market_turnover_amount": None,
                "median_amount_20d": None,
            },
            "northbound": {
                "net_flow_5d": None,
                "status": "unknown",
            },
        },
        "account_context": {
            "mode": "new_entry",
            "available_cash": None,
            "total_equity": None,
            "current_gross_exposure": None,
            "positions": [],
        },
        "state_refs": {
            "positions": "state/a_short/positions.json",
            "veto_log": "state/a_short/veto_log.json",
            "circuit_breaker": "state/a_short/circuit_breaker.json",
            "execution_log": "state/a_short/execution_log.csv",
        },
        "candidates": candidates,
    }

    analysis_path = os.path.join(out_dir, "analysis_input.json")
    snapshot_path = os.path.join(out_dir, "snapshot.json")
    candidates_path = os.path.join(out_dir, "candidates.csv")

    validate_analysis_input_contract(analysis_input, label=f"analysis_input export {latest_td}")
    write_json_atomic(analysis_path, analysis_input)

    candidates_df = watch_df.copy()
    candidates_df["run_date"] = latest_td
    candidates_df["run_id"] = run_identity["run_id"]
    candidates_df["candidate_digest"] = run_identity["candidate_digest"]
    write_csv_atomic(candidates_df, candidates_path, index=False, encoding="utf-8-sig")

    snapshot = {
        "schema_name": "snapshot",
        "schema_version": "1.0.0",
        "generated_at": analysis_input["generated_at"],
        "trade_date": latest_td,
        "preset": "a_short",
        "analysis_input": os.path.relpath(analysis_path, project_root),
        "candidates": os.path.relpath(candidates_path, project_root),
        "source_files": analysis_input["source"]["input_files"],
        "run_identity": run_identity,
        "counts": analysis_input["universe_summary"],
        "columns": {
            "watch": list(candidates_df.columns),
            "full": list(df_full.columns),
        },
    }
    write_json_atomic(snapshot_path, snapshot)

    return analysis_path, snapshot_path, candidates_path, analysis_input


def export_stage3_selection_snapshot(top50, tier1_final, latest_td, run_identity,
                                     red_dict, unlock_set, output_root=None):
    """Write the same-run P4a source proof without changing official selection.

    P4a must compare only the already formed Stage3 pool.  This sidecar records
    the pre-rank eligible pool after the two production Stage3 exclusions and
    the actual short ``tier1_final`` output.  It does not rerun a veto, add a
    candidate, alter Top5, or affect the EGS publish decision.
    """
    project_root = os.path.dirname(SCRIPT_DIR)
    base_root = (output_root if output_root and os.path.isabs(output_root) else
                 os.path.join(project_root, output_root) if output_root else
                 os.path.join(project_root, "result", "a_short"))
    out_path = os.path.join(base_root, latest_td, "stage3_selection_snapshot.json")
    required = ("ts_code", "final_score", "l1_name", "l2_name", "tier", "overheat_flag", "chasing_high")

    def _rows(frame, label):
        if frame is None:
            frame = pd.DataFrame()
        missing = [column for column in required if column not in frame.columns]
        if missing:
            raise ValueError(f"P4a {label} snapshot missing columns: {','.join(missing)}")
        rows = []
        for _, row in frame.iterrows():
            score = pd.to_numeric(row.get("final_score"), errors="coerce")
            if pd.isna(score) or not str(row.get("ts_code") or ""):
                raise ValueError(f"P4a {label} snapshot has invalid score/code")
            rows.append({"ts_code": str(row["ts_code"]), "final_score": float(score),
                         "l1_name": str(row.get("l1_name")), "l2_name": str(row.get("l2_name")),
                         "tier": str(row.get("tier")), "overheat_flag": bool(row.get("overheat_flag", False)),
                         "chasing_high": bool(row.get("chasing_high", False))})
        if len({row["ts_code"] for row in rows}) != len(rows):
            raise ValueError(f"P4a {label} snapshot has duplicate ts_code")
        return rows

    top50_rows = _rows(top50, "top50")
    veto_10d = set((red_dict or {}).get("veto_10d", set()))
    eligible_rows = [row for row in top50_rows if row["ts_code"] not in veto_10d and row["ts_code"] not in set(unlock_set or set())]
    import hashlib
    governance_path = os.path.join(project_root, "presets", "egs_industry_heat_governance_20260611.json")
    governance = load_governance()
    active_profile = str(governance.get("active_profile") or "")
    active_weights = (governance.get("profiles") or {}).get(active_profile)
    if not active_profile or not isinstance(active_weights, dict) or not active_weights:
        raise ValueError("P4a Stage3 snapshot requires a valid active industry profile")
    with open(governance_path, "rb") as _governance_handle:
        governance_sha256 = hashlib.sha256(_governance_handle.read()).hexdigest()
    runtime_lineage = runtime_configuration_lineage(_RUNTIME_CONFIGURATION)
    screening_policies = [policy for policy in runtime_lineage.get("policies", [])
                          if policy.get("schema_name") == "a_short_screening_runtime_policy"]
    if len(screening_policies) != 1 or any(not isinstance(screening_policies[0].get(key), str) or not screening_policies[0][key]
                                           for key in ("policy_id", "schema_version", "path", "sha256")):
        raise ValueError("P4a Stage3 snapshot requires one complete screening runtime recipe")
    screening_runtime_recipe = {key: screening_policies[0][key]
                                for key in ("policy_id", "schema_version", "path", "sha256")}
    payload = {"schema_name": "a_short_p4_stage3_selection_snapshot", "schema_version": "1.0.0",
               "as_of": latest_td, "run_id": str((run_identity or {}).get("run_id") or ""),
               "candidate_digest": str((run_identity or {}).get("candidate_digest") or ""),
               "active_industry_weight_profile": {"active_profile": active_profile, "weights": active_weights,
                                                   "governance_sha256": governance_sha256},
               "screening_runtime_recipe": screening_runtime_recipe,
               "top50": top50_rows, "stage3_eligible_pool": eligible_rows,
               "official_tier1_final": _rows(tier1_final, "tier1_final"),
               "boundary": {"comparison_only": True, "changes_final_score_or_tier": False,
                            "changes_official_top5": False, "automatic_promotion": False}}
    if not payload["run_id"] or not payload["candidate_digest"]:
        raise ValueError("P4a Stage3 snapshot requires complete EGS run identity")
    write_json_atomic(out_path, payload)
    return out_path


DATA_HEALTH_SCHEMA_VERSION = "1.5.0"
DATA_HEALTH_SCHEMA_PATH = os.path.join(PROJECT_ROOT, "schemas", "data_health.schema.json")


def _health_issue(check, message, **metrics):
    issue = {"check": check, "message": message}
    issue.update({k: _json_value(v) for k, v in metrics.items()})
    return issue


def _comparison_sidecar_warning(sidecar_name, exc):
    return _health_issue(
        f"comparison_sidecar_{sidecar_name}",
        f"comparison-only {sidecar_name} sidecar failed: {safe_exception_summary(exc)}; "
        "formal selection unchanged",
    )


def _missing_or_nonpositive_count(df, column):
    if df is None or df.empty:
        return None
    if column not in df.columns:
        return int(len(df))
    values = pd.to_numeric(df[column], errors="coerce")
    return int(values.isna().sum() + (values <= 0).sum())


def _missing_count(df, column):
    if df is None or df.empty:
        return None
    if column not in df.columns:
        return int(len(df))
    values = df[column]
    missing = values.isna()
    if values.dtype == object:
        missing = missing | values.astype(str).str.strip().isin(["", "nan", "None", "未知"])
    return int(missing.sum())


def _candidate_quality_scores(analysis_input):
    scores = []
    for candidate in (analysis_input or {}).get("candidates", []):
        data_quality = candidate.get("data_quality", {}) if isinstance(candidate, dict) else {}
        score = _json_float(data_quality.get("completeness_score"))
        if score is not None:
            scores.append(score)
    return scores


def _rank_reconciliation_not_observed():
    return {
        "status": "not_observed",
        "l0_count": 0,
        "ranked_count": 0,
        "expected_excluded_count": 0,
        "unexpected_excluded_count": 0,
        "unexpected_added_count": 0,
        "unexpected_stage_change_count": 0,
        "accounted_count": 0,
        "unaccounted_count": 0,
        "duplicate_code_count": 0,
        "accounting_balanced": True,
        "detail_row_count": 0,
        "source_coverage_failure_count": 0,
        "stage_counts": [],
        "source_coverage": {},
    }


def build_watch_pool_reconciliation(actual_count, eligible_count, target_count):
    """Account for a short watch pool without treating a target cap as a minimum.

    Production exports every eligible Tier1 row up to ``target_count``.  A pool
    below the target is therefore healthy when the eligible pool itself is
    exhausted; only an unexplained export count mismatch is a data-health error.
    """
    actual_count = max(int(actual_count), 0)
    eligible_count = max(int(eligible_count), 0)
    target_count = max(int(target_count), 0)
    expected_count = min(eligible_count, target_count)
    status = "pass" if actual_count == expected_count else "fail"
    if status == "fail":
        reason = "output_count_mismatch"
    elif actual_count >= target_count:
        reason = "target_met"
    else:
        reason = "eligible_pool_exhausted"
    return {
        "status": status,
        "reason": reason,
        "target_count": target_count,
        "eligible_count": eligible_count,
        "expected_count": expected_count,
        "actual_count": actual_count,
        "shortfall_count": max(target_count - actual_count, 0),
    }


def build_data_health(df_full, watch_df, tier1_final, analysis_input, latest_td,
                      analysis_path, snapshot_path, candidates_path,
                      tier1_csv_path, full_csv_path, rank_reconciliation=None,
                      rank_reconciliation_path=None, watch_eligible_count=None,
                      sidecar_warnings=None):
    errors = []
    warnings_ = []
    watch_count = int(len(watch_df)) if watch_df is not None else 0
    final_count = int(min(len(tier1_final), CONF["final_n"])) if tier1_final is not None else 0
    tier1_count = 0
    if watch_df is not None and not watch_df.empty and "tier" in watch_df.columns:
        tier1_count = int((watch_df["tier"].astype(str) == "Tier1").sum())

    quality_scores = _candidate_quality_scores(analysis_input)
    low_quality_count = int(sum(score < 95 for score in quality_scores))
    severe_quality_count = int(sum(score < 75 for score in quality_scores))
    close_bad_count = _missing_or_nonpositive_count(watch_df, "close")
    pe_missing_count = _missing_count(watch_df, "pe_ttm")
    if pe_missing_count is None or pe_missing_count == watch_count:
        pe_missing_count = _missing_count(watch_df, "pe")
    pb_missing_count = _missing_count(watch_df, "pb")
    l1_unknown_count = _missing_count(watch_df, "l1_name")
    l2_unknown_count = _missing_count(watch_df, "l2_name")
    full_l2_unknown_count = _missing_count(df_full, "l2_name")
    rank_reconciliation = dict(
        rank_reconciliation or _rank_reconciliation_not_observed()
    )
    if watch_eligible_count is None:
        watch_eligible_count = watch_count
    watch_pool_reconciliation = build_watch_pool_reconciliation(
        actual_count=watch_count,
        eligible_count=watch_eligible_count,
        target_count=CONF["watch_n"],
    )
    sw_industry_membership = _current_sw_industry_source_observation()

    if not isinstance(analysis_input, dict):
        errors.append(_health_issue("analysis_input", "analysis_input is not a JSON object"))
    else:
        source = analysis_input.get("source", {})
        if analysis_input.get("schema_name") != "analysis_input":
            errors.append(_health_issue("analysis_input_schema", "analysis_input schema_name mismatch"))
        if analysis_input.get("schema_version") != ANALYSIS_INPUT_SCHEMA_VERSION:
            errors.append(_health_issue(
                "analysis_input_version",
                f"analysis_input schema_version is not {ANALYSIS_INPUT_SCHEMA_VERSION}",
                schema_version=analysis_input.get("schema_version"),
            ))
        if source.get("screening_engine_version") != EGS_VERSION:
            errors.append(_health_issue(
                "engine_version",
                "analysis_input was not produced by the current EGS version",
                expected=EGS_VERSION,
                actual=source.get("screening_engine_version"),
            ))
        expected_data_provider = (
            "mixed" if source.get("l3_provider") == HITHINK_L3_SOURCE_ID else "tushare"
        )
        if source.get("data_provider") != expected_data_provider:
            errors.append(_health_issue(
                "data_provider",
                "analysis_input data_provider does not match its L3 provider composition",
                expected=expected_data_provider,
                actual=source.get("data_provider"),
            ))

    checked_paths = {
        "analysis_input": analysis_path,
        "snapshot": snapshot_path,
        "candidates": candidates_path,
        "tier1_csv": tier1_csv_path,
        "full_rank": full_csv_path,
    }
    if rank_reconciliation_path:
        checked_paths["rank_universe_reconciliation"] = rank_reconciliation_path
    missing_outputs = [role for role, path in checked_paths.items() if path and not os.path.exists(path)]
    if missing_outputs:
        errors.append(_health_issue(
            "output_files",
            "one or more expected output files are missing",
            missing=",".join(missing_outputs),
        ))

    if df_full is None or df_full.empty:
        errors.append(_health_issue("full_universe", "full ranked universe is empty"))
    if rank_reconciliation.get("status") == "fail":
        if int(rank_reconciliation.get("source_coverage_failure_count", 0)):
            errors.append(_health_issue(
                "rank_source_coverage",
                "one or more rank inputs are below their required symbol coverage",
                failure_count=rank_reconciliation.get("source_coverage_failure_count"),
            ))
        if (
            int(rank_reconciliation.get("unexpected_stage_change_count", 0))
            or int(rank_reconciliation.get("unaccounted_count", 0))
            or int(rank_reconciliation.get("duplicate_code_count", 0))
            or not bool(rank_reconciliation.get("accounting_balanced", False))
        ):
            errors.append(_health_issue(
                "rank_universe_reconciliation",
                "post-L0 rank universe has an unexplained row change",
                unexpected_stage_change_count=rank_reconciliation.get("unexpected_stage_change_count"),
                unaccounted_count=rank_reconciliation.get("unaccounted_count"),
                duplicate_code_count=rank_reconciliation.get("duplicate_code_count"),
            ))
    elif rank_reconciliation.get("status") == "pass" and df_full is not None:
        if int(rank_reconciliation.get("ranked_count", -1)) != int(len(df_full)):
            errors.append(_health_issue(
                "rank_universe_reconciliation",
                "reconciled ranked count does not match full-rank output",
                reconciled_ranked_count=rank_reconciliation.get("ranked_count"),
                full_count=len(df_full),
            ))

    if watch_count == 0:
        errors.append(_health_issue("watch_pool", "watch pool is empty"))
    if watch_pool_reconciliation["status"] == "fail":
        errors.append(_health_issue(
            "watch_pool_reconciliation",
            "exported watch pool count does not match the eligible capped pool",
            actual_count=watch_pool_reconciliation["actual_count"],
            expected_count=watch_pool_reconciliation["expected_count"],
            eligible_count=watch_pool_reconciliation["eligible_count"],
            target_count=watch_pool_reconciliation["target_count"],
        ))

    if final_count == 0:
        errors.append(_health_issue("final_pool", "final recommendation pool is empty"))
    elif final_count < int(CONF["final_n"]):
        warnings_.append(_health_issue(
            "final_pool",
            "final recommendation count is below configured final_n",
            final_count=final_count,
            final_n=CONF["final_n"],
        ))

    if tier1_count == 0:
        errors.append(_health_issue("tier1_count", "no Tier1 candidates in watch pool"))
    elif tier1_count < int(CONF["final_n"]):
        warnings_.append(_health_issue(
            "tier1_count",
            "Tier1 count is below final_n",
            tier1_count=tier1_count,
            final_n=CONF["final_n"],
        ))

    if close_bad_count:
        errors.append(_health_issue(
            "close",
            "watch pool contains missing or non-positive close values",
            bad_count=close_bad_count,
        ))
    for column_name, missing_count in (("pe_ttm_or_pe", pe_missing_count), ("pb", pb_missing_count)):
        if missing_count and watch_count and missing_count / watch_count > 0.2:
            warnings_.append(_health_issue(
                column_name,
                "watch pool valuation missing rate is above 20%",
                missing_count=missing_count,
                watch_count=watch_count,
            ))
    for column_name, unknown_count in (("l1_name", l1_unknown_count), ("l2_name", l2_unknown_count)):
        if unknown_count:
            warnings_.append(_health_issue(
                column_name,
                "watch pool contains unknown industry labels",
                unknown_count=unknown_count,
            ))
    if severe_quality_count:
        errors.append(_health_issue(
            "completeness_score",
            "candidate data completeness score below 75",
            severe_count=severe_quality_count,
        ))
    elif low_quality_count:
        warnings_.append(_health_issue(
            "completeness_score",
            "candidate data completeness score below 95",
            low_count=low_quality_count,
        ))

    warnings_.extend(sidecar_warnings or [])
    status = "error" if errors else ("warn" if warnings_ else "ok")
    return {
        "schema_name": "data_health",
        "schema_version": DATA_HEALTH_SCHEMA_VERSION,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "trade_date": latest_td,
        "preset": "a_short",
        "market": "A",
        "source": {
            "screening_engine": "egs_main.py",
            "screening_engine_version": EGS_VERSION,
            "data_provider": (
                "mixed" if CONF.get("l3_provider") == HITHINK_L3_SOURCE_ID else "tushare"
            ),
            "api_families": list(EGS_API_FAMILIES),
            "l3_mode": CONF.get("l3_mode", "today"),
            "l3_pit_strict": bool(CONF.get("l3_pit_strict", False)),
            "l3_provider": CONF.get("l3_provider"),
        },
        "overall_status": status,
        "errors": errors,
        "warnings": warnings_,
        "metrics": {
            "full_count": int(len(df_full)) if df_full is not None else 0,
            "watch_count": watch_count,
            "final_count": final_count,
            "tier1_count": tier1_count,
            "close_missing_or_nonpositive_count": close_bad_count,
            "pe_ttm_or_pe_missing_count": pe_missing_count,
            "pb_missing_count": pb_missing_count,
            "watch_l1_unknown_count": l1_unknown_count,
            "watch_l2_unknown_count": l2_unknown_count,
            "full_l2_unknown_count": full_l2_unknown_count,
            "watch_pool_reconciliation": watch_pool_reconciliation,
            "sw_industry_membership": sw_industry_membership,
            "rank_universe_reconciliation": rank_reconciliation,
            "suspend_daily_coverage": _current_suspend_daily_coverage_observation(),
            "completeness_score_min": min(quality_scores) if quality_scores else None,
            "completeness_score_below_95_count": low_quality_count,
            "completeness_score_below_75_count": severe_quality_count,
        },
        "outputs_checked": checked_paths,
        "limitations": [
            "This is an internal Tushare output health check, not a second-source reconciliation.",
            "It checks structural integrity, source coverage, post-L0 rank reconciliation, key field coverage, counts, industry labels, and data completeness.",
            "AKShare canary (runners/data_canary.py) now uses sina (stock_zh_a_spot) by default and works reliably; "
            "the em source needs VPN split-tunnel for *.eastmoney.com if pe/pb reconciliation is required.",
        ],
    }


def validate_data_health_consistency(health):
    metrics = health.get("metrics", {}) if isinstance(health, dict) else {}
    watch = metrics.get("watch_pool_reconciliation")
    if not isinstance(watch, dict):
        raise ValueError("data_health missing watch_pool_reconciliation")
    expected_watch = build_watch_pool_reconciliation(
        actual_count=watch.get("actual_count"),
        eligible_count=watch.get("eligible_count"),
        target_count=watch.get("target_count"),
    )
    if watch != expected_watch:
        raise ValueError("data_health watch_pool_reconciliation is internally inconsistent")

    sw = metrics.get("sw_industry_membership")
    if not isinstance(sw, dict):
        raise ValueError("data_health missing sw_industry_membership")
    status = sw.get("status")
    source = sw.get("source")
    fast_path_used = bool(sw.get("fast_path_used"))
    fallback_used = bool(sw.get("fallback_used"))
    cache_hit = bool(sw.get("cache_hit"))
    if sum((fast_path_used, fallback_used, cache_hit)) > 1:
        raise ValueError("data_health SW source flags are mutually exclusive")
    if status == "not_observed":
        if source is not None or sw.get("active_count") is not None or any(
            (fast_path_used, fallback_used, cache_hit)
        ):
            raise ValueError("data_health unobserved SW source carries observed values")
    elif status == "pass":
        active_count = int(sw.get("active_count"))
        min_active = int(sw.get("min_active"))
        if active_count < min_active:
            raise ValueError("data_health passing SW source is below minimum coverage")
        expected_flags = {
            "cache": (False, False, True),
            "index_member_all_l1_current": (True, False, False),
            "index_member_l2_history": (False, True, False),
        }
        if source not in expected_flags or (
            fast_path_used,
            fallback_used,
            cache_hit,
        ) != expected_flags[source]:
            raise ValueError("data_health passing SW source does not match its source flags")
    return health


def export_data_health(df_full, watch_df, tier1_final, analysis_input, latest_td,
                       analysis_path, snapshot_path, candidates_path,
                       tier1_csv_path, full_csv_path, rank_reconciliation=None,
                       rank_reconciliation_path=None, watch_eligible_count=None,
                       sidecar_warnings=None):
    health = build_data_health(
        df_full=df_full,
        watch_df=watch_df,
        tier1_final=tier1_final,
        analysis_input=analysis_input,
        latest_td=latest_td,
        analysis_path=analysis_path,
        snapshot_path=snapshot_path,
        candidates_path=candidates_path,
        tier1_csv_path=tier1_csv_path,
        full_csv_path=full_csv_path,
        rank_reconciliation=rank_reconciliation,
        rank_reconciliation_path=rank_reconciliation_path,
        watch_eligible_count=watch_eligible_count,
        sidecar_warnings=sidecar_warnings,
    )
    validate_json_schema(health, schema_path=DATA_HEALTH_SCHEMA_PATH, label=f"data_health export {latest_td}")
    validate_data_health_consistency(health)
    health_path = os.path.join(os.path.dirname(analysis_path), "data_health.json")
    write_json_atomic(health_path, health)
    return health_path, health


def publish_egs_run_manifest(analysis_input, health, paths):
    """Publish the sole official marker after every EGS artifact validates.

    Files written by a failed/staged run are not official without this marker;
    because the marker itself is atomically replaced last, the prior official
    run identity remains authoritative on any upstream failure.
    """
    if health.get("overall_status") == "error":
        raise RuntimeError("data_health is error; refusing official EGS publish")
    import hashlib
    run_identity = dict(((analysis_input.get("source") or {}).get("run_identity") or {}))
    if not run_identity:
        raise RuntimeError("analysis_input missing run_identity; refusing official EGS publish")
    files = {}
    for role, path in paths.items():
        with open(path, "rb") as handle:
            files[role] = {"path": os.path.basename(path), "sha256": hashlib.sha256(handle.read()).hexdigest()}
    manifest = {
        "schema_name": "a_short_egs_official_publish",
        "schema_version": "1.0.0",
        "published_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "trade_date": analysis_input["trade_date"],
        "run_id": run_identity["run_id"],
        "candidate_digest": run_identity["candidate_digest"],
        "stage_status": "complete",
        "files": files,
    }
    marker = os.path.join(os.path.dirname(paths["analysis_input"]), "official_publish.json")
    write_json_atomic(marker, manifest)
    return marker, manifest


def log_data_health_summary(health_path, health):
    status = str(health.get("overall_status", "error")).upper()
    errors_count = len(health.get("errors", []))
    warnings_count = len(health.get("warnings", []))
    metrics = health.get("metrics", {})
    message = (
        f"[DATA_HEALTH] {status}: errors={errors_count}, warnings={warnings_count}, "
        f"watch={metrics.get('watch_count')}, tier1={metrics.get('tier1_count')}, "
        f"final={metrics.get('final_count')} -> {health_path}"
    )
    print(message)
    log.info(message)


def safe_api(fn, *a, default=None, retries=3, **kw):
    for i in range(retries):
        try:
            time.sleep(CONF["request_delay"])
            res = fn(*a, **kw)
            if res is not None and len(res) > 0: return res
            return default
        except Exception as e:
            log.warning(f"API异常 [{fn.__name__ if hasattr(fn,'__name__') else fn}] 第{i+1}次: {e}")
            if i == retries - 1: return default
            time.sleep(2 ** i)
    return default

L3_SNAPSHOT_SCHEMA = 1  # bump if dict shape changes
L3_SNAPSHOT_PREFIX = "l3_snapshot_"
L3_SNAPSHOT_SUFFIX = ".pkl"


def _l3_snapshot_path(date_str):
    return os.path.join(L3_SNAPSHOT_DIR, f"{L3_SNAPSHOT_PREFIX}{date_str}{L3_SNAPSHOT_SUFFIX}")


def _write_l3_snapshot(date_str, concepts_df, stock_concepts, concept_members,
                       l3_source=None, coverage=None):
    """Atomic single-file snapshot. Same-day re-runs overwrite cleanly; crash
    mid-write cannot produce inconsistent multi-file state (H1).

    date_str should be the real-world date the data was observed (datetime.now,
    not the as_of override) so PIT lookups have correct semantics.
    """
    os.makedirs(L3_SNAPSHOT_DIR, exist_ok=True)
    path = _l3_snapshot_path(date_str)
    payload = {
        "schema": L3_SNAPSHOT_SCHEMA,
        "snap_date": date_str,
        "concepts_df": concepts_df,
        "stock_concepts": stock_concepts,
        "concept_members": concept_members,
    }
    if l3_source is not None:
        payload["l3_source"] = str(l3_source)
    if coverage is not None:
        payload["coverage"] = dict(coverage)
    tmp = path + ".tmp"
    with open(tmp, "wb") as f:
        pickle.dump(payload, f)
    os.replace(tmp, path)


def _list_l3_snapshot_dates():
    if not os.path.isdir(L3_SNAPSHOT_DIR):
        return []
    dates = []
    for name in os.listdir(L3_SNAPSHOT_DIR):
        if not (name.startswith(L3_SNAPSHOT_PREFIX) and name.endswith(L3_SNAPSHOT_SUFFIX)):
            continue
        d = name[len(L3_SNAPSHOT_PREFIX):-len(L3_SNAPSHOT_SUFFIX)]
        if len(d) != 8 or not d.isdigit():
            continue
        dates.append(d)
    return sorted(dates)


def _main_board_l3_membership(concept_members):
    scoped_members = {}
    stock_concepts = {}
    for concept_id, member_codes in (concept_members or {}).items():
        members = sorted({
            str(code) for code in (member_codes or []) if is_a_share_main_board(code)
        })
        scoped_members[str(concept_id)] = members
        for code in members:
            stock_concepts.setdefault(code, []).append(str(concept_id))
    for code in stock_concepts:
        stock_concepts[code] = sorted(set(stock_concepts[code]))
    return stock_concepts, scoped_members


def _load_l3_snapshot(as_of, include_metadata=False):
    """Return (concepts_df, stock_concepts, concept_members, snap_date) for the
    latest complete snapshot with snap_date <= as_of, or None if none exist or
    the file is corrupt (caller handles fallback)."""
    dates = _list_l3_snapshot_dates()
    eligible = [d for d in dates if d <= as_of]
    if not eligible:
        return None
    snap_date = eligible[-1]
    path = _l3_snapshot_path(snap_date)
    try:
        with open(path, "rb") as f:
            payload = pickle.load(f)
    except (pickle.UnpicklingError, EOFError, OSError) as e:
        log.warning(f"L3 snapshot {snap_date} unreadable ({type(e).__name__}: {e}); treating as missing")
        return None
    if not isinstance(payload, dict) or payload.get("schema") != L3_SNAPSHOT_SCHEMA:
        log.warning(f"L3 snapshot {snap_date} schema mismatch (got {payload.get('schema') if isinstance(payload, dict) else type(payload).__name__}); treating as missing")
        return None
    required = {"concepts_df", "stock_concepts", "concept_members"}
    missing = required - set(payload.keys())
    if missing:
        log.warning(f"L3 snapshot {snap_date} missing keys {missing}; treating as missing")
        return None
    stock_concepts, concept_members = _main_board_l3_membership(payload["concept_members"])
    result = (payload["concepts_df"], stock_concepts, concept_members, snap_date)
    if include_metadata:
        return result + (payload.get("l3_source"), payload.get("coverage"))
    return result


def _is_complete_hithink_snapshot(concepts_df, concept_members, coverage):
    if not isinstance(coverage, dict):
        return False
    catalog_codes = _catalog_codes_from_snapshot(concepts_df)
    main_board_member_pair_count = sum(
        len(set(member_codes or [])) for member_codes in (concept_members or {}).values()
    )
    return (
        coverage.get("source") == HITHINK_L3_SOURCE_ID
        and coverage.get("scoring_universe") == "a_share_main_board"
        and coverage.get("complete") is True
        and len(catalog_codes) >= MIN_CONCEPT_CATALOG_BOARD_COUNT
        and coverage.get("catalog_digest") == catalog_digest(catalog_codes)
        and coverage.get("catalog_board_count") == len(catalog_codes)
        and coverage.get("catalog_board_count") == coverage.get("received_board_count")
        and set(concept_members or {}) == catalog_codes
        and coverage.get("main_board_member_pair_count") == main_board_member_pair_count
    )


def _catalog_codes_from_snapshot(concepts_df):
    if not isinstance(concepts_df, pd.DataFrame) or "code" not in concepts_df.columns:
        return set()
    return set(concepts_df["code"].dropna().astype(str))


def set_asof(date_str):
    global TODAY, TODAY_DT
    try:
        asof_dt = datetime.strptime(date_str, "%Y%m%d")
    except ValueError as exc:
        raise ValueError("--as-of must use YYYYMMDD format") from exc

    cal = safe_api(
        pro.trade_cal,
        exchange="SSE",
        start_date=date_str,
        end_date=date_str,
        is_open="1",
        fields="cal_date,is_open",
    )
    if cal is None or cal.empty:
        raise ValueError(f"--as-of {date_str} is not an A-share trading day")

    TODAY = date_str
    TODAY_DT = asof_dt
    if CONF.get("cache_policy") == "disabled":
        log.info(f"[ASOF] running EGS as of {TODAY}; cache_policy=disabled; cache_ttl=ignored")
        return
    CONF["cache_ttl"] = BACKTEST_CACHE_TTL
    log.info(f"[ASOF] running EGS as of {TODAY}; cache_ttl={CONF['cache_ttl']}s")

def _guard_historical_asof_l3_mode(as_of, l3_mode, allow_historical_live_l3=False, run_date=None):
    """l3=today 只对**真·过去**的 as_of 拦截（须显式 pit/neutralize）；当前/前瞻 as_of 放行。

    判据从「as_of != run_date」放宽为「as_of < run_date」：weekly canonical 解析器（2026-06-22）会把
    周末/周一盘前运行解析成「即将到来的周一」as_of（> run_date 的前瞻交易日，EOD 尚未发布 → 价格回退到
    上一已结算交易日，与周一盘前实盘同一条已验证路径）。这类前瞻 live 运行用 l3=today 正确（L3 概念数据
    同价格一样取当前最佳），不该被当历史回放拦死。真·过去回放（as_of < run_date）仍须 pit/neutralize。
    """
    if not as_of:
        return
    effective_run_date = run_date or a_share_market_date()
    if str(as_of) < str(effective_run_date) and l3_mode == "today" and not allow_historical_live_l3:
        raise SystemExit(
            f"[FATAL] Historical --as-of {as_of} predates the run date {effective_run_date} "
            "and cannot run with --l3-mode=today by default. Use --l3-mode=pit --l3-pit-strict "
            "for PIT snapshots, --l3-mode=neutralize for an L3-neutral replay, or add "
            "--allow-historical-live-l3 only for an explicitly non-evidence live-concept smoke run."
        )

def dstr(days_ago=0):
    return (TODAY_DT - timedelta(days=days_ago)).strftime("%Y%m%d")

def dfuture(days_ahead=30):
    return (TODAY_DT + timedelta(days=days_ahead)).strftime("%Y%m%d")

def to_chunks(lst, size):
    return [lst[i:i+size] for i in range(0, len(lst), size)]


# ═══════════════════════════════════════════════════
# §2 数据拉取
# ═══════════════════════════════════════════════════
def get_trade_dates(n=25):
    key = f"trade_dates_{TODAY}_official_v2"
    cached = load_cache(key)
    if cached is not None and len(cached) >= n:
        return cached[:n]
    # 缓存不存在或数量不足 n，重新拉取；dstr(100) 覆盖约70个交易日
    df = safe_api(pro.trade_cal, exchange="SSE", start_date=dstr(100), end_date=TODAY,
                  is_open="1", fields="cal_date")
    if df is None or df.empty:
        raise RuntimeError("official trade calendar source unavailable; refusing weekday fallback")
    if "cal_date" not in df.columns:
        raise RuntimeError("official trade calendar payload has no cal_date column")
    raw_dates = df["cal_date"].dropna().astype(str).tolist()
    dates = []
    for value in raw_dates:
        try:
            parsed = datetime.strptime(value, "%Y%m%d")
        except ValueError as exc:
            raise RuntimeError(f"official trade calendar contains invalid cal_date {value!r}") from exc
        if parsed > TODAY_DT:
            raise RuntimeError(f"official trade calendar contains future cal_date {value!r}")
        dates.append(value)
    dates = sorted(set(dates), reverse=True)[:max(n, 70)]
    if len(dates) < n:
        raise RuntimeError(f"official trade calendar returned only {len(dates)} dates; need {n}")
    save_cache(key, dates)
    return dates[:n]

def _namechange_date(value, field_name):
    """Return one optional canonical Tushare date or fail before historical replay."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    if not re.fullmatch(r"\d{8}", text):
        raise RuntimeError(f"namechange has invalid {field_name}: {value!r}")
    return text


def _historical_name_map(df_stocks):
    """Resolve each universe member's name active and knowable at TODAY."""
    namechange = getattr(pro, "namechange", None)
    if not callable(namechange):
        raise RuntimeError("historical namechange source unavailable; refusing current-name fallback")
    fields = "ts_code,name,start_date,end_date,change_reason"
    history = safe_api(namechange, fields=fields)
    if history is None or history.empty:
        raise RuntimeError("historical namechange source returned no rows; refusing current-name fallback")
    required = {"ts_code", "name", "start_date", "end_date"}
    missing_columns = required - set(history.columns)
    if missing_columns:
        raise RuntimeError(
            "historical namechange payload missing columns: " + ",".join(sorted(missing_columns))
        )

    history = history.copy()
    history["ts_code"] = history["ts_code"].fillna("").astype(str).str.strip()
    history["name"] = history["name"].fillna("").astype(str).str.strip()
    for column in ("start_date", "end_date"):
        history[column] = history[column].map(lambda value, c=column: _namechange_date(value, c))
    if (history["ts_code"] == "").any() or (history["name"] == "").any():
        raise RuntimeError("historical namechange payload has blank ts_code or name")
    if (history["start_date"] == "").any():
        raise RuntimeError("historical namechange payload has blank start_date")

    active = history[
        (history["start_date"] <= TODAY)
        & ((history["end_date"] == "") | (history["end_date"] >= TODAY))
    ].copy()
    duplicate_codes = active.loc[active["ts_code"].duplicated(keep=False), "ts_code"].unique().tolist()
    if duplicate_codes:
        raise RuntimeError(
            "historical namechange has overlapping active names: " + ",".join(sorted(duplicate_codes)[:10])
        )

    names = active.set_index("ts_code")["name"].to_dict()
    universe_codes = set(df_stocks["ts_code"].dropna().astype(str))
    uncovered = sorted(universe_codes - set(names))
    if uncovered:
        raise RuntimeError(
            "historical namechange PIT coverage incomplete; refusing current-name fallback: "
            + ",".join(uncovered[:10])
        )
    return names


def get_stock_list():
    """As-of-aware universe.

    Pulls L (listed), D (delisted), and P (paused) so that historical as_of dates
    can reach stocks that have since delisted -- otherwise the backtest carries
    a survivorship bias. Keeps a row iff list_date <= TODAY AND
    (delist_date is empty OR delist_date > TODAY).
    """
    # Historical rows have PIT-replaced names; current rows deliberately keep
    # stock_basic names.  They must never share a cache entry.
    mode = "hist" if _historical_replay_mode() else "cur"
    key = f"stock_list_{TODAY}_v4_{mode}"
    cached = load_cache(key)
    if cached is not None:
        return cached
    log.info("拉取股票基础信息(L+D+P,按as_of过滤)...")
    fields = "ts_code,symbol,name,list_date,delist_date,market,list_status"
    frames = []
    for status in ("L", "D", "P"):
        part = safe_api(pro.stock_basic, exchange="", list_status=status, fields=fields)
        if part is not None and not part.empty:
            frames.append(part)
    if not frames:
        raise RuntimeError("股票总表获取失败")
    df = pd.concat(frames, ignore_index=True)
    required_columns = {"ts_code", "name", "list_date", "delist_date", "list_status"}
    missing_columns = sorted(required_columns - set(df.columns))
    if missing_columns:
        raise RuntimeError(
            "stock_basic payload missing delisting-safety columns: " + ",".join(missing_columns)
        )
    # Tushare can expose the same code in multiple status queries.  Prefer the
    # most restrictive status before applying the as-of date filter, otherwise
    # the initial L row can hide a D/P row and let a delisting escape L0.
    status_priority = {"D": 0, "P": 1, "L": 2}
    df["_status_priority"] = (
        df["list_status"].fillna("").astype(str).str.upper().map(status_priority).fillna(-1)
    )
    df = (
        df.sort_values(["ts_code", "_status_priority"], kind="stable")
        .drop_duplicates(subset=["ts_code"], keep="first")
        .drop(columns=["_status_priority"])
        .reset_index(drop=True)
    )
    if "list_date" in df.columns:
        df = df[df["list_date"].fillna("") <= TODAY].copy()
    if "delist_date" in df.columns:
        delist = df["delist_date"].fillna("")
        df = df[(delist == "") | (delist > TODAY)].copy()
    if mode == "hist":
        historical_names = _historical_name_map(df)
        df["name"] = df["ts_code"].map(historical_names)
    save_cache(key, df)
    return df

SW_INDUSTRY_MIN_ACTIVE = 3000
SW_INDEX_MEMBER_ALL_ROW_LIMIT = 2000
_LAST_SW_INDUSTRY_SOURCE_OBSERVATION = None


def _sw_industry_source_not_observed():
    return {
        "status": "not_observed",
        "source": None,
        "as_of": None,
        "active_count": None,
        "min_active": int(SW_INDUSTRY_MIN_ACTIVE),
        "request_group_count": 0,
        "fast_path_used": False,
        "fallback_used": False,
        "cache_hit": False,
        "message": None,
    }


def _record_sw_industry_source_observation(**details):
    global _LAST_SW_INDUSTRY_SOURCE_OBSERVATION
    payload = _sw_industry_source_not_observed()
    payload.update(details)
    _LAST_SW_INDUSTRY_SOURCE_OBSERVATION = payload
    return dict(payload)


def _current_sw_industry_source_observation():
    if _LAST_SW_INDUSTRY_SOURCE_OBSERVATION is None:
        return _sw_industry_source_not_observed()
    return dict(_LAST_SW_INDUSTRY_SOURCE_OBSERVATION)


def _sw_mapping_is_usable(mapping):
    return isinstance(mapping, dict) and len(mapping) >= SW_INDUSTRY_MIN_ACTIVE


def _normalize_member_dates(df):
    """Ensure in_date / out_date columns exist and have no NaN. Tushare's
    index_member_all vs index_member may not return identical schemas; guard
    against missing columns so downstream PIT filter doesn't KeyError."""
    if df is None or df.empty:
        return df
    for col in ("in_date", "out_date"):
        if col not in df.columns:
            df[col] = ""
        else:
            df[col] = df[col].fillna("")
    return df


def _normalize_sw_member_columns(df, source):
    """Normalize the two documented Tushare SW member response shapes."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["con_code", "index_code", "in_date", "out_date"])
    normalized = df.copy()
    if source == "index_member_all":
        aliases = {}
        if "con_code" not in normalized.columns and "ts_code" in normalized.columns:
            aliases["ts_code"] = "con_code"
        if "index_code" not in normalized.columns and "l2_code" in normalized.columns:
            aliases["l2_code"] = "index_code"
        normalized = normalized.rename(columns=aliases)
    missing = [name for name in ("con_code", "index_code") if name not in normalized.columns]
    if missing:
        log.warning(
            f"SW industry member source={source} missing required columns {missing}; "
            "rejecting this source shape"
        )
        return pd.DataFrame(columns=["con_code", "index_code", "in_date", "out_date"])
    return _normalize_member_dates(normalized)


def _apply_pit_window(df, l2_codes, as_of_date, source="index_member"):
    """Filter to L2 codes of interest and keep only rows active at as_of.

    Defends against Tushare returning a DataFrame without the expected schema
    (`con_code` / `index_code`); falls back to empty DataFrame so caller's
    threshold checks fire instead of KeyError aborting the whole run.
    """
    if df is None or df.empty:
        return df
    df = _normalize_sw_member_columns(df, source)
    if df.empty:
        return df
    df = df[df["index_code"].isin(l2_codes)].copy()
    df = _normalize_member_dates(df)
    mask_in = (df["in_date"] == "") | (df["in_date"] <= as_of_date)
    mask_out = (df["out_date"] == "") | (df["out_date"] > as_of_date)
    return df[mask_in & mask_out]


def get_sw_industry_map():
    """As-of-aware SW industry membership.

    Tushare's pro.index_member exposes in_date / out_date per (con_code, index_code).
    We fetch the full membership history and pick the row whose interval contains
    TODAY, so each as_of date sees the industry assignment that was active at the
    time (instead of today's snapshot with is_new=1).
    """
    # v4 caches written before SW_INDUSTRY_MIN_ACTIVE safeguard had ~75-stock
    # coverage (mostly l2_name="未知"). v5 fixed L2 coverage but missed L1
    # parent mapping because Tushare L2 parent_code points to L1 industry_code,
    # not index_code. v6 invalidates both broken cache generations.
    key = f"sw_industry_map_v6_{TODAY}"
    cached = load_cache(key)
    if cached is not None:
        if _sw_mapping_is_usable(cached):
            _record_sw_industry_source_observation(
                status="pass",
                source="cache",
                as_of=TODAY,
                active_count=int(len(cached)),
                min_active=int(SW_INDUSTRY_MIN_ACTIVE),
                request_group_count=0,
                fast_path_used=False,
                fallback_used=False,
                cache_hit=True,
                message=None,
            )
            return cached
        cached_size = len(cached) if isinstance(cached, dict) else "invalid"
        log.warning(f"SW industry cache {key} coverage too low ({cached_size}); refetching")
    log.info("拉取申万行业分类(按as_of时点)...")
    df_l2 = safe_api(pro.index_classify, level="L2")
    df_l1 = safe_api(pro.index_classify, level="L1")
    if df_l2 is None or df_l2.empty:
        raise RuntimeError("SW L2 industry classification unavailable")

    l1_map = {}
    if df_l1 is not None and not df_l1.empty:
        l1_map.update(dict(zip(df_l1["index_code"], df_l1["industry_name"])))
        if "industry_code" in df_l1.columns:
            l1_map.update(dict(zip(df_l1["industry_code"].astype(str), df_l1["industry_name"])))
    l2_info = {row["index_code"]: row for _, row in df_l2.iterrows()}

    member_fields = "con_code,index_code,in_date,out_date"
    l2_codes = set(l2_info.keys())

    def _fetch_l2_batch():
        """L2-by-L2 batching fallback. Slow but covers when full-market endpoints
        return incomplete data. Returns a concat'd raw DataFrame (unfiltered)."""
        frames = []
        l2_code_list = df_l2["index_code"].dropna().astype(str).drop_duplicates().tolist()
        for l2_code in tqdm(l2_code_list, desc="SW industry L2 batching"):
            t = safe_api(pro.index_member, index_code=l2_code, fields=member_fields)
            if t is not None and not t.empty:
                frames.append(t)
        if not frames:
            return pd.DataFrame(), len(l2_code_list)
        return pd.concat(frames, ignore_index=True), len(l2_code_list)

    def _fetch_current_by_l1(index_member_all):
        """Use the documented current-members endpoint in bounded L1 groups.

        The unfiltered endpoint is capped and therefore cannot prove full-market
        completeness.  Each L1 response must be non-empty, below the documented
        row limit, and have the official ts_code/l2_code shape; otherwise the
        caller discards the whole fast-path result and uses the PIT history path.
        """
        if df_l1 is None or df_l1.empty or "index_code" not in df_l1.columns:
            return pd.DataFrame(), 0, "l1_classification_unavailable"
        l1_codes = df_l1["index_code"].dropna().astype(str).drop_duplicates().tolist()
        frames = []
        fields = "ts_code,l2_code,in_date,out_date,is_new"
        for l1_code in l1_codes:
            raw = safe_api(
                index_member_all,
                l1_code=l1_code,
                is_new="Y",
                fields=fields,
                retries=1,
            )
            if raw is None or raw.empty:
                return pd.DataFrame(), len(frames) + 1, f"empty_l1_response:{l1_code}"
            if len(raw) >= SW_INDEX_MEMBER_ALL_ROW_LIMIT:
                return pd.DataFrame(), len(frames) + 1, f"row_limit_hit:{l1_code}:{len(raw)}"
            normalized = _normalize_sw_member_columns(raw, "index_member_all")
            if normalized.empty:
                return pd.DataFrame(), len(frames) + 1, f"bad_shape:{l1_code}"
            frames.append(normalized)
        if not frames:
            return pd.DataFrame(), 0, "no_l1_groups"
        return pd.concat(frames, ignore_index=True), len(l1_codes), None

    df_mem = pd.DataFrame()
    member_source = None
    request_group_count = 0
    fallback_reason = None

    # Current production run: query the official endpoint by L1 group.  Never
    # trust the capped unfiltered response, and never use current-only rows for
    # a historical PIT replay.
    index_member_all = getattr(pro, "index_member_all", None)
    wall_date = a_share_market_date()
    if TODAY == wall_date and callable(index_member_all):
        candidate, request_group_count, fallback_reason = _fetch_current_by_l1(index_member_all)
        if not candidate.empty:
            candidate = _apply_pit_window(
                candidate,
                l2_codes,
                TODAY,
                source="index_member_all",
            )
            fast_count = candidate["con_code"].nunique()
            if fast_count >= SW_INDUSTRY_MIN_ACTIVE:
                df_mem = candidate
                member_source = "index_member_all_l1_current"
            else:
                fallback_reason = f"coverage_below_min:{fast_count}<{SW_INDUSTRY_MIN_ACTIVE}"
    elif TODAY != wall_date:
        fallback_reason = "historical_as_of_requires_pit_history"
    else:
        fallback_reason = "index_member_all_unavailable"

    if df_mem.empty:
        log.warning(
            "SW industry L1 current fast path unavailable or incomplete "
            f"({fallback_reason}); fetching PIT history by L2 index_code"
        )
        batched, l2_group_count = _fetch_l2_batch()
        request_group_count += l2_group_count
        if batched.empty:
            _record_sw_industry_source_observation(
                status="fail",
                source="index_member_l2_history",
                as_of=TODAY,
                request_group_count=request_group_count,
                fallback_used=True,
                message=fallback_reason,
            )
            raise RuntimeError("SW industry member fetch failed; cannot build reliable L2 map")
        df_mem = _apply_pit_window(batched, l2_codes, TODAY, source="index_member")
        member_source = "index_member_l2_history"

    active_count = df_mem["con_code"].nunique()
    if active_count < SW_INDUSTRY_MIN_ACTIVE:
        _record_sw_industry_source_observation(
            status="fail",
            source=member_source,
            as_of=TODAY,
            active_count=int(active_count),
            request_group_count=request_group_count,
            fast_path_used=member_source == "index_member_all_l1_current",
            fallback_used=member_source == "index_member_l2_history",
            message=fallback_reason,
        )
        raise RuntimeError(
            f"SW industry map coverage too low: {active_count} active stocks "
            f"(source={member_source}); aborting to avoid invalid Tier1 output"
        )

    # If a stock has multiple active L2 rows, prefer the most recent in_date.
    df_mem = df_mem.sort_values(["con_code", "in_date"], ascending=[True, False]) \
                   .drop_duplicates(subset=["con_code"], keep="first")

    mapping = {}
    for _, row in df_mem.iterrows():
        l2_code = row["index_code"]
        stock   = row["con_code"]
        l2_row  = l2_info[l2_code]
        raw_parent = str(l2_row.get("parent_code", ""))

        # parent_code 与 l1_map key 格式可能不一致，尝试多种匹配
        l1_name = l1_map.get(raw_parent)
        if l1_name is None:
            stripped = raw_parent.split(".")[0]
            l1_name = l1_map.get(stripped)
        if l1_name is None:
            for suffix in [".SI", ".SZ", ".SH"]:
                l1_name = l1_map.get(raw_parent + suffix)
                if l1_name: break
        if l1_name is None:
            l1_name = "未知"

        mapping[stock] = {
            "l2_name": l2_row["industry_name"], "l2_code": l2_code,
            "l1_name": l1_name, "l1_code": raw_parent,
        }
    if not _sw_mapping_is_usable(mapping):
        raise RuntimeError(
            f"SW industry map coverage too low after mapping: {len(mapping)} stocks; aborting to avoid invalid Tier1 output"
        )
    _record_sw_industry_source_observation(
        status="pass",
        source=member_source,
        as_of=TODAY,
        active_count=int(len(mapping)),
        min_active=int(SW_INDUSTRY_MIN_ACTIVE),
        request_group_count=int(request_group_count),
        fast_path_used=member_source == "index_member_all_l1_current",
        fallback_used=member_source == "index_member_l2_history",
        cache_hit=False,
        message=fallback_reason,
    )
    save_cache(key, mapping)
    return mapping

def get_csi300_return(trade_dates):
    key = f"csi300_{trade_dates[0]}"
    if (cached := load_cache(key)) is not None: return cached
    start = trade_dates[-1] if len(trade_dates) >= 20 else dstr(35)
    df = safe_api(pro.index_daily, ts_code="000300.SH", start_date=start, end_date=trade_dates[0],
                  fields="trade_date,close")
    if df is None or len(df) < 2:
        log.warning("沪深300获取失败，L1基准将使用市场中位数")
        return None
    df  = df.sort_values("trade_date", ascending=False).reset_index(drop=True)
    ret = (float(df.iloc[0]["close"]) / float(df.iloc[-1]["close"]) - 1) * 100
    save_cache(key, ret)
    return ret

def _daily_cache_days(df):
    if df is None or df.empty or "trade_date" not in df.columns: return 0
    return int(df["trade_date"].nunique())


# EGS used to cache raw ``pro.daily`` bars and then label the resulting
# cross-day indicators as qfq.  Keep the raw columns for exchange-mechanism
# checks, but make every cross-day price statistic consume this explicit,
# as-of-anchored qfq view.
DAILY_ALL_QFQ_CACHE_VERSION = "qfq_v1"
DAILY_ALL_RAW_OHLC_COLUMNS = ("open", "high", "low", "close")
DAILY_ALL_QFQ_OHLC_COLUMNS = tuple(f"qfq_{column}" for column in DAILY_ALL_RAW_OHLC_COLUMNS)


def _daily_all_qfq_cache_key(as_of):
    return f"daily_all_qfq_{as_of}_60d_{DAILY_ALL_QFQ_CACHE_VERSION}"


def _date8(value, label):
    text = str(value).strip()
    if not re.fullmatch(r"\d{8}", text):
        raise RuntimeError(f"{label} must be YYYYMMDD")
    return text


def _validate_ohlc(frame, columns, label):
    for column in columns:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    values = frame[list(columns)].to_numpy(dtype=float)
    if not np.isfinite(values).all() or (values <= 0).any():
        raise RuntimeError(f"{label} contains non-finite or non-positive OHLC")
    if (frame[columns["high"]] < frame[[columns["open"], columns["close"]]].max(axis=1)).any():
        raise RuntimeError(f"{label} high is below open/close")
    if (frame[columns["low"]] > frame[[columns["open"], columns["close"]]].min(axis=1)).any():
        raise RuntimeError(f"{label} low is above open/close")


def _normalize_daily_all_raw(frame, as_of, expected_dates):
    required = {"ts_code", "trade_date", *DAILY_ALL_RAW_OHLC_COLUMNS, "pre_close", "pct_chg", "vol", "amount"}
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        raise RuntimeError("daily price payload is empty")
    missing = required - set(frame.columns)
    if missing:
        raise RuntimeError(f"daily price payload missing required fields: {sorted(missing)}")
    result = frame.copy()
    result["ts_code"] = result["ts_code"].astype(str).str.strip()
    if result["ts_code"].eq("").any():
        raise RuntimeError("daily price payload contains blank ts_code")
    result["trade_date"] = result["trade_date"].map(lambda value: _date8(value, "daily trade_date"))
    if (result["trade_date"] > as_of).any() or set(result["trade_date"]) - set(expected_dates):
        raise RuntimeError("daily price payload contains a future or unexpected trade_date")
    if result.duplicated(["ts_code", "trade_date"]).any():
        raise RuntimeError("daily price payload contains duplicate ts_code/trade_date")
    _validate_ohlc(result, {
        "open": "open", "high": "high", "low": "low", "close": "close",
    }, "daily price payload")
    for column in ("pre_close", "pct_chg", "vol", "amount"):
        result[column] = pd.to_numeric(result[column], errors="coerce")
    return result


def _normalize_adj_factors(frame, as_of, expected_dates):
    required = {"ts_code", "trade_date", "adj_factor"}
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        raise RuntimeError("adj_factor payload is empty")
    missing = required - set(frame.columns)
    if missing:
        raise RuntimeError(f"adj_factor payload missing required fields: {sorted(missing)}")
    result = frame.copy()
    result["ts_code"] = result["ts_code"].astype(str).str.strip()
    if result["ts_code"].eq("").any():
        raise RuntimeError("adj_factor payload contains blank ts_code")
    result["trade_date"] = result["trade_date"].map(lambda value: _date8(value, "adj_factor trade_date"))
    if (result["trade_date"] > as_of).any() or set(result["trade_date"]) - set(expected_dates):
        raise RuntimeError("adj_factor payload contains a future or unexpected trade_date")
    if result.duplicated(["ts_code", "trade_date"]).any():
        raise RuntimeError("adj_factor payload contains duplicate ts_code/trade_date")
    if result["adj_factor"].map(lambda value: isinstance(value, (bool, np.bool_))).any():
        raise RuntimeError("adj_factor payload contains boolean factor")
    result["adj_factor"] = pd.to_numeric(result["adj_factor"], errors="coerce")
    factors = result["adj_factor"].to_numpy(dtype=float)
    if not np.isfinite(factors).all() or (factors <= 0).any():
        raise RuntimeError("adj_factor payload contains non-finite or non-positive factor")
    return result


def _build_qfq_daily_all(daily_frames, factor_frames, as_of, expected_dates):
    """Join provider-observed factors and derive one qfq view anchored at ``as_of``.

    Every returned factor is tied to the same stock/date key as a raw bar.  No
    forward-fill, default factor, or future factor may enter the EGS batch.
    """
    as_of = _date8(as_of, "daily qfq as_of")
    expected_dates = [_date8(value, "expected trade_date") for value in expected_dates]
    if not expected_dates or len(daily_frames) != len(expected_dates) or len(factor_frames) != len(expected_dates):
        raise RuntimeError("daily qfq coverage is incomplete")
    raw = _normalize_daily_all_raw(pd.concat(daily_frames, ignore_index=True), as_of, expected_dates)
    factors = _normalize_adj_factors(pd.concat(factor_frames, ignore_index=True), as_of, expected_dates)
    if set(raw["trade_date"]) != set(expected_dates):
        raise RuntimeError("daily price coverage does not match the requested trade-date window")
    if set(factors["trade_date"]) != set(expected_dates):
        raise RuntimeError("adj_factor coverage does not match the requested trade-date window")
    result = raw.merge(factors, on=["ts_code", "trade_date"], how="left", validate="one_to_one")
    if result["adj_factor"].isna().any():
        missing = result.loc[result["adj_factor"].isna(), ["ts_code", "trade_date"]].head(3).to_dict("records")
        raise RuntimeError(f"adj_factor coverage missing for daily price rows: {missing}")
    anchors = (
        result.sort_values(["ts_code", "trade_date"])
        .groupby("ts_code", as_index=False)
        .tail(1)[["ts_code", "trade_date", "adj_factor"]]
        .rename(columns={"trade_date": "qfq_anchor_trade_date", "adj_factor": "qfq_anchor_factor"})
    )
    result = result.merge(anchors, on="ts_code", how="left", validate="many_to_one")
    if (result["qfq_anchor_trade_date"] > as_of).any():
        raise RuntimeError("qfq anchor uses a future factor")
    for raw_column, qfq_column in zip(DAILY_ALL_RAW_OHLC_COLUMNS, DAILY_ALL_QFQ_OHLC_COLUMNS):
        result[qfq_column] = result[raw_column] * result["adj_factor"] / result["qfq_anchor_factor"]
    _validate_ohlc(result, {
        "open": "qfq_open", "high": "qfq_high", "low": "qfq_low", "close": "qfq_close",
    }, "qfq daily price payload")
    result = result.sort_values(["ts_code", "trade_date"], ascending=[True, False]).reset_index(drop=True)
    result["qfq_pre_close"] = result.groupby("ts_code")["qfq_close"].shift(-1)
    result["qfq_pct_chg"] = (result["qfq_close"] / result["qfq_pre_close"] - 1.0) * 100.0
    result["adj_factor_observed"] = True
    result["adj_factor_source"] = "tushare.adj_factor"
    result["qfq_price_basis"] = "qfq_anchored_as_of"
    result["qfq_as_of"] = as_of
    return result


def _validate_cached_qfq_daily_all(cached, as_of, expected_dates):
    required = {
        "ts_code", "trade_date", "adj_factor", "adj_factor_observed", "adj_factor_source",
        "qfq_anchor_trade_date", "qfq_anchor_factor", "qfq_pre_close", "qfq_pct_chg",
        "qfq_price_basis", "qfq_as_of", *DAILY_ALL_RAW_OHLC_COLUMNS, *DAILY_ALL_QFQ_OHLC_COLUMNS,
    }
    if not isinstance(cached, pd.DataFrame):
        raise RuntimeError("daily qfq cache is not a DataFrame")
    missing = required - set(cached.columns)
    if missing:
        raise RuntimeError(f"daily qfq cache lacks required fields: {sorted(missing)}")
    result = _normalize_daily_all_raw(cached, as_of, expected_dates)
    _normalize_adj_factors(result[["ts_code", "trade_date", "adj_factor"]], as_of, expected_dates)
    _validate_ohlc(result, {
        "open": "qfq_open", "high": "qfq_high", "low": "qfq_low", "close": "qfq_close",
    }, "daily qfq cache")
    if set(result["trade_date"]) != set(expected_dates):
        raise RuntimeError("daily qfq cache coverage does not match the requested trade-date window")
    if (result["qfq_as_of"].astype(str) != as_of).any() or (result["qfq_price_basis"] != "qfq_anchored_as_of").any():
        raise RuntimeError("daily qfq cache identity does not match this as_of")
    if not result["adj_factor_observed"].eq(True).all() or not result["adj_factor_source"].eq("tushare.adj_factor").all():
        raise RuntimeError("daily qfq cache contains non-observed adjustment factors")
    result["qfq_anchor_trade_date"] = result["qfq_anchor_trade_date"].map(
        lambda value: _date8(value, "daily qfq cache anchor trade_date")
    )
    result["qfq_anchor_factor"] = pd.to_numeric(result["qfq_anchor_factor"], errors="coerce")
    anchors = result["qfq_anchor_factor"].to_numpy(dtype=float)
    if not np.isfinite(anchors).all() or (anchors <= 0).any():
        raise RuntimeError("daily qfq cache contains invalid anchor factor")
    if (result["qfq_anchor_trade_date"] > as_of).any():
        raise RuntimeError("daily qfq cache contains future anchors")
    anchor_keys = result[["ts_code", "qfq_anchor_trade_date", "qfq_anchor_factor"]].drop_duplicates()
    observed_anchor = result[["ts_code", "trade_date", "adj_factor"]].rename(
        columns={"trade_date": "qfq_anchor_trade_date", "adj_factor": "observed_anchor_factor"}
    )
    anchor_check = anchor_keys.merge(observed_anchor, on=["ts_code", "qfq_anchor_trade_date"], how="left")
    if (len(anchor_check) != len(anchor_keys) or anchor_check["observed_anchor_factor"].isna().any()
            or not np.isclose(
                anchor_check["qfq_anchor_factor"].to_numpy(dtype=float),
                anchor_check["observed_anchor_factor"].to_numpy(dtype=float),
                rtol=1e-12, atol=1e-12,
            ).all()):
        raise RuntimeError("daily qfq cache anchor does not match an observed factor")
    for raw_column, qfq_column in zip(DAILY_ALL_RAW_OHLC_COLUMNS, DAILY_ALL_QFQ_OHLC_COLUMNS):
        expected = result[raw_column] * result["adj_factor"] / result["qfq_anchor_factor"]
        if not np.isclose(result[qfq_column].to_numpy(dtype=float), expected.to_numpy(dtype=float),
                          rtol=1e-12, atol=1e-12).all():
            raise RuntimeError("daily qfq cache price does not match observed adjustment factors")
    expected_pre_close = result.groupby("ts_code")["qfq_close"].shift(-1)
    expected_pct_chg = (result["qfq_close"] / expected_pre_close - 1.0) * 100.0
    if (not np.isclose(result["qfq_pre_close"].to_numpy(dtype=float), expected_pre_close.to_numpy(dtype=float),
                       rtol=1e-12, atol=1e-12, equal_nan=True).all()
            or not np.isclose(result["qfq_pct_chg"].to_numpy(dtype=float), expected_pct_chg.to_numpy(dtype=float),
                              rtol=1e-12, atol=1e-12, equal_nan=True).all()):
        raise RuntimeError("daily qfq cache cross-day fields are inconsistent")
    return result


def get_daily_all(trade_dates):
    if not trade_dates:
        raise RuntimeError("daily qfq fetch requires trade dates")
    n_days = min(60, len(trade_dates))
    expected_dates = [str(value) for value in trade_dates[:n_days]]
    as_of = _date8(expected_dates[0], "daily qfq as_of")
    key = _daily_all_qfq_cache_key(as_of)
    cached = load_cache(key)
    if cached is not None:
        try:
            return _validate_cached_qfq_daily_all(cached, as_of, expected_dates)
        except RuntimeError as exc:
            log.warning(f"日线 qfq 缓存不可用，将重新拉取：{exc}")
    log.info(f"拉取日线和 provider-observed 复权因子（近{n_days}交易日）...")
    daily_frames, factor_frames = [], []
    for d in tqdm(expected_dates, desc="日线/复权因子"):
        df = safe_api(pro.daily, trade_date=d,
                      fields="ts_code,trade_date,open,high,low,close,pre_close,pct_chg,vol,amount")
        factor = safe_api(getattr(pro, "adj_factor", None), trade_date=d,
                          fields="ts_code,trade_date,adj_factor")
        if df is None or df.empty or factor is None or factor.empty:
            raise RuntimeError(f"daily qfq coverage unavailable for {d}; aborting entire EGS batch")
        daily_frames.append(df)
        factor_frames.append(factor)
    result = _build_qfq_daily_all(daily_frames, factor_frames, as_of, expected_dates)
    save_cache(key, result)
    return result

# ── [崩溃修复①] ─────────────────────────────────────────────────────────────
def get_daily_basic(trade_date, fallback_dates=None):
    """
    拉取每日行情基本面。
    若 trade_date 当日无数据（午夜运行 / 非交易日），
    依次尝试 fallback_dates[1..3] 回退到前一个有效交易日。
    """
    requested_trade_date = trade_date
    key = f"daily_basic_{trade_date}_source_v2"
    if (cached := load_cache(key)) is not None:
        if "source_trade_date" not in cached.columns:
            raise RuntimeError("daily_basic cache lacks source_trade_date provenance")
        return cached

    df = safe_api(pro.daily_basic, trade_date=trade_date,
                  fields="ts_code,close,pe,pe_ttm,pb,roe,turnover_rate,total_mv,circ_mv")

    if (df is None or len(df) == 0) and fallback_dates:
        for fb in fallback_dates[1:4]:
            log.warning(f"daily_basic {trade_date} 无数据，回退至 {fb}")
            df = safe_api(pro.daily_basic, trade_date=fb,
                          fields="ts_code,close,pe,pe_ttm,pb,roe,turnover_rate,total_mv,circ_mv")
            if df is not None and len(df) > 0:
                trade_date = fb
                key = f"daily_basic_{trade_date}_source_v2"
                break

    if df is None or len(df) == 0:
        raise RuntimeError(
            f"daily_basic source unavailable for {requested_trade_date}; "
            "refusing quote/unlock inference without an observed market date"
        )

    for col in ["close","pe","pe_ttm","pb","roe","turnover_rate","total_mv","circ_mv"]:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.copy()
    df["source_trade_date"] = trade_date
    save_cache(key, df)
    return df
# ─────────────────────────────────────────────────────────────────────────────

def _validated_suspend_traded_codes(daily_td, all_codes, trade_date, as_of=None, attempted_trade_dates=None):
    if daily_td is None or daily_td.empty:
        return None
    if "ts_code" not in daily_td.columns:
        raise RuntimeError(
            f"suspend daily payload for {trade_date} has no ts_code column; "
            "abort to avoid unsafe suspend inference"
        )
    universe = {str(code) for code in all_codes if pd.notna(code)}
    if not universe:
        raise RuntimeError("stock list is empty; cannot infer suspended stocks safely")

    traded_codes = set(daily_td["ts_code"].dropna().astype(str).tolist())
    in_universe_traded = traded_codes & universe
    coverage = len(in_universe_traded) / len(universe)
    min_coverage = float(CONF["suspend_daily_min_coverage"])
    suspended_count = len(universe - in_universe_traded)
    _record_suspend_daily_coverage_observation(
        as_of=as_of or trade_date,
        trade_date=trade_date,
        status="pass" if coverage >= min_coverage else "fail_low_coverage",
        stock_universe_count=len(universe),
        daily_payload_row_count=len(daily_td),
        traded_in_universe_count=len(in_universe_traded),
        suspended_count=suspended_count,
        coverage_ratio=coverage,
        min_coverage=min_coverage,
        attempted_trade_dates=attempted_trade_dates,
        message=(
            "daily payload coverage is sufficient for suspend inference"
            if coverage >= min_coverage
            else "daily payload coverage is below suspend_daily_min_coverage"
        ),
    )
    if coverage < min_coverage:
        raise RuntimeError(
            f"suspend daily completeness too low for {trade_date}: "
            f"{len(in_universe_traded)}/{len(universe)} in-universe rows "
            f"({coverage:.2%}) below suspend_daily_min_coverage={min_coverage:.2%}; "
            "abort to avoid treating a partial daily response as suspended stocks"
        )
    return in_universe_traded


def get_suspend_info(trade_dates):
    """
    判断停牌股票。
    若当日数据未就绪（开盘前运行），回退到前一交易日。
    若所有候选日均无数据，阻断本次运行；未知源不得伪装成已知空集。
    """
    key = f"suspend_{trade_dates[0]}_v3"
    global _LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION
    _LAST_SUSPEND_DAILY_COVERAGE_OBSERVATION = None
    if (cached := load_cache(key)) is not None:
        if not isinstance(cached, dict) or cached.get("status") not in {"known_clear", "known_hit"}:
            raise RuntimeError("suspend cache lacks tri-state provenance")
        members = set(cached.get("members") or [])
        _record_hard_veto_source(
            "suspension", cached["status"], cached.get("observed_at"),
            source="local_cache", hit_count=len(members),
        )
        _record_suspend_daily_coverage_observation(
            as_of=trade_dates[0],
            trade_date=cached.get("observed_at"),
            status="cache_hit_coverage_not_observed",
            suspended_count=len(members),
            attempted_trade_dates=trade_dates[:3],
            source="local suspend cache",
            message="suspend set loaded from cache; no fresh daily coverage measurement in this run",
        )
        return members
    all_codes = set(get_stock_list()["ts_code"].dropna().astype(str))

    for td in trade_dates[:3]:
        daily_td = safe_api(pro.daily, trade_date=td, fields="ts_code")
        traded_codes = _validated_suspend_traded_codes(
            daily_td,
            all_codes,
            td,
            as_of=trade_dates[0],
            attempted_trade_dates=trade_dates[:3],
        )
        if traded_codes is not None:
            suspended    = all_codes - traded_codes
            log.info(
                f"停牌数据取自 {td}，daily覆盖 {len(traded_codes)}/{len(all_codes)}，"
                f"停牌股 {len(suspended)} 只"
            )
            source_status = "known_hit" if suspended else "known_clear"
            _record_hard_veto_source(
                "suspension", source_status, td,
                source="tushare.daily", hit_count=len(suspended),
            )
            save_cache(key, {
                "status": source_status,
                "observed_at": td,
                "members": sorted(suspended),
            })
            return suspended

    _record_suspend_daily_coverage_observation(
        as_of=trade_dates[0],
        trade_date=None,
        status="no_daily_payload_blocked",
        stock_universe_count=len(all_codes),
        min_coverage=float(CONF["suspend_daily_min_coverage"]),
        attempted_trade_dates=trade_dates[:3],
        message="all candidate pro.daily payloads were empty; actionable output is blocked",
    )
    _record_hard_veto_source(
        "suspension", "unknown", None, source="tushare.daily", hit_count=None,
    )
    raise RuntimeError("suspend source unavailable; refusing to treat unknown as clear")

def _lookback_cutoff_trade_date(trade_dates, lookback):
    lookback = max(1, int(lookback))
    cutoff_idx = min(len(trade_dates) - 1, lookback - 1)
    return trade_dates[cutoff_idx]

def get_relisted_stocks(trade_dates):
    key = f"relisted_{trade_dates[0]}_v2"
    if (cached := load_cache(key)) is not None: return cached
    all_daily = get_daily_all(trade_dates)
    if all_daily.empty:
        save_cache(key, set())          # [小修复④]
        return set()
    first_trade = all_daily.groupby("ts_code")["trade_date"].min()
    cutoff   = _lookback_cutoff_trade_date(trade_dates, CONF["suspend_lookback"])
    relisted = set(first_trade[first_trade >= cutoff].index)
    save_cache(key, relisted)
    return relisted

def _latest_quarters(n=4):
    y, m = TODAY_DT.year, TODAY_DT.month
    quars = []
    for yr in [y-2, y-1, y]:
        for qe, avail_m in [("0331",5),("0630",9),("0930",11),("1231",13)]:
            avail_y  = yr if avail_m <= 12 else yr+1
            avail_mr = avail_m if avail_m <= 12 else avail_m-12
            qdate = f"{yr}{qe}"
            if qdate < TODAY and (y*12+m >= avail_y*12+avail_mr):
                quars.append(qdate)
    result = sorted(quars, reverse=True)[:n]
    if len(result) < n:
        all_q  = [f"{yr}{qe}" for yr in [y-2,y-1,y] for qe in ["0331","0630","0930","1231"]]
        result = sorted([q for q in all_q if q < TODAY], reverse=True)[:n]
    return result

def get_financial_data(ts_codes):
    key = f"financial_{TODAY}_{len(ts_codes)}"
    if (cached := load_cache(key)) is not None: return cached

    quarters  = _latest_quarters(4)   # 4季：配额充足（每日上限10万次），TTM精度最高
    code_list = list(ts_codes)
    fin_chunk_size = int(CONF.get("financial_chunk_size", CONF["chunk_size"]))
    chunks    = to_chunks(code_list, fin_chunk_size)
    log.info(f"财务矩阵拉取：{len(code_list)}只 / {len(chunks)}批 / {len(quarters)}季")

    fina_frames = []
    def _fetch_fina_indicator(chunk, q):
        chunk = list(chunk)
        chunk_str = ",".join(chunk)
        fi = safe_api(
            pro.fina_indicator,
            ts_code=chunk_str,
            period=q,
            fields=("ts_code,ann_date,end_date,dt_netprofit_yoy,"
                    "tr_yoy,ocf_to_profit,profit_dedt,dtprofit_to_profit,roe"),
            retries=2,
        )
        if fi is not None:
            return fi
        min_chunk = int(CONF.get("financial_min_chunk", 10))
        if len(chunk) <= min_chunk:
            return None
        mid = max(len(chunk) // 2, 1)
        log.warning(f"财报切片 {q} 批次{len(chunk)}只失败，自动拆分为 {mid}/{len(chunk)-mid} 只重试")
        parts = []
        left = _fetch_fina_indicator(chunk[:mid], q)
        if left is not None:
            parts.append(left)
        right = _fetch_fina_indicator(chunk[mid:], q)
        if right is not None:
            parts.append(right)
        return pd.concat(parts, ignore_index=True) if parts else None

    for q in quarters:
        for chunk in tqdm(chunks, desc=f"财报切片 {q}"):
            # income 接口强制要求单个 ts_code，不支持批量，故从 fina_indicator 补充净利润字段
            fi = _fetch_fina_indicator(chunk, q)
            if fi is not None:
                fi["quarter"] = q
                fina_frames.append(fi)

    if not fina_frames: return pd.DataFrame()
    df_fi = pd.concat(fina_frames, ignore_index=True)

    df_fi["ann_date"] = pd.to_datetime(df_fi["ann_date"], format="%Y%m%d", errors="coerce")
    df_fi = df_fi[df_fi["ann_date"] <= pd.Timestamp(TODAY_DT)]
    # Keep the explicit disclosure date in a serialisable canonical form for
    # the downstream Rule6 revenue-growth PIT gate.
    df_fi["ann_date"] = df_fi["ann_date"].dt.strftime("%Y%m%d")
    df_fi = df_fi.sort_values("ann_date", ascending=False).drop_duplicates(subset=["ts_code","quarter"])
    log.info(f"fina_indicator 合并后：{len(df_fi)} 行")

    q0, q1 = quarters[0], (quarters[1] if len(quarters) > 1 else None)
    uniq      = sorted(set(df_fi["ts_code"].tolist()))
    df_merged = pd.DataFrame({"ts_code": uniq})

    fi0_cols = ["ts_code", "ann_date", "end_date", "dt_netprofit_yoy", "tr_yoy",
                "profit_dedt", "dtprofit_to_profit", "roe"]
    fi0_cols = [c for c in fi0_cols if c in df_fi.columns]
    fi0 = df_fi[df_fi["quarter"]==q0][fi0_cols].copy()
    rename_map = {"ann_date": "q0_ann_date", "end_date": "q0_end_date",
                  "dt_netprofit_yoy":"q0_dt_yoy", "tr_yoy": "q0_revenue_yoy",
                  "profit_dedt":"q0_profit_dedt",
                  "dtprofit_to_profit":"q0_dt_profit_ratio", "roe":"roe"}
    fi0.columns = [rename_map.get(c, c) for c in fi0.columns]
    df_merged = df_merged.merge(fi0, on="ts_code", how="left")

    if q1:
        fi1_cols = [column for column in ["ts_code", "ann_date", "end_date", "dt_netprofit_yoy", "tr_yoy"]
                    if column in df_fi.columns]
        fi1 = df_fi[df_fi["quarter"]==q1][fi1_cols].copy()
        fi1.columns = [{"ann_date": "q1_ann_date", "end_date": "q1_end_date", "dt_netprofit_yoy": "q1_dt_yoy",
                        "tr_yoy": "q1_revenue_yoy"}.get(column, column) for column in fi1.columns]
        df_merged = df_merged.merge(fi1, on="ts_code", how="left")
    else:
        df_merged["q1_dt_yoy"] = np.nan
    for column in ("q0_ann_date", "q1_ann_date", "q0_end_date", "q1_end_date",
                   "q0_revenue_yoy", "q1_revenue_yoy"):
        if column not in df_merged.columns:
            df_merged[column] = np.nan

    # TTM 扣非净利润（fina_indicator 批量可用）
    fi_4q = df_fi[df_fi["quarter"].isin(quarters[:4])].copy()
    fi_4q["profit_dedt"] = pd.to_numeric(fi_4q["profit_dedt"], errors="coerce")
    ttm_dedt = fi_4q.groupby("ts_code")["profit_dedt"].sum().reset_index()
    ttm_dedt.columns = ["ts_code","ttm_profit_dedt"]
    df_merged = df_merged.merge(ttm_dedt, on="ts_code", how="left")

    # income 接口不支持批量，净利润改由 dtprofit_to_profit 比率替代质量检查
    df_merged["ttm_net_income"] = np.nan
    df_merged["q0_net_income"]  = np.nan

    fi0_ocf = df_fi[df_fi["quarter"]==q0][["ts_code","ocf_to_profit"]].copy()
    fi0_ocf.columns = ["ts_code","ttm_ocf_ratio"]
    df_merged = df_merged.merge(fi0_ocf, on="ts_code", how="left")

    for col in ["q0_dt_yoy","q1_dt_yoy","q0_revenue_yoy","q1_revenue_yoy","q0_profit_dedt","q0_dt_profit_ratio",
                "ttm_profit_dedt","ttm_ocf_ratio","roe"]:
        if col in df_merged.columns:
            df_merged[col] = pd.to_numeric(df_merged[col], errors="coerce")

    save_cache(key, df_merged)
    return df_merged

def get_moneyflow(trade_dates):
    key = f"moneyflow_{trade_dates[0]}"
    if (cached := load_cache(key)) is not None: return cached
    frames, collected = [], 0
    for d in tqdm(trade_dates[:5], desc="资金流"):
        df = safe_api(pro.moneyflow, trade_date=d,
                      fields="ts_code,trade_date,"
                             "buy_elg_amount,sell_elg_amount,"
                             "buy_lg_amount,sell_lg_amount,"
                             "buy_md_amount,sell_md_amount,"
                             "buy_sm_amount,sell_sm_amount,"
                             "net_mf_amount",
                      retries=2)
        if df is not None and len(df) > 0:
            frames.append(df)
            collected += 1
            if collected >= 5: break
        else:
            log.warning(f"资金流日期 {d} 无数据，已跳过")
    if not frames:
        log.error("近5日资金流全部拉取失败，大单流向加分将缺失")
        save_cache(key, pd.DataFrame())
        return pd.DataFrame()
    result = pd.concat(frames, ignore_index=True)
    # trade_amount 不是 Tushare 有效字段，本地从各档位买卖额之和推算
    buy_cols  = ["buy_elg_amount","buy_lg_amount","buy_md_amount","buy_sm_amount"]
    sell_cols = ["sell_elg_amount","sell_lg_amount","sell_md_amount","sell_sm_amount"]
    for col in buy_cols + sell_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0)
    result["trade_amount"] = (
        sum(result[c] for c in buy_cols  if c in result.columns) +
        sum(result[c] for c in sell_cols if c in result.columns)
    )
    save_cache(key, result)
    return result

def get_margin(trade_dates):
    # A ten-session change requires today's observation plus the observation
    # ten trading sessions earlier; v3 separates the old ten-row cache.
    key = f"margin_{trade_dates[0]}_rule6_v3"
    if (cached := load_cache(key)) is not None: return cached
    frames = []
    for d in tqdm(trade_dates[:11], desc="两融"):
        df = safe_api(pro.margin_detail, trade_date=d, fields="ts_code,trade_date,rzye,rqye")
        if df is not None: frames.append(df)
    result = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    for column in ("rzye", "rqye"):
        if not result.empty and column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce")
    save_cache(key, result)
    return result


def _rule6_fetch_dataframe(fn, **kwargs):
    """Preserve an empty successful response; ``safe_api`` intentionally cannot."""
    if not callable(fn):
        return None
    for attempt in range(3):
        try:
            time.sleep(CONF["request_delay"])
            result = fn(**kwargs)
            return result if isinstance(result, pd.DataFrame) else None
        except Exception as exc:
            log.warning(f"Rule6 API exception [{getattr(fn, '__name__', 'unknown')}] #{attempt + 1}: {exc}")
            if attempt < 2:
                time.sleep(2 ** attempt)
    return None


def _rule6_cache_suffix(ts_codes):
    canonical = ",".join(sorted({str(code) for code in ts_codes if str(code)}))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


def get_rule6_balancesheets(ts_codes):
    """Fetch candidate-only PIT-safe balancesheets for Rule6; missing coverage stays unknown."""
    codes = sorted({str(code) for code in ts_codes if str(code)})
    if not codes:
        return {}
    key = f"rule6_balancesheet_{TODAY}_{_rule6_cache_suffix(codes)}"
    if (cached := load_cache(key)) is not None:
        return cached
    fields = "ts_code,ann_date,end_date,money_cap,st_borr,total_assets,accounts_receiv,contract_liab"
    result = {}
    endpoint = getattr(pro, "balancesheet", None)
    for code in tqdm(codes, desc="Rule6资产负债表"):
        df = _rule6_fetch_dataframe(endpoint, ts_code=code, report_type="1", fields=fields)
        required = {"ts_code", "ann_date", "end_date", "money_cap", "st_borr", "total_assets",
                    "accounts_receiv", "contract_liab"}
        if df is None or not required.issubset(set(df.columns)):
            result[code] = None
            continue
        valid_rows = []
        malformed = False
        for _, row in df.iterrows():
            ann_date_value, end_date_value = _json_value(row.get("ann_date")), _json_value(row.get("end_date"))
            ann_date = str(ann_date_value) if ann_date_value is not None else ""
            end_date = str(end_date_value) if end_date_value is not None else ""
            if not (re.fullmatch(r"\d{8}", ann_date or "") and re.fullmatch(r"\d{8}", end_date or "")):
                malformed = True
                break
            if ann_date > TODAY or end_date > TODAY:
                continue
            valid_rows.append({
                "ann_date": ann_date, "end_date": end_date,
                "money_cap": _json_float(row.get("money_cap")),
                "st_borr": _json_float(row.get("st_borr")),
                "total_assets": _json_float(row.get("total_assets")),
                "accounts_receiv": _json_float(row.get("accounts_receiv")),
                "contract_liab": _json_float(row.get("contract_liab")),
            })
        if malformed or not valid_rows:
            result[code] = None
            continue
        valid_rows.sort(key=lambda item: (item["end_date"], item["ann_date"]), reverse=True)
        deduped = {}
        for item in valid_rows:
            deduped.setdefault(item["end_date"], item)
        result[code] = list(deduped.values())
    save_cache(key, result)
    return result


def get_rule6_block_trades(trade_dates):
    """Fetch exactly the Rule6 ten-trading-day block-trade coverage window."""
    window = [str(day) for day in list(trade_dates)[:10]]
    if len(window) != 10:
        return {day: None for day in window}
    key = f"rule6_block_trade_{window[0]}_{window[-1]}"
    if (cached := load_cache(key)) is not None:
        return cached
    endpoint = getattr(pro, "block_trade", None)
    result = {}
    required = {"ts_code", "trade_date", "price", "vol"}
    for trade_date in tqdm(window, desc="Rule6大宗交易"):
        df = _rule6_fetch_dataframe(endpoint, trade_date=trade_date,
                                    fields="trade_date,ts_code,price,vol,amount")
        if df is None or not required.issubset(set(df.columns)):
            result[trade_date] = None
            continue
        if df.empty:
            result[trade_date] = []
            continue
        result[trade_date] = [
            {"ts_code": str(_json_value(row.get("ts_code")) or ""),
             "trade_date": str(_json_value(row.get("trade_date")) or ""),
             "price": _json_float(row.get("price")), "vol": _json_float(row.get("vol"))}
            for _, row in df.iterrows()
        ]
    save_cache(key, result)
    return result


def _rule6_date8(value):
    value = _json_value(value)
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text if re.fullmatch(r"\d{8}", text) else None


def _rule6_daily_rows(all_daily, ts_code, dates, *, price_basis="raw"):
    if all_daily is None or all_daily.empty or not {"ts_code", "trade_date"}.issubset(all_daily.columns):
        return []
    if price_basis not in {"raw", "qfq"}:
        raise ValueError(f"unsupported Rule6 price basis: {price_basis}")
    qfq_required = {"qfq_high", "qfq_low", "qfq_close", "qfq_pct_chg"}
    if price_basis == "qfq" and not qfq_required.issubset(all_daily.columns):
        return []
    rows = []
    for trade_date in dates:
        match = all_daily[(all_daily["ts_code"].astype(str) == str(ts_code))
                          & (all_daily["trade_date"].astype(str) == str(trade_date))]
        if len(match) != 1:
            return []
        row = match.iloc[0].to_dict()
        if price_basis == "qfq":
            row.update({
                "high": row["qfq_high"],
                "low": row["qfq_low"],
                "close": row["qfq_close"],
                "pct_chg": row["qfq_pct_chg"],
            })
        rows.append(row)
    return rows


def _rule6_margin_value(margin_df, ts_code, trade_date, field):
    if margin_df is None or margin_df.empty or not {"ts_code", "trade_date", field}.issubset(margin_df.columns):
        return None
    match = margin_df[(margin_df["ts_code"].astype(str) == str(ts_code))
                      & (margin_df["trade_date"].astype(str) == str(trade_date))]
    return _json_float(match.iloc[0].get(field)) if len(match) == 1 else None


def _rule6_revenue_periods(df_fin, ts_code):
    if df_fin is None or df_fin.empty or "ts_code" not in df_fin.columns:
        return []
    match = df_fin[df_fin["ts_code"].astype(str) == str(ts_code)]
    if len(match) != 1:
        return []
    row = match.iloc[0]
    periods = []
    for prefix in ("q0", "q1"):
        period = _rule6_date8(row.get(f"{prefix}_end_date"))
        ann_date = _rule6_date8(row.get(f"{prefix}_ann_date"))
        revenue_yoy = _json_float(row.get(f"{prefix}_revenue_yoy"))
        if period is None or ann_date is None or revenue_yoy is None:
            return []
        periods.append({"period": period, "ann_date": ann_date, "revenue_yoy_pct": revenue_yoy})
    return periods


# The Shanghai+Shenzhen 融资融券 (margin) target universe is a large, stable set
# (thousands of securities for years).  A fetched reference-date universe far
# below this floor is a partial / garbage / truncated provider response, not a
# real universe; treating a candidate's absence from it as "non-margin" would
# fail OPEN.  Below the floor the two margin Rule6 checks stay `unknown`
# (fail-closed), per R-ASHORT-RULE6-MARGIN-ELIGIBILITY-DISPOSITION.
MARGIN_ELIGIBILITY_MIN_UNIVERSE = 1000


_ASHARE_TS_CODE_RE = re.compile(r"^\d{6}\.(SH|SZ|BJ)$")


def _canonical_ashare_ts_code(value):
    """Canonical stripped/upper ``NNNNNN.XX`` A-share code, or None if malformed.

    Margin-universe codes and candidate codes are compared in this single
    canonical namespace so provider whitespace/case/suffix/float drift cannot
    masquerade as a different (non-matching) universe and silently clear the
    margin vetoes; non-canonical shapes are dropped (they never reach the floor).
    """
    text = str(value).strip().upper()
    return text if _ASHARE_TS_CODE_RE.match(text) else None


def _clean_margin_ts_codes(frame):
    """Distinct CANONICAL A-share ts_codes in a margin frame (drops malformed/garbage)."""
    if not (isinstance(frame, pd.DataFrame) and not frame.empty and "ts_code" in frame.columns):
        return set()
    codes = set()
    for value in frame["ts_code"].dropna():
        canon = _canonical_ashare_ts_code(value)
        if canon is not None:
            codes.add(canon)
    return codes


def _collect_rule6_evaluations(watch_df, all_daily, margin_df, trade_dates, red_dict, df_fin):
    """Evaluate every EGS-computable Rule6 item for the final candidate set.

    This function intentionally runs after ranking: the newly authorized
    balancesheet and block-trade calls are candidate-only and never influence
    the screening scores or rank order.
    """
    if watch_df is None or watch_df.empty or "ts_code" not in watch_df.columns:
        return {}
    codes = [str(code) for code in watch_df["ts_code"].dropna().astype(str).tolist()]
    if len(trade_dates) < 11:
        return {code: {} for code in codes}
    balance_by_code = get_rule6_balancesheets(codes)
    block_by_date = get_rule6_block_trades(trade_dates[:10])
    holder_events = (red_dict or {}).get("rule6_holder_events")
    # Margin-eligibility (两融标的) is inferred from the fetched margin universe:
    # a stock absent from a reference-date universe that is both COMPLETE
    # (>= MARGIN_ELIGIBILITY_MIN_UNIVERSE canonical codes) and CLEAN (no malformed
    # ts_code) is a genuine non-margin target and the two margin Rule6 checks do
    # not apply to it.  Any provider anomaly must never become a clear result
    # (R-ASHORT-RULE6-MARGIN-ELIGIBILITY-DISPOSITION): an empty / sub-floor /
    # truncated response fails the size gate, and any garbage or malformed
    # (null / non-canonical) reference-date ts_code fails the clean gate -- so a
    # selective corruption of one candidate's own row cannot masquerade as its
    # absence.  In every anomaly eligibility stays None → fail-closed `unknown`.
    margin_ref_codes = set()
    margin_ref_has_malformed = False
    if (isinstance(margin_df, pd.DataFrame) and not margin_df.empty
            and {"ts_code", "trade_date", "rzye", "rqye"}.issubset(margin_df.columns)):
        ref_frame = margin_df[margin_df["trade_date"].astype(str) == str(trade_dates[0])]
        for value in ref_frame["ts_code"].dropna():
            canon = _canonical_ashare_ts_code(value)
            if canon is None:
                margin_ref_has_malformed = True   # a corrupt row cannot be trusted to prove absence
            else:
                margin_ref_codes.add(canon)
    margin_universe_present = (
        len(margin_ref_codes) >= MARGIN_ELIGIBILITY_MIN_UNIVERSE
        and not margin_ref_has_malformed
    )
    margin_eligible_codes = _clean_margin_ts_codes(margin_df) if margin_universe_present else None
    evaluations_by_code = {}
    for code in codes:
        daily_11 = _rule6_daily_rows(all_daily, code, trade_dates[:11], price_basis="qfq")
        daily_6 = daily_11[:6] if len(daily_11) >= 6 else []
        latest_daily = daily_11[0] if len(daily_11) == 11 else {}
        oldest_daily = daily_11[10] if len(daily_11) == 11 else {}
        if not isinstance(holder_events, list):
            code_holder_events = None
        else:
            code_holder_events = [
                event for event in holder_events
                if not isinstance(event, dict) or str(event.get("ts_code")) == code
            ]
        code_balances = balance_by_code.get(code)
        latest_balance = code_balances[0] if isinstance(code_balances, list) and code_balances else None
        balance_by_period = {
            item["end_date"]: item for item in (code_balances or [])
            if isinstance(item, dict) and item.get("end_date")
        }
        block_records = {}
        for trade_date in trade_dates[:10]:
            source_rows = block_by_date.get(str(trade_date)) if isinstance(block_by_date, dict) else None
            if source_rows is None:
                block_records[str(trade_date)] = None
                continue
            candidate_rows = []
            daily_row = _rule6_daily_rows(all_daily, code, [trade_date], price_basis="raw")
            close = _json_float(daily_row[0].get("close")) if daily_row else None
            for item in source_rows:
                if not isinstance(item, dict) or str(item.get("ts_code")) != code:
                    continue
                if str(item.get("trade_date")) != str(trade_date):
                    candidate_rows.append({"price": None, "vol": None, "close": None})
                else:
                    candidate_rows.append({"price": item.get("price"), "vol": item.get("vol"), "close": close})
            block_records[str(trade_date)] = candidate_rows
        canonical_code = _canonical_ashare_ts_code(code)
        is_margin_eligible = (
            None if (margin_eligible_codes is None or canonical_code is None)
            else (canonical_code in margin_eligible_codes)
        )
        evaluations_by_code[code] = {
            "rule6_holder_below_5pct": evaluate_holder_below_5pct(code_holder_events, TODAY),
            "rule6_volume_stall": evaluate_volume_stall(daily_6),
            "rule6_margin_extreme_accumulation": evaluate_margin_extreme_accumulation(
                _rule6_margin_value(margin_df, code, trade_dates[0], "rzye"),
                _rule6_margin_value(margin_df, code, trade_dates[10], "rzye"),
                _json_float(latest_daily.get("close")), _json_float(oldest_daily.get("close")),
                is_margin_eligible=is_margin_eligible,
            ),
            "rule6_short_selling_surge": evaluate_short_selling_surge(
                _rule6_margin_value(margin_df, code, trade_dates[0], "rqye"),
                _rule6_margin_value(margin_df, code, trade_dates[6], "rqye"),
                hedge_announcement_status=None,
                is_margin_eligible=is_margin_eligible,
            ),
            "rule6_cash_debt_double_high": evaluate_cash_debt_double_high(
                latest_balance, TODAY,
            ),
            "rule6_ar_growth_gt_revenue_growth": evaluate_ar_growth_gt_revenue_growth(
                _rule6_revenue_periods(df_fin, code), balance_by_period, TODAY,
            ),
            "rule6_block_trade_discount": evaluate_block_trade_discount(trade_dates[:10], block_records),
        }
    return evaluations_by_code

# ── [崩溃修复②] ─────────────────────────────────────────────────────────────
def get_unlock_future(stock_list, daily_basic_df):
    key = f"unlock_future_{TODAY}_v2"
    global _LAST_UNLOCK_DETAILS
    _LAST_UNLOCK_DETAILS = {}
    if (cached := load_cache(key)) is not None:
        if not isinstance(cached, dict) or cached.get("status") not in {"known_clear", "known_hit"}:
            raise RuntimeError("unlock cache lacks tri-state provenance")
        _LAST_UNLOCK_DETAILS = dict(cached.get("details") or {})
        members = set(cached.get("members") or [])
        _record_hard_veto_source(
            "unlock", cached["status"], cached.get("observed_at"),
            source="local_cache", hit_count=len(members),
        )
        return members

    if daily_basic_df.empty or "close" not in daily_basic_df.columns:
        _record_hard_veto_source("unlock", "unknown", None, source="tushare.share_float")
        raise RuntimeError("unlock denominator source unavailable; refusing unknown as clear")

    df_float = safe_api(getattr(pro, "share_float", None), start_date=TODAY, end_date=dfuture(30),
                        fields="ts_code,ann_date,float_date,float_share,float_ratio")
    if df_float is None:
        _record_hard_veto_source("unlock", "unknown", None, source="tushare.share_float")
        raise RuntimeError("unlock source unavailable; refusing unknown as clear")
    if df_float.empty:
        payload = {"status": "known_clear", "observed_at": TODAY, "members": [], "details": {}}
        _record_hard_veto_source("unlock", "known_clear", TODAY, source="tushare.share_float", hit_count=0)
        save_cache(key, payload)
        return set()
    required = {"ts_code", "ann_date", "float_date", "float_share"}
    missing = required - set(df_float.columns)
    if missing:
        _record_hard_veto_source("unlock", "unknown", None, source="tushare.share_float")
        raise RuntimeError(f"unlock source missing required fields: {sorted(missing)}")
    ann_dates = df_float["ann_date"].astype(str)
    parsed_ann = pd.to_datetime(ann_dates, format="%Y%m%d", errors="coerce")
    if parsed_ann.isna().any() or (ann_dates > TODAY).any():
        _record_hard_veto_source("unlock", "unknown", None, source="tushare.share_float")
        raise RuntimeError("unlock source PIT violation: ann_date must be valid and <= as_of")

    db = daily_basic_df[["ts_code","close","circ_mv"]].copy()
    db["close"]   = pd.to_numeric(db["close"],   errors="coerce")
    db["circ_mv"] = pd.to_numeric(db["circ_mv"], errors="coerce")
    # circ_mv(万元) / close(元/股) = 万股（正确单位，不需要额外除以10000）
    db["circ_share"] = db["circ_mv"] / db["close"]

    df = df_float.merge(db, on="ts_code", how="left")
    df["float_share"] = pd.to_numeric(df["float_share"], errors="coerce")
    df["unlock_pct"] = np.nan
    mask_circ = df["circ_share"].notna() & (df["circ_share"] > 0)
    df.loc[mask_circ & df["float_share"].notna(), "unlock_pct"] = (
        df["float_share"] / df["circ_share"] * 100
    )
    # Missing a real circulating-share denominator is unknown, not clear.  Block
    # the affected symbol conservatively; float_ratio is provider-relative and
    # must never be converted with a guessed coefficient.
    unknown_denominator = set(df[df["unlock_pct"].isna()]["ts_code"].dropna().astype(str))
    large = set(df[df["unlock_pct"] > CONF["unlock_ratio"]]["ts_code"].dropna().astype(str))
    blocked = large | unknown_denominator
    details = {}
    for _, item in df.iterrows():
        code = str(item.get("ts_code"))
        details[code] = {
            "status": "unknown" if code in unknown_denominator else ("known_hit" if code in large else "known_clear"),
            "observed_at": str(item.get("ann_date")),
            "unlock_date": str(item.get("float_date")),
            "unlock_pct": _json_float(item.get("unlock_pct")),
            "denominator": "circ_mv_div_close" if code not in unknown_denominator else "missing",
        }
    _LAST_UNLOCK_DETAILS = details
    source_status = "known_hit" if blocked else "known_clear"
    payload = {"status": source_status, "observed_at": TODAY, "members": sorted(blocked), "details": details}
    _record_hard_veto_source(
        "unlock", source_status, TODAY, source="tushare.share_float",
        hit_count=len(large), unknown_denominator_count=len(unknown_denominator),
    )
    save_cache(key, payload)
    return blocked
# ─────────────────────────────────────────────────────────────────────────────

def get_holder_reductions():
    key = f"reductions_{TODAY}_rule6_v3"
    if (cached := load_cache(key)) is not None:
        if not isinstance(cached, dict) or cached.get("source_status") not in {"known_clear", "known_hit"}:
            raise RuntimeError("holder-reduction cache lacks tri-state provenance")
        result = {
            "veto_10d": set(cached.get("veto_10d") or []),
            "deduct_30d": set(cached.get("deduct_30d") or []),
            "rule6_holder_events": cached.get("rule6_holder_events"),
        }
        _record_hard_veto_source(
            "holder_reduction", cached["source_status"], cached.get("observed_at"),
            source="local_cache", hit_count=len(result["veto_10d"] | result["deduct_30d"]),
        )
        return result
    df = safe_api(getattr(pro, "stk_holdertrade", None), start_date=dstr(30), end_date=TODAY,
                  fields="ts_code,ann_date,in_de,after_ratio")
    if df is None:
        _record_hard_veto_source("holder_reduction", "unknown", None, source="tushare.stk_holdertrade")
        raise RuntimeError("holder-reduction source unavailable; refusing unknown as clear")
    if len(df) == 0:
        empty_res = {"veto_10d": set(), "deduct_30d": set(), "rule6_holder_events": []}
        _record_hard_veto_source("holder_reduction", "known_clear", TODAY, source="tushare.stk_holdertrade", hit_count=0)
        save_cache(key, {"source_status": "known_clear", "observed_at": TODAY,
                         "veto_10d": [], "deduct_30d": [], "rule6_holder_events": []})
        return empty_res
    required = {"ts_code", "ann_date", "in_de"}
    missing = required - set(df.columns)
    if missing:
        raise RuntimeError(f"holder-reduction source missing required fields: {sorted(missing)}")
    ann_dates = df["ann_date"].astype(str)
    parsed_ann = pd.to_datetime(ann_dates, format="%Y%m%d", errors="coerce")
    if parsed_ann.isna().any() or (ann_dates > TODAY).any():
        _record_hard_veto_source("holder_reduction", "unknown", None, source="tushare.stk_holdertrade")
        raise RuntimeError("holder-reduction PIT violation: ann_date must be valid and <= as_of")
    df = df[df["in_de"]=="DE"].copy()
    df["ann_dt"] = pd.to_datetime(df["ann_date"], format="%Y%m%d", errors="coerce")
    cut10 = TODAY_DT - timedelta(days=10)
    v10   = set(df[df["ann_dt"] >= cut10]["ts_code"])
    v30   = set(df["ts_code"]) - v10
    # PIT 修复(未来函数 look-ahead):移除原"未来30日"第二段查询。原 start=TODAY / end=dfuture(30) 按
    # ann_date 抓 as_of **之后**才公告的减持 → 历史 --as-of / 回测 look-ahead(把 as_of 时点尚不可知的未来
    # 减持并入 veto_10d、提前剔除候选、回测虚高)。对实盘(as_of==今天)该段恒**冗余**:今日公告 ann_date==as_of
    # 已被上面第一段(as_of-30 ≤ ann_date ≤ as_of)纳入 v10,故移除后**实盘行为不变、仅消除历史污染**。
    # 已公告且执行在未来的减持计划由第一段按 ann_date 捕获(ann_date≤as_of);未来才公告的不属 PIT 可知。
    rule6_events = None
    if "after_ratio" in df.columns:
        rule6_events = [
            {"ts_code": str(row["ts_code"]), "ann_date": str(row["ann_date"]),
             "in_de": str(row["in_de"]), "after_ratio": _json_float(row["after_ratio"])}
            for _, row in df.iterrows()
        ]
    result = {"veto_10d": v10, "deduct_30d": v30, "rule6_holder_events": rule6_events}
    source_status = "known_hit" if (v10 or v30) else "known_clear"
    _record_hard_veto_source(
        "holder_reduction", source_status, TODAY, source="tushare.stk_holdertrade",
        hit_count=len(v10 | v30),
    )
    save_cache(key, {
        "source_status": source_status, "observed_at": TODAY,
        "veto_10d": sorted(v10), "deduct_30d": sorted(v30),
        "rule6_holder_events": rule6_events,
    })
    return result


# ═══════════════════════════════════════════════════
# §3 预计算统计量
# ═══════════════════════════════════════════════════
def _neutral_stats_df(codes):
    basic = pd.DataFrame({"ts_code": list(codes)})
    for col in ["pct_20d","pct_5d","pct_60d","avg_amount_20d","avg_amount_5d","high_20d","low_20d","drawdown_20d"]:
        basic[col] = np.nan
    basic["vol_confirm"]    = True
    basic["limit_20d"]      = 0
    basic["limit_10d"]      = 0
    basic["is_lock"]        = False
    basic["is_breakout"]    = False
    basic["limit_breakout_legacy"] = False
    basic["has_crash_veto"] = False
    return basic


def precompute_stock_stats(codes, all_daily):
    min_rows = int(CONF["daily_stats_min_rows"])
    if all_daily.empty:
        raise RuntimeError(
            "daily stats coverage too low: all_daily is empty; "
            "abort to avoid neutral pass-through stats"
        )
    if len(all_daily) < min_rows:
        raise RuntimeError(
            f"daily stats coverage too low: {len(all_daily)} rows below "
            f"daily_stats_min_rows={min_rows}; abort to avoid neutral pass-through stats"
        )

    qfq_required = {"qfq_open", "qfq_high", "qfq_low", "qfq_close"}
    missing_qfq = qfq_required - set(all_daily.columns)
    if missing_qfq:
        raise RuntimeError(
            f"daily stats require qfq OHLC; missing {sorted(missing_qfq)}; "
            "refusing raw-price fallback"
        )

    ad = all_daily[all_daily["ts_code"].isin(codes)].copy()
    if len(ad) < min_rows:
        raise RuntimeError(
            f"daily stats coverage too low after stock-universe match: {len(ad)} rows below "
            f"daily_stats_min_rows={min_rows}; abort to avoid neutral pass-through stats"
        )
    ad = ad.sort_values(["ts_code","trade_date"], ascending=[True, False])

    rows = []
    for code, grp in tqdm(ad.groupby("ts_code"), desc="全市场量能预计算"):
        grp    = grp.reset_index(drop=True)
        closes = grp["qfq_close"].dropna()
        if len(closes) < 1: continue

        pct_20d = float((closes.iloc[0] / closes.iloc[min(19, len(closes)-1)] - 1) * 100) if len(closes) >= 2 else np.nan
        pct_5d  = float((closes.iloc[0] / closes.iloc[min(4,  len(closes)-1)] - 1) * 100) if len(closes) >= 2 else np.nan
        pct_60d = float((closes.iloc[0] / closes.iloc[min(59, len(closes)-1)] - 1) * 100) if len(closes) >= 2 else np.nan
        avg_20d = grp.head(20)["amount"].mean() * 1000 if "amount" in grp.columns else np.nan
        avg_5d  = grp.head(5) ["amount"].mean() * 1000 if "amount" in grp.columns else np.nan

        r5     = grp.head(5)
        up_amt = r5.loc[r5["pct_chg"] >  0, "amount"].mean() if "pct_chg" in r5.columns else np.nan
        dn_amt = r5.loc[r5["pct_chg"] <= 0, "amount"].mean() if "pct_chg" in r5.columns else np.nan
        if pd.isna(up_amt):
            vol_confirm = False       # 近5日无上涨日，量能不确认
        elif pd.isna(dn_amt):
            vol_confirm = True        # 近5日全部上涨，量能强
        else:
            vol_confirm = bool(up_amt > dn_amt)

        # Exchange mechanisms remain on raw daily prices.  qfq is only for
        # cross-day price structures below; it must not redefine limit-up.
        grp["limit_price"] = (grp["pre_close"] * 1.10).round(2)
        grp["is_limit"]    = (np.abs(grp["close"] - grp["limit_price"]) < 0.01) & (grp["high"] == grp["close"])
        is_lock     = len(grp) >= 2 and bool(grp.loc[0,"is_limit"]) and bool(grp.loc[1,"is_limit"])
        limit_20d   = int(grp.head(20)["is_limit"].sum())
        limit_10d   = int(grp.head(10)["is_limit"].sum())
        # 旧口径(近20日涨停≥3 且 近10日涨停≥1)保留为审计字段,便于看新旧 is_breakout 差异;不进 analysis_input、不进评分。
        limit_breakout_legacy = (limit_20d >= 3) and (limit_10d >= 1)
        # is_breakout 改用 v14.2 spec §M3.2 突破型口径:现价站稳 MA10 上方 且 当日量 > 5日均量×1.2(成交额代理)。
        ma10 = float(closes.head(10).mean()) if len(closes) >= 10 else np.nan
        amt0 = float(grp.iloc[0]["amount"]) if ("amount" in grp.columns and pd.notna(grp.iloc[0]["amount"])) else np.nan
        amt5 = float(grp.head(5)["amount"].mean()) if "amount" in grp.columns else np.nan
        is_breakout = bool(
            len(closes) >= 10 and not pd.isna(ma10) and float(closes.iloc[0]) >= ma10
            and not pd.isna(amt0) and not pd.isna(amt5) and amt5 > 0 and amt0 > amt5 * 1.2
        )

        # 闪崩/断头铡刀检测。**有意偏离 v14.2 Rule6「5日内放量跌>8%」**(2026-06-19 审查决定保留):此处用更稳健的
        # **价格结构**口径——最近 5 个已有次日确认的交易日内，单日跌>5% ∧ 收在当日振幅下 20%(收得弱)
        # ∧ 次日收盘<(pre_close+close)/2(不修复)。最新交易日尚无次日确认，故不纳入本次窗口。
        # 不依赖含糊的「放量」量级,且加「弱收 + 次日不修复」两道结构确认,对 risk_filter_only 系统更保守可靠。
        # 阈值 −5(非 −8)+ 结构门是 deliberate(非把 8 打错成 5);改 −8 / 加放量 = 放松一条硬否决,故不改。
        # 行为由 tests/phase6/test_egs_main_board_and_holder_pit.py::HasCrashVetoSpecDeviationTest 钉住。
        has_crash_veto = False
        for i in range(1, min(6, len(grp))):
            # A corporate action must not look like a crash.  The threshold
            # and recovery comparison are therefore derived from qfq bars.
            day_chg = (closes.iloc[i] / closes.iloc[i + 1] - 1.0) * 100 if i + 1 < len(closes) else np.nan
            if day_chg < -5:
                high_p  = grp.iloc[i].get("qfq_high",  0)
                low_p   = grp.iloc[i].get("qfq_low",   0)
                close_p = grp.iloc[i].get("qfq_close", 0)
                if high_p > low_p and (close_p - low_p) / (high_p - low_p) <= 0.2:
                    pre_c        = grp.iloc[i + 1].get("qfq_close", np.nan)
                    recover_line = (pre_c + close_p) / 2.0
                    next_close   = grp.iloc[i-1].get("qfq_close", 0)
                    if next_close < recover_line:
                        has_crash_veto = True
                        break

        high_20d    = grp.head(20)["qfq_high"].max()
        low_20d     = grp.head(20)["qfq_low"].min()
        drawdown_20d = float((closes.iloc[0] / high_20d - 1) * 100) if (
            not pd.isna(high_20d) and high_20d > 0 and not pd.isna(closes.iloc[0])
        ) else np.nan

        rows.append({
            "ts_code":        code,
            "qfq_close":      float(grp.iloc[0]["qfq_close"]),
            "qfq_source_trade_date": str(grp.iloc[0]["trade_date"]),
            "pct_20d":        pct_20d,
            "pct_5d":         pct_5d,
            "pct_60d":        pct_60d,
            "avg_amount_20d": avg_20d,
            "avg_amount_5d":  avg_5d,
            "vol_confirm":    vol_confirm,
            "limit_20d":      limit_20d,
            "limit_10d":      limit_10d,
            "is_lock":        is_lock,
            "is_breakout":    is_breakout,
            "limit_breakout_legacy": limit_breakout_legacy,   # 审计:旧涨停口径(诊断,不进 analysis_input)
            "high_20d":       high_20d,
            "low_20d":        low_20d,
            "drawdown_20d":   drawdown_20d,
            "has_crash_veto": has_crash_veto,
        })
    if not rows:
        raise RuntimeError(
            "daily stats coverage too low: no valid close rows for stock universe; "
            "abort to avoid neutral pass-through stats"
        )
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════
# §4 L0 过滤
# ═══════════════════════════════════════════════════
def filter_l0(df_stocks, stats_df, unlock_set, red_dict, suspended_set, relisted_set):
    missing_columns = sorted({"ts_code", "name"} - set(df_stocks.columns))
    if missing_columns:
        raise RuntimeError("L0 input missing delisting-safety columns: " + ",".join(missing_columns))
    df = df_stocks.copy()
    n0 = len(df)

    # 主板 only(用户硬口径,**strict INCLUSION**):只保留 A 股主板规范码(SSE 600/601/603/605 + SZSE 000/001/002/003,
    # 6 位 ASCII 数字)。用 is_a_share_main_board 一次性排除 创业板(300/301)/科创板(688/689)/北交所(920/.BJ)/
    # **B股(沪900·深200)**/畸形码——比逐个加排除前缀(原 `~startswith(300/301/688/689/920/900/200)`)更彻底:
    # 畸形码(如 600ABC.SH)、未来新非主板前缀都拒,_board_from_code 之后只会见到主板码。
    df = df[df["ts_code"].map(is_a_share_main_board)].copy()
    delisting = df.apply(
        lambda row: derive_delisting_flags(row, historical=_historical_replay_mode()),
        axis=1,
        result_type="expand",
    )
    # Unknown status is never a safe pass-through.  Historical rows use only
    # the already PIT-resolved name; live rows additionally require list_status.
    safe_status = (
        delisting["known"].astype(bool)
        & ~delisting["st_flag"].fillna(True).astype(bool)
        & ~delisting["delisting_warning"].fillna(True).astype(bool)
    )
    df = df[safe_status].copy()
    df = df[~df["name"].fillna("").astype(str).str.contains("暂停上市", regex=False)].copy()
    if not _historical_replay_mode() and "list_status" in df.columns:
        df = df[df["list_status"].fillna("").astype(str).str.upper().ne("P")].copy()

    if suspended_set: df = df[~df["ts_code"].isin(suspended_set)].copy()
    if relisted_set:  df = df[~df["ts_code"].isin(relisted_set)].copy()
    if unlock_set:    df = df[~df["ts_code"].isin(unlock_set)].copy()

    veto = red_dict.get("veto_10d", set())
    if veto: df = df[~df["ts_code"].isin(veto)].copy()

    if not stats_df.empty and "avg_amount_20d" in stats_df.columns:
        df = df.merge(stats_df[["ts_code","avg_amount_20d"]], on="ts_code", how="left")
        df["avg_amount_20d"] = df["avg_amount_20d"].fillna(0)
        if df["avg_amount_20d"].gt(0).any():
            df = df[df["avg_amount_20d"] >= CONF["min_avg_amount"]].copy()
        else:
            log.warning("avg_amount_20d 全为 NaN/0（日线数据缺失），跳过成交额过滤")

    if "pct_20d" in stats_df.columns and stats_df["pct_20d"].notna().any():
        df = df.merge(stats_df[["ts_code","pct_20d"]], on="ts_code", how="left")
        df = df[df["pct_20d"].notna()].copy()
    elif "pct_20d" in stats_df.columns:
        log.warning("pct_20d 全为 NaN（数据应急兜底），跳过涨跌幅过滤")
        df = df.merge(stats_df[["ts_code","pct_20d"]], on="ts_code", how="left")

    log.info(f"L0 过滤：{n0} -> {len(df)}")
    return df.reset_index(drop=True)


# ═══════════════════════════════════════════════════
# §5 构建主表
# ═══════════════════════════════════════════════════
def build_master(df_l0, stats_df, df_db, df_fin, sw_map, red_dict):
    df = df_l0.copy()
    df["l2_name"] = df["ts_code"].map(lambda c: sw_map.get(c, {}).get("l2_name", "未知"))
    df["l1_name"] = df["ts_code"].map(lambda c: sw_map.get(c, {}).get("l1_name", "未知"))
    df["l2_code"] = df["ts_code"].map(lambda c: sw_map.get(c, {}).get("l2_code", ""))
    df["l1_code"] = df["ts_code"].map(lambda c: sw_map.get(c, {}).get("l1_code", ""))

    if not df_db.empty:
        cols = [c for c in ["ts_code","close","pe","pe_ttm","pb","roe",
                             "turnover_rate","total_mv","circ_mv","source_trade_date"] if c in df_db.columns]
        daily_basic = df_db[cols].copy().rename(columns={"close": "raw_close"})
        df = df.merge(daily_basic, on="ts_code", how="left")

    if not stats_df.empty:
        st_cols = [c for c in ["ts_code","avg_amount_5d","vol_confirm","limit_20d","limit_10d",
                                "is_lock","is_breakout","limit_breakout_legacy","high_20d","low_20d","pct_5d","pct_60d",
                                "drawdown_20d","has_crash_veto","qfq_close","qfq_source_trade_date"]
                   if c in stats_df.columns]
        df = df.merge(stats_df[st_cols], on="ts_code", how="left")

    required_price_binding = {"raw_close", "source_trade_date", "qfq_close", "qfq_source_trade_date"}
    missing_price_binding = required_price_binding - set(df.columns)
    if missing_price_binding:
        raise RuntimeError(
            f"master table lacks raw/qfq price binding fields: {sorted(missing_price_binding)}"
        )
    if df[list(required_price_binding)].isna().any().any():
        raise RuntimeError("master table lacks a same-day raw/qfq candidate price")
    if (df["source_trade_date"].astype(str) != df["qfq_source_trade_date"].astype(str)).any():
        raise RuntimeError("daily_basic raw quote date does not match qfq candidate price date")
    df["close"] = pd.to_numeric(df["qfq_close"], errors="coerce")
    if not np.isfinite(df["close"].to_numpy(dtype=float)).all() or (df["close"] <= 0).any():
        raise RuntimeError("master table qfq candidate price is non-finite or non-positive")

    for bool_col in ["is_lock","is_breakout","limit_breakout_legacy","has_crash_veto"]:
        if bool_col in df.columns:
            df[bool_col] = df[bool_col].fillna(False)

    df["pct_20d_n"] = pd.to_numeric(df["pct_20d"], errors="coerce")
    med = df["pct_20d_n"].median()
    df["market_med_20d"] = 0.0 if pd.isna(med) else med

    if not df_fin.empty:
        fin_cols = [c for c in ["ts_code","q0_dt_yoy","q1_dt_yoy","q0_profit_dedt",
                                 "q0_dt_profit_ratio","q0_net_income",
                                 "ttm_net_income","ttm_profit_dedt","ttm_ocf_ratio","roe",
                                 "q0_end_date","q1_end_date","q0_revenue_yoy","q1_revenue_yoy"]
                    if c in df_fin.columns]
        df = df.merge(df_fin[fin_cols], on="ts_code", how="left", suffixes=("","_fin"))
        if "roe_fin" in df.columns:
            df["roe"] = df["roe"].fillna(df["roe_fin"])
            df.drop(columns=["roe_fin"], inplace=True, errors="ignore")

    df["reduce_deduct"] = df["ts_code"].isin(red_dict.get("deduct_30d", set())).astype(int)
    return df


# ═══════════════════════════════════════════════════
# §6 L1~L5 评分
# ═══════════════════════════════════════════════════
def score_l1(df, csi300_ret, exclusion_reasons=None):
    df = df.copy()
    if exclusion_reasons is None:
        exclusion_reasons = {}
    df["pct_20d_n"]    = pd.to_numeric(df["pct_20d"], errors="coerce")
    df["total_mv_n"]   = pd.to_numeric(df.get("total_mv",    pd.Series(dtype=float)), errors="coerce")
    df["avg_amt_5d_n"] = pd.to_numeric(df.get("avg_amount_5d", pd.Series(dtype=float)), errors="coerce")

    mkt_med   = df["pct_20d_n"].median()
    benchmark = (csi300_ret + 2.0) if csi300_ret is not None else (0.0 if pd.isna(mkt_med) else mkt_med + 2.0)

    def leader_score(g):
        cap3 = g.nlargest(min(3, len(g)), "total_mv_n")
        ret2 = g.nlargest(min(2, len(g)), "pct_20d_n")
        cand = pd.concat([cap3, ret2]).drop_duplicates("ts_code") if "ts_code" in g.columns else cap3
        qual = cand[cand["avg_amt_5d_n"] > 1e8]
        if qual.empty: return 0.5
        avg_ret = qual["pct_20d_n"].mean()
        return 1.0 if (not pd.isna(avg_ret) and avg_ret >= benchmark) else 0.0

    try:
        # pandas >= 2.2：需要显式传入 include_groups=False，否则将来会崩溃
        l1_ldr = df.groupby("l2_name").apply(leader_score, include_groups=False)
    except TypeError:
        # pandas < 2.2：不支持该参数，回退旧写法
        l1_ldr = df.groupby("l2_name").apply(leader_score)
    l1_ldr = l1_ldr.rename("l1_leader").reset_index()
    df = df.merge(l1_ldr, on="l2_name", how="left")
    df["l1_leader"] = df["l1_leader"].fillna(0.5)
    df["l1_score"]  = 0.5 + df["l1_leader"]

    df["l1_flag"] = ""
    df.loc[df["l1_score"] >= 1.5, "l1_flag"] = "PASS"
    df.loc[df["l1_score"] == 1.0, "l1_flag"] = "ITF-2"
    df.loc[df["l1_score"] <  1.0, "l1_flag"] = "ELIM"
    l1_elim = df["l1_flag"] == "ELIM"
    for _, row in df.loc[l1_elim, ["ts_code", "l2_name"]].iterrows():
        exclusion_reasons[str(row["ts_code"])] = (
            "l1_unknown_industry_elim"
            if str(row["l2_name"]).strip() in {"", "未知", "nan", "None"}
            else "l1_industry_leader_elim"
        )
    df = df[df["l1_flag"] != "ELIM"].copy()

    # ── [逻辑修复③] ITF-ADJ：直接用原始 "pe" 列，pe_n 在 score_l2 才创建 ──
    df["itf_adj"] = False
    if "pe" in df.columns:
        pe_temp = pd.to_numeric(df["pe"], errors="coerce")
        if pe_temp.notna().sum() > 0:
            df["_pe_tmp"]       = pe_temp
            ind_pe_median       = df.groupby("l2_name")["_pe_tmp"].median()
            ind_pe_rank         = ind_pe_median.rank(pct=True)
            df["l2_pe_mkt_pct"] = df["l2_name"].map(ind_pe_rank)
            df.loc[
                (df["l1_flag"] == "ITF-2") &
                (df["l2_pe_mkt_pct"] < 0.50) &
                (df["pct_20d_n"] > -15),
                "itf_adj"
            ] = True
            df.drop(columns=["_pe_tmp"], inplace=True, errors="ignore")
    # ─────────────────────────────────────────────────────────────────────────

    log.info(f"L1 筛选后剩余 {len(df)} 只")
    return df.reset_index(drop=True)


def score_l2(df, mg_df, trade_dates, global_ind_med, exclusion_reasons=None):
    df = df.copy()
    if exclusion_reasons is None:
        exclusion_reasons = {}

    if "has_crash_veto" in df.columns:
        n_before = len(df)
        crash_mask = df["has_crash_veto"] == True
        for ts_code in df.loc[crash_mask, "ts_code"].astype(str):
            exclusion_reasons[ts_code] = "l2_crash_veto"
        df = df[df["has_crash_veto"] == False].copy()
        log.info(f"异常大跌一票否决：剔除 {n_before - len(df)} 只")

    df["q0_dt_yoy_n"] = pd.to_numeric(df.get("q0_dt_yoy", pd.Series(dtype=float)), errors="coerce")
    df["q1_dt_yoy_n"] = pd.to_numeric(df.get("q1_dt_yoy", pd.Series(dtype=float)), errors="coerce")
    df["ind_med"]     = df["l2_name"].map(global_ind_med)

    def _surp(g, med):
        if pd.isna(g) or g is None: return None, "DATA-INC"
        if med is None or pd.isna(med): return (g > 0), ("DATA-INC" if g <= 0 else "")
        if med > 0:     return ((g / med) > 1.5), ""
        elif med >= -3: return (g > 0 and g > 5), ""
        else:           return (g > 0 and g > abs(med)), ""

    def calc_esp(row):
        lg  = row["q0_dt_yoy_n"]
        pg  = row["q1_dt_yoy_n"]
        med = row["ind_med"]
        flags = []

        if pd.isna(lg):
            flags.append("DATA-INC")
            return 0.0, flags

        if pd.isnull(med):
            flags.append("COV-LOW")
            return (lg * 0.5 if lg > 0 else 0.0), flags

        s0, tag0 = _surp(lg, med)
        s1, _    = _surp(pg, med)

        if tag0 == "DATA-INC": flags.append("DATA-INC")

        if s0 and s1:
            raw = abs(med) + (lg - med)
        elif s0:
            flags.append("ESP-Q")
            raw = abs(med) + (lg - med)
        elif med <= 0:
            # 行业负增长区间：允许负值，区分"在下行行业中的相对强弱"
            raw = lg * 0.5
        else:
            # 未超越行业中位数：改为允许负值（lg - med < 0）
            # 原 max(0.0, lg-med) 导致 71% 股票 raw=0，Z-Score 输入分布崩塌
            # 修改后 Z-Score 双侧连续分布，esp_score 从81%集中在35分扩散到全区间
            raw = lg - med

        # 非经常性损益检查：dtprofit_to_profit = 扣非/净利润（Tushare 以百分比存储）
        dt_ratio = pd.to_numeric(row.get("q0_dt_profit_ratio", np.nan), errors="coerce")
        if not pd.isna(dt_ratio):
            dt_ratio_norm = dt_ratio / 100.0 if abs(dt_ratio) > 2 else dt_ratio
            if dt_ratio_norm < 0.80 and "ESP-Q" not in flags:
                flags.append("ESP-Q")

        # OCF 质量检查：用 ttm_profit_dedt 作为规模代理（替代不可用的 ttm_net_income）
        ttm_dt  = pd.to_numeric(row.get("ttm_profit_dedt", np.nan), errors="coerce")
        ttm_ocf = pd.to_numeric(row.get("ttm_ocf_ratio",   np.nan), errors="coerce")
        if not pd.isna(ttm_ocf) and not pd.isna(ttm_dt):
            threshold = 0 if abs(ttm_dt) <= 100 else 0.7
            if ttm_ocf < threshold and "ESP-Q" not in flags:
                flags.append("ESP-Q")

        return raw, flags

    results        = df.apply(calc_esp, axis=1)
    df["esp_raw"]  = [r[0] for r in results]
    df["l2_flags"] = ["|".join(r[1]) if r[1] else "" for r in results]

    for col in ["pe","pb","roe"]:
        df[f"{col}_n"] = pd.to_numeric(df.get(col, pd.Series(dtype=float)), errors="coerce")

    df["peg_n"] = np.nan
    mask_peg = df["pe_n"].notna() & df["q0_dt_yoy_n"].notna() & (df["q0_dt_yoy_n"] > 0)
    df.loc[mask_peg, "peg_n"] = df["pe_n"] / df["q0_dt_yoy_n"]

    ind_pe_mean = df.groupby("l2_name")["pe_n"].transform("mean")
    ind_roe_p70 = df.groupby("l2_name")["roe_n"].transform(lambda x: x.quantile(0.70))
    ind_pe_p30  = df.groupby("l2_name")["pe_n"].transform(lambda x: x.quantile(0.30))
    df["pe_mean"] = ind_pe_mean

    df["val_bonus"]   = 0.0
    df["val_penalty"] = 0.0
    df.loc[(df["pe_n"] < ind_pe_p30) & (df["roe_n"] > ind_roe_p70), "val_bonus"] += 10
    df.loc[(df["pb_n"] < 1.0) & (df["roe_n"] > 10), "val_bonus"] += 10
    df["val_bonus"] = df["val_bonus"].clip(upper=20)

    mask_pe2    = df["pe_n"] > ind_pe_mean * 2
    mask_pe3    = df["pe_n"] > ind_pe_mean * 3
    mask_peg_ok = df["peg_n"].notna()
    df.loc[mask_pe2 & mask_peg_ok & (df["peg_n"] > 1.5), "val_penalty"] = 10
    df.loc[mask_pe3 & mask_peg_ok & (df["peg_n"] > 2),   "val_penalty"] = 15
    df.loc[mask_pe3 & mask_peg_ok & (df["peg_n"] > 2),   "l2_flags"]   += "|VAL-X"

    if "high_20d" in df.columns and "close" in df.columns:
        df["close_n"]    = pd.to_numeric(df.get("close",   pd.Series(dtype=float)), errors="coerce")
        df["high_20d_n"] = pd.to_numeric(df["high_20d"], errors="coerce")
        mask_high = (df["close_n"] >= df["high_20d_n"] * 0.95) & mask_pe2
        df.loc[mask_high, "val_penalty"] = df.loc[mask_high, "val_penalty"].clip(lower=10)

    mask_espq   = df["l2_flags"].str.contains("ESP-Q", na=False)
    mask_triple = mask_espq & mask_pe3 & mask_peg_ok & (df["peg_n"] > 2)
    for ts_code in df.loc[mask_triple, "ts_code"].astype(str):
        exclusion_reasons[ts_code] = "l2_espq_valuation_veto"
    df = df[~mask_triple].copy()

    if not mg_df.empty and "rzye" in mg_df.columns and len(trade_dates) >= 10:
        d_latest = trade_dates[0]
        d_oldest = trade_dates[9]
        mg_l = mg_df[mg_df["trade_date"]==d_latest][["ts_code","rzye"]].rename(columns={"rzye":"rzye_l"})
        mg_o = mg_df[mg_df["trade_date"]==d_oldest][["ts_code","rzye"]].rename(columns={"rzye":"rzye_o"})
        if not mg_l.empty and not mg_o.empty:
            mg_chg = mg_l.merge(mg_o, on="ts_code")
            mg_chg["chg"] = (mg_chg["rzye_l"] - mg_chg["rzye_o"]) / mg_chg["rzye_o"].abs().clip(lower=1)
            mg_chg = mg_chg[mg_chg["chg"] > 0.15]

            exempt_vol = set()
            if "avg_amount_5d" in df.columns and "avg_amount_20d" in df.columns:
                df["vol_shrink"] = df["avg_amount_5d"] < df["avg_amount_20d"] * 0.8
                exempt_vol = set(df[df["vol_shrink"]]["ts_code"])

            exempt_ret  = set(df[df["pct_20d_n"] >= 3]["ts_code"])
            veto_margin = set(mg_chg["ts_code"]) - exempt_ret - exempt_vol
            for ts_code in veto_margin & set(df["ts_code"].astype(str)):
                exclusion_reasons[str(ts_code)] = "l2_margin_growth_veto"
            df = df[~df["ts_code"].isin(veto_margin)].copy()

    df["reduce_penalty"] = df["reduce_deduct"].fillna(0) * 10
    return df


def score_l3(df, trade_dates, all_daily):
    df = df.copy()
    df["cat_score"] = 50.0
    df["cat_flag"]  = ""
    CONF["l3_snapshot_date"] = None  # M2: reset every call; pit branch will set if applicable
    CONF["l3_provider"] = None
    CONF["l3_coverage"] = None

    l3_mode = CONF.get("l3_mode", "today")
    if l3_mode == "neutralize":
        CONF["l3_provider"] = "neutralized"
        log.info("L3 mode=neutralize: cat_score=50.0 for all candidates, skipping L3 API calls")
        return df

    if all_daily.empty:
        if l3_mode == "today":
            raise SystemExit(
                "[FATAL] L3 requires usable market daily data to calculate concept intensity; "
                "no selection will be published."
            )
        df["cat_flag"] = "COV-LOW"
        return df

    # 取近5日日线数据
    recent_dates = set(trade_dates[:5])
    ad5 = all_daily[all_daily["trade_date"].isin(recent_dates)].copy()
    for col in ["pct_chg", "amount"]:
        if col in ad5.columns:
            ad5[col] = pd.to_numeric(ad5[col], errors="coerce")
    ad5 = ad5[["ts_code", "trade_date", "pct_chg", "amount"]].dropna(subset=["pct_chg", "amount"])

    if ad5.empty:
        if l3_mode == "today":
            raise SystemExit(
                "[FATAL] L3 has no usable recent daily rows to calculate concept intensity; "
                "no selection will be published."
            )
        df["cat_flag"] = "COV-LOW"
        return df

    if l3_mode == "pit":
        snapshot = _load_l3_snapshot(TODAY, include_metadata=True)
        if snapshot is None:
            if CONF.get("l3_pit_strict", False):
                raise SystemExit(
                    f"[FATAL] L3 mode=pit requires a snapshot dated <= {TODAY} in "
                    f"{L3_SNAPSHOT_DIR}; none found. Seed snapshots by running with "
                    "--l3-mode=today on or before this date, or use --l3-mode=neutralize."
                )
            log.warning(f"L3 mode=pit: no snapshot <= {TODAY}, falling back to cat_score=50.0")
            return df
        concepts_df, stock_concepts, concept_members, snap_date, l3_source, coverage = snapshot
        CONF["l3_snapshot_date"] = snap_date
        CONF["l3_provider"] = l3_source or "legacy_tushare_snapshot"
        CONF["l3_coverage"] = coverage
        gap_days = (datetime.strptime(TODAY, "%Y%m%d") - datetime.strptime(snap_date, "%Y%m%d")).days
        if gap_days > 14:
            log.warning(
                f"L3 mode=pit: using snapshot from {snap_date} ({gap_days}d behind as_of {TODAY}; >14d)"
            )
        else:
            log.info(f"L3 mode=pit: using snapshot from {snap_date} (gap={gap_days}d)")
    elif l3_mode == "today":
        real_today = datetime.now().strftime("%Y%m%d")
        prior_snapshot = _load_l3_snapshot(real_today, include_metadata=True)
        reuse_l3_cache = CONF.get("l3_cache_mode") == "reuse"

        if reuse_l3_cache:
            if prior_snapshot is None:
                raise SystemExit(
                    "[FATAL] --reuse-l3-cache requires an existing complete HiThink L3 snapshot; "
                    "no provider call was made."
                )
            concepts_df, stock_concepts, concept_members, snap_date, l3_source, coverage = prior_snapshot
            if l3_source != HITHINK_L3_SOURCE_ID or not _is_complete_hithink_snapshot(
                concepts_df, concept_members, coverage
            ):
                raise SystemExit(
                    "[FATAL] --reuse-l3-cache found no reusable complete HiThink main-board snapshot; "
                    "no provider call was made."
                )
            reuse_gap_days = (datetime.strptime(real_today, "%Y%m%d") -
                              datetime.strptime(snap_date, "%Y%m%d")).days
            if reuse_gap_days > 14:
                if not CONF.get("l3_allow_stale_cache", False):
                    raise SystemExit(
                        f"[FATAL] --reuse-l3-cache snapshot {snap_date} is {reuse_gap_days}d behind "
                        f"run date {real_today} (>14d); refresh L3 or use --allow-stale-l3-cache for testing only. "
                        "No provider call was made."
                    )
                log.warning(
                    f"L3 testing cache reused stale snapshot {snap_date} "
                    f"({reuse_gap_days}d behind run date {real_today}; explicit test-only override)"
                )
            CONF["l3_snapshot_date"] = snap_date
            CONF["l3_provider"] = HITHINK_L3_SOURCE_ID
            CONF["l3_coverage"] = dict(coverage)
            log.info(
                f"L3 testing cache reused: snapshot={snap_date}, "
                f"catalog={coverage['catalog_board_count']}, provider call skipped"
            )
        else:
            expected_catalog_codes = None
            if prior_snapshot is not None and prior_snapshot[4] == HITHINK_L3_SOURCE_ID:
                expected_catalog_codes = _catalog_codes_from_snapshot(prior_snapshot[0]) or None
            try:
                l3_graph = fetch_complete_concept_graph(
                    expected_catalog_codes=expected_catalog_codes,
                )
            except HiThinkL3SourceError as exc:
                raise SystemExit(
                    "[FATAL] HiThink L3 concept catalog is incomplete or unavailable; "
                    f"no selection will be published. {exc}"
                ) from None

            concepts_df = l3_graph.concepts_df
            concept_members = l3_graph.concept_members
            stock_concepts = l3_graph.stock_concepts
            coverage = l3_graph.coverage
            CONF["l3_provider"] = HITHINK_L3_SOURCE_ID
            CONF["l3_coverage"] = dict(coverage)

            # Snapshot persistence is part of the source receipt. A complete live
            # graph that cannot be durably recorded is not publishable.
            try:
                if TODAY != real_today:
                    log.info(
                        f"L3 snapshot tagged with real-world date {real_today} "
                        f"(not as_of {TODAY}); future pit reads find it by snap_date"
                    )
                _write_l3_snapshot(
                    real_today,
                    concepts_df,
                    stock_concepts,
                    concept_members,
                    l3_source=HITHINK_L3_SOURCE_ID,
                    coverage=coverage,
                )
            except Exception as exc:
                raise SystemExit(
                    "[FATAL] Complete HiThink L3 graph could not be snapshotted; "
                    f"no selection will be published ({type(exc).__name__})."
                ) from None
            CONF["l3_snapshot_date"] = real_today
            log.info(
                f"L3 complete graph saved -> state/l3_snapshots/{L3_SNAPSHOT_PREFIX}{real_today}{L3_SNAPSHOT_SUFFIX} "
                f"(boards={coverage['received_board_count']}/{coverage['catalog_board_count']}, "
                f"verified_empty={coverage['verified_empty_board_count']}, "
                f"stocks={len(stock_concepts)})"
            )
    else:
        raise ValueError(f"unsupported L3 mode {l3_mode!r}")
    # Re-validate concepts_df (defensive for both pit and today branches).
    if concepts_df is None or concepts_df.empty:
        df["cat_flag"] = "COV-LOW"
        return df
    all_concept_ids = set(concepts_df["code"].dropna().tolist())
    if not all_concept_ids:
        df["cat_flag"] = "COV-LOW"
        return df

    # Step 4: 合成虚拟概念指数，用成交额加权涨跌幅作为5日资金流入强度代理
    concept_intensity = {}
    for cid, member_codes in concept_members.items():
        sub = ad5[ad5["ts_code"].isin(member_codes)]
        if sub.empty:
            continue
        total_amt = sub["amount"].sum()
        if total_amt <= 0:
            continue
        concept_intensity[cid] = float((sub["pct_chg"] * sub["amount"]).sum() / total_amt)

    if not concept_intensity:
        if l3_mode == "today":
            raise SystemExit(
                "[FATAL] Complete HiThink L3 membership produced no usable concept intensity; "
                "no selection will be published."
            )
        log.warning("L3: 概念强度全部计算失败，cat_score 默认 50")
        df["cat_flag"] = "COV-LOW"
        return df

    # Step 5: 对所有概念强度做百分位排名（0~100）
    intensity_s = pd.Series(concept_intensity)
    pct_rank    = intensity_s.rank(pct=True) * 100  # index=concept_id, value=0~100

    # 取候选股所属概念中百分位最高的值作为 cat_score，并打标
    for idx, row in df.iterrows():
        code  = row["ts_code"]
        cids  = stock_concepts.get(code, [])
        ranks = [pct_rank[c] for c in cids if c in pct_rank.index]

        if not ranks:
            df.at[idx, "cat_flag"] = "COV-LOW"
            continue

        cat_score = float(max(ranks))
        df.at[idx, "cat_score"] = cat_score

        # Step 6: 打标
        pct_20d = pd.to_numeric(row.get("pct_20d_n", np.nan), errors="coerce")
        if cat_score < 40:
            df.at[idx, "cat_flag"] = "CAT-0"
        elif cat_score >= 70 and not pd.isna(pct_20d) and float(pct_20d) > 15:
            df.at[idx, "cat_flag"] = "CHASE"

    log.info(
        f"L3 完成：{(df['cat_flag']=='').sum()} 正常 / "
        f"{(df['cat_flag']=='CAT-0').sum()} CAT-0 / "
        f"{(df['cat_flag']=='CHASE').sum()} CHASE / "
        f"{(df['cat_flag']=='COV-LOW').sum()} COV-LOW"
    )
    return df


def score_l4(df, mf_df):
    df = df.copy()

    df["mom_rank"]   = df.groupby("l2_name")["pct_20d_n"].rank(pct=True)
    df["l4_mom_ok"]  = (df["mom_rank"] >= 0.70).astype(int)
    mkt_med          = df["pct_20d_n"].median()
    df["l4_rel_ok"]  = (df["pct_20d_n"] > mkt_med + 1).astype(int)
    df["l4_score"]   = (df["l4_mom_ok"] + df["l4_rel_ok"]) * 50.0

    if "vol_confirm" in df.columns:
        df.loc[df["vol_confirm"] == False, "l4_score"] *= 0.7
    df["l4_score"] = df["l4_score"].clip(0, 100)

    if "l4_mom_ok" in df.columns:
        # 每个二级行业中满足动量条件的股票数，map 回全量 df 避免 pandas 2.x 索引对齐报错
        # 同时正确地对行业内所有股票（而非仅 l4_mom_ok==1 的股票）发放加分
        ind_mom_map = df[df["l4_mom_ok"]==1].groupby("l2_name")["ts_code"].count()
        df["ind_mom_cnt"] = df["l2_name"].map(ind_mom_map).fillna(0)
        df.loc[df["ind_mom_cnt"] >= 3, "l4_score"] += 5
        df["l4_score"] = df["l4_score"].clip(0, 100)

    ind_ret_med       = df.groupby("l2_name")["pct_20d_n"].transform("median")
    df["alpha_excess"] = df["pct_20d_n"] - ind_ret_med
    df["alpha_flag"]   = df["alpha_excess"] > 5.0

    if "pct_5d" in df.columns:
        df["pct_5d_n"] = pd.to_numeric(df["pct_5d"], errors="coerce")
        ind_5d_ret     = df.groupby("l2_name")["pct_5d_n"].transform("median")
    else:
        ind_5d_ret = ind_ret_med / 4.0

    df.loc[df["alpha_flag"] & (ind_5d_ret > -2), "l4_score"] += 10  # v7.3: ALPHA+行业不跌→+10
    df["l4_score"] = df["l4_score"].clip(0, 100)

    def limit_flag(row):
        tags = []
        if row.get("is_lock",     False): tags.append("LOCK")
        if row.get("is_breakout", False): tags.append("突破型")
        return "|".join(tags)

    df["l4_flag"] = df.apply(limit_flag, axis=1)

    if not mf_df.empty:
        for col in ["buy_elg_amount","sell_elg_amount","buy_lg_amount","sell_lg_amount","trade_amount"]:
            if col in mf_df.columns:
                mf_df[col] = pd.to_numeric(mf_df[col], errors="coerce").fillna(0)
        needed = ["buy_elg_amount","sell_elg_amount","buy_lg_amount","sell_lg_amount","trade_amount"]
        if all(c in mf_df.columns for c in needed):
            mf_agg = mf_df.groupby("ts_code").agg(
                elg_buy  = ("buy_elg_amount","sum"),
                elg_sell = ("sell_elg_amount","sum"),
                lg_buy   = ("buy_lg_amount","sum"),
                lg_sell  = ("sell_lg_amount","sum"),
                total    = ("trade_amount","sum"),
            ).reset_index()
            mf_agg["big_net"]   = (mf_agg["elg_buy"] - mf_agg["elg_sell"]) + (mf_agg["lg_buy"] - mf_agg["lg_sell"])
            mf_agg["big_ratio"] = mf_agg["big_net"] / mf_agg["total"].clip(lower=1)
            df = df.merge(mf_agg[["ts_code","big_ratio"]], on="ts_code", how="left")
            df.loc[df["big_ratio"] > 0.15, "l4_score"] += 5
            df["l4_score"] = df["l4_score"].clip(0, 100)
    else:
        log.warning("资金流数据为空，大单流向加分跳过")

    # 所有加分完成后再判定 TIER2_FORCED，避免资金流加分后仍被冤枉降级
    df["l4_flag"] = df.apply(
        lambda r: ("" if not r["l4_flag"] else r["l4_flag"] + "|") + "TIER2_FORCED"
        if r["l4_score"] == 0 else r["l4_flag"], axis=1
    )

    df["chasing_high"] = df["pct_20d_n"] > 15

    # 周频过热标记：5日涨幅过快 或 20日大涨且仍在高位（drawdown < 3%）
    if "pct_5d_n" not in df.columns:
        df["pct_5d_n"] = pd.to_numeric(df.get("pct_5d", pd.Series(dtype=float)), errors="coerce")
    dd20 = pd.to_numeric(df.get("drawdown_20d", pd.Series(np.nan, index=df.index)), errors="coerce")
    oh_5d  = df["pct_5d_n"]  > CONF["overheat_5d"]
    oh_20d = (df["pct_20d_n"] > CONF["overheat_20d"]) & (dd20 > -3)
    df["overheat_flag"] = oh_5d | oh_20d
    df["overheat_flag"] = df["overheat_flag"].fillna(False)
    df.loc[df["overheat_flag"], "l4_flag"] = df.loc[df["overheat_flag"], "l4_flag"].apply(
        lambda x: (x + "|" if x else "") + "OVERHEAT"
    )
    return df


def score_l5(df, sw_map):
    df = df.copy()

    global_l2 = pd.Series([v["l2_name"] for v in sw_map.values() if v["l2_name"] != "未知"])
    global_l1 = pd.Series([v["l1_name"] for v in sw_map.values() if v["l1_name"] != "未知"])
    g_l2_cnt  = global_l2.value_counts()
    g_l1_cnt  = global_l1.value_counts()

    def get_z_group(row):
        l2n = g_l2_cnt.get(row["l2_name"], 0)
        if l2n >= 10: return row["l2_name"]
        elif l2n >= 5: return "全市场"
        else:
            l1n = g_l1_cnt.get(row["l1_name"], 0)
            return row["l1_name"] if l1n >= 5 else "独立池"

    df["z_group"] = df.apply(get_z_group, axis=1)

    valid_esp = df["esp_raw"].dropna()
    if not valid_esp.empty:
        p1, p99 = valid_esp.quantile(0.01), valid_esp.quantile(0.99)
        upper = min(float(p99), float(CONF["esp_raw_cap"]))
        df["esp_raw_w"] = df["esp_raw"].clip(p1, upper)
    else:
        df["esp_raw_w"] = df["esp_raw"].copy()
    df["esp_raw_w"] = df["esp_raw_w"].fillna(0.0)
    df["low_base_growth_flag"] = pd.to_numeric(df["esp_raw"], errors="coerce") > float(CONF["esp_raw_cap"])

    cov_mask = df["l2_flags"].str.contains("COV-LOW", na=False)
    df_cov   = df[cov_mask].copy()
    df_main  = df[~cov_mask].copy()

    def _calc_z(sub):
        sub["esp_z"] = 0.0
        for grp_name, idx in sub.groupby("z_group").groups.items():
            if grp_name == "独立池": continue
            vals = sub.loc[idx, "esp_raw_w"]
            if len(vals) < 2: continue
            mu, sig = vals.mean(), vals.std()
            if pd.isna(sig) or sig < 1e-9: continue
            sub.loc[idx, "esp_z"] = (sub.loc[idx, "esp_raw_w"] - mu) / sig
        return sub

    df_main = _calc_z(df_main)
    if not df_cov.empty:
        df_cov = _calc_z(df_cov)
    else:
        df_cov["esp_z"] = 0.0

    df = pd.concat([df_main, df_cov]).loc[df.index]

    indep_mask      = df["z_group"] == "独立池"
    df["esp_score"] = 0.0

    if indep_mask.any():
        x   = df.loc[indep_mask, "esp_raw_w"]
        rng = x.max() - x.min()
        df.loc[indep_mask, "esp_score"] = ((x - x.min()) / rng * 50) if rng > 0 else 25.0

    def z2s(z):
        """v7.3 官方映射表：负值四档梯度，正值宽桶设计"""
        if pd.isna(z): return 50.0
        if z < -2.5: return 5
        if z < -1.5: return 12
        if z < -0.5: return 22
        if z <  0:   return 32
        if z <  1:   return 50
        if z <  2:   return 68
        if z <  3:   return 82
        return min(100, 95 + (z - 3) * 3)

    non_indep = ~indep_mask
    df.loc[non_indep, "esp_score"] = df.loc[non_indep, "esp_z"].apply(z2s)
    df.loc[cov_mask & non_indep, "esp_score"] = df.loc[cov_mask & non_indep, "esp_score"].clip(upper=50)

    # 行业热度(SW L2)+ 可治理权重 profile —— 打分尾段(egs_base→mult→deduct→final_score→tier→准入降级)
    # 抽到 engine/egs_industry_heat.py(单一真相源:egs_main 与每周非生产对比 diff(同模块 write_weight_comparison)共用,杜绝漂移)。
    # 生产 active_profile=balanced(esp.20/cat.25/l4.40/ind.15)= 已生效:行业/赛道权重已提高、选股已改变。
    # legacy(esp.20/cat.30/l4.50/ind 0 = 改前原式)仅作一键回滚锚 + 回归基准(翻回 legacy 即还原)。
    # 行业热度只加分排序,绝不救回 hard_veto/停牌/涨停锁/ST/减持/闪崩;chasing_high·overheat·未知行业 降级原样保留。
    df["industry_heat_score"] = compute_industry_heat_score(df)
    _ih_profile, _ih_weights = get_active_weights()
    df, _score_info = final_score_and_tier(df, _ih_weights)
    fin_coverage = _score_info["fin_coverage"]
    esp_neg_mask = df["esp_raw"].fillna(0) <= 0          # 供下游 downgrade_reasons 标注复用
    if fin_coverage >= 0.70:
        log.info(f"Tier1 准入硬条件激活（财务覆盖率{fin_coverage:.0%}≥70%）：{_score_info['esp_neg_demoted']} 只降级")
    else:
        log.info(f"Tier1 准入硬条件暂停（财务覆盖率{fin_coverage:.0%}<70%，等待财报完整披露）")
    log.info(f"egs_base 权重 profile={_ih_profile}（active 生效；行业热度权重={_ih_weights['industry_heat']}；legacy 仅回滚锚）")

    ch_mask = df.get("chasing_high", pd.Series(False, index=df.index)).fillna(False)
    oh_mask = df.get("overheat_flag", pd.Series(False, index=df.index)).fillna(False)

    def _join_reasons(reasons):
        reasons = [r for r in reasons if r]
        return "|".join(reasons) if reasons else ""

    df["downgrade_reasons"] = ""
    df["score_penalty_reasons"] = ""
    df.loc[ch_mask, "downgrade_reasons"] = df.loc[ch_mask].apply(
        lambda r: _join_reasons([r.get("downgrade_reasons"), "chasing_high"]), axis=1)
    df.loc[oh_mask, "downgrade_reasons"] = df.loc[oh_mask].apply(
        lambda r: _join_reasons([r.get("downgrade_reasons"), "overheat"]), axis=1)
    df.loc[df["l4_flag"].str.contains("TIER2_FORCED", na=False), "downgrade_reasons"] = \
        df.loc[df["l4_flag"].str.contains("TIER2_FORCED", na=False)].apply(
            lambda r: _join_reasons([r.get("downgrade_reasons"), "l4_score_zero"]), axis=1)
    df.loc[df["l2_name"] == "未知", "downgrade_reasons"] = df.loc[df["l2_name"] == "未知"].apply(
        lambda r: _join_reasons([r.get("downgrade_reasons"), "unknown_industry"]), axis=1)
    if fin_coverage >= 0.70:
        df.loc[esp_neg_mask, "downgrade_reasons"] = df.loc[esp_neg_mask].apply(
            lambda r: _join_reasons([r.get("downgrade_reasons"), "esp_raw_non_positive"]), axis=1)
    df.loc[df["low_base_growth_flag"], "score_penalty_reasons"] = "esp_raw_cap_200"

    # One governed selector serves production and all future P5 arms.  It
    # preserves the prior Tier1/order/concentration behavior while preventing
    # a comparison-only profile from silently using a different pool shape.
    top_df = select_profile_watch_pool(df, top_n=CONF["top_n"])
    return df, top_df


# ═══════════════════════════════════════════════════
# §7 第三级漏斗(减持/解禁 production veto + cninfo 监管 advisory;POL-RISK 已于 Slice3 移除)
# ═══════════════════════════════════════════════════
# POL-RISK-VETO(DeepSeek 行业政策硬否决)及其 prompt-injection helpers 已于 Slice 3 reconciliation
# (2026-06-20)移除——违反「web/LLM 绝不自动 hard-veto」原则,且主源新浪 roll 失效后基本失效。非生产政策/
# 语义 advisory 已由 weekly M6.7 的 DeepSeek adapter(advisory-only)承担;cninfo 监管检查降为 advisory(见
# stage3_ai_clearing,不删生产候选)。要做成真正的生产监管硬否决须另开 opt-in slice(修请求形态 + PIT + governance + 测)。
def stage3_ai_clearing(top50_df, red_dict, unlock_set, backtest_mode=False):
    import requests as _requests

    if top50_df.empty:
        return top50_df, {}
    df = top50_df.copy()

    def _finalize_stage3(final_df, cninfo_checked):
        if final_df.empty:
            log.info("[Stage3] 三级漏斗完成，无满足条件股票")
            return final_df, cninfo_checked

        final_df = final_df.sort_values("final_score", ascending=False).reset_index(drop=True)
        pool, l1c, l2c = [], {}, {}
        for _, row in final_df.iterrows():
            l1  = row.get("l1_name","未知")
            l2  = row.get("l2_name","未知")
            l1k = l2 if l1 == "未知" else l1
            n   = max(len(pool), 1)
            if l1c.get(l1k, 0) / n > 0.4: continue
            if l2c.get(l2,  0) / n > 0.3: continue
            pool.append(row)
            l1c[l1k] = l1c.get(l1k, 0) + 1
            l2c[l2]  = l2c.get(l2,  0) + 1

        result = pd.DataFrame(pool).head(5)
        log.info(f"[Stage3] 三级漏斗完成，最终Tier1候选池 {len(result)} 只")
        return result, cninfo_checked

    # ① 大股东减持一票否决
    veto_10d = red_dict.get("veto_10d", set())
    mask_red = df["ts_code"].isin(veto_10d)
    for _, row in df[mask_red].iterrows():
        log.info(f"[Stage3] {row['ts_code']} {row.get('name','')} → REDUCTION-VETO（10日内大股东减持）")
    df = df[~mask_red].copy()

    # ② 解禁一票否决
    mask_unlock = df["ts_code"].isin(unlock_set)
    for _, row in df[mask_unlock].iterrows():
        log.info(f"[Stage3] {row['ts_code']} {row.get('name','')} → UNLOCK-VETO（未来30日大额解禁）")
    df = df[~mask_unlock].copy()

    if backtest_mode:
        cninfo_checked = {code: "回测跳过" for code in df["ts_code"].tolist()} if "ts_code" in df.columns else {}
        df["cninfo_flag"] = "回测跳过"
        log.info("[Stage3] backtest mode: skip cninfo advisory check")
        return _finalize_stage3(df, cninfo_checked)

    # ③ 巨潮资讯监管公告检查
    REGULATOR_KEYWORDS = ["问询函","立案调查","监管关注","警示函"]

    def _cninfo_check(ts_code):
        stock_code = ts_code.split(".")[0]
        market     = "sz" if ts_code.endswith(".SZ") else "sh"
        column     = "szse" if market == "sz" else "sse"
        d30        = dstr(30)
        start_str  = f"{d30[:4]}-{d30[4:6]}-{d30[6:]}"
        end_str    = f"{TODAY[:4]}-{TODAY[4:6]}-{TODAY[6:]}"
        try:
            resp = _requests.post(
                "http://www.cninfo.com.cn/new/hisAnnouncement/query",
                data={
                    "stock":     f"{stock_code},{market}",
                    "tabName":   "fulltext",
                    "pageSize":  30,
                    "pageNum":   1,
                    "column":    column,
                    "category":  "",
                    "plate":     market,
                    "seDate":    f"{start_str} ~ {end_str}",
                    "searchkey": "",
                    "secid":     "",
                    "sortName":  "",
                    "sortType":  "",
                    "isHLtitle": True,
                },
                headers={
                    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Referer":      "http://www.cninfo.com.cn/new/disclosure/stock",
                },
                timeout=10,
            )
            if resp.status_code != 200:
                return None, None
            anns = resp.json().get("announcements") or []
            if not anns:
                return None, None   # 200 但空公告:legacy 形态(stock=code,market 非契约 code,orgId)可能失败 → 无法证明 clear, 标「未核查」非「通过」(修假清白)
            for ann in anns:
                title = ann.get("announcementTitle","")
                for kw in REGULATOR_KEYWORDS:
                    if kw in title:
                        return True, f"{kw}（{title[:60]}）"
            return False, None
        except Exception as e:
            log.warning(f"[Stage3] 巨潮爬取失败 {ts_code}: {e}")
            return None, None

    # cninfo 监管检查降为 advisory-only(Slice 3 reconciliation, 2026-06-20):**不再删生产候选**。
    # legacy 曾对命中 REGULATOR_KEYWORDS 的候选硬删除(production veto);与「web/LLM/语义 advisory-only、
    # 生产硬否决须 opt-in 重建」原则一致,降为仅写 cninfo_flag(advisory 展示,进 M6.7),不改生产候选池。
    cninfo_checked = {}   # ts_code -> cninfo_flag，供调用方回写 watch_df(advisory 展示)
    df["cninfo_flag"] = "未检查"
    for idx, row in df.iterrows():
        hit, reason = _cninfo_check(row["ts_code"])
        if hit is True:
            df.at[idx, "cninfo_flag"] = reason
            cninfo_checked[row["ts_code"]] = reason
            log.info(f"[Stage3] {row['ts_code']} {row.get('name','')} → REGULATOR-ADVISORY（{reason}; advisory, 不删生产候选）")
        elif hit is False:
            df.at[idx, "cninfo_flag"] = "通过"
            cninfo_checked[row["ts_code"]] = "通过"
        # hit is None(HTTP 失败/异常/200 空公告)→ 无法证明 clear, 保留默认「未检查」(不伪装通过)

    # ④ DeepSeek 行业政策 POL-RISK-VETO 已于 Slice 3 reconciliation(2026-06-20)**整段移除**(见上方注释):
    # web/LLM 绝不自动生产硬否决;非生产政策/语义 advisory 由 weekly M6.7 DeepSeek adapter 承担。
    return _finalize_stage3(df, cninfo_checked)


# ═══════════════════════════════════════════════════
# §8 市场环境
# ═══════════════════════════════════════════════════
def market_environment(trade_dates, stats_df):
    north_flow = None
    try:
        df_hsgt = safe_api(pro.moneyflow_hsgt, start_date=trade_dates[4], end_date=trade_dates[0])
        if df_hsgt is not None and not df_hsgt.empty and "north_money" in df_hsgt.columns:
            df_hsgt["north_money"] = pd.to_numeric(df_hsgt["north_money"], errors="coerce")
            north_flow = df_hsgt["north_money"].sum()
    except Exception:
        pass

    env = []
    if north_flow is not None:
        env.append(f"北向资金近一周净流入: {north_flow/1e8:.2f} 亿")
        if north_flow < -50e8:
            env.append("[!!] 北向资金大幅流出，防御信号。")
    else:
        env.append("北向资金数据不可用")

    csi300_ret = get_csi300_return(trade_dates)
    if csi300_ret is not None and csi300_ret < -10 and (north_flow is not None and north_flow < 0):
        env.append("[静默] 市场进入防御/收缩期：建议静默，禁止开新仓。")
    elif csi300_ret is not None and csi300_ret < -5:
        env.append("[警] 市场偏弱，注意仓位控制。")
    else:
        env.append("[OK] 市场环境正常。")

    # 动量因子有效性判断
    cross_std = None
    if stats_df is not None and not stats_df.empty and "pct_20d" in stats_df.columns:
        pct_series = pd.to_numeric(stats_df["pct_20d"], errors="coerce").dropna()
        if len(pct_series) > 10:
            cross_std = float(pct_series.std())

    if csi300_ret is not None and cross_std is not None:
        if csi300_ret > 0 and cross_std > CONF["momentum_std_threshold"]:
            env.append("[OK] 动量因子当前有效，趋势分化明显")
        elif csi300_ret > 0 and cross_std <= CONF["momentum_std_threshold"]:
            env.append("[!!] 市场普涨但分化不足，动量因子效果有限")
        else:
            env.append("[!!] 市场下行，动量因子参考价值降低，建议谨慎")
    else:
        env.append("动量因子有效性：数据不足，无法判断")

    env.append("融资过热判断：待接入两融余额历史分位")
    env.append("全市场风险提示：请结合当日政策新闻判断")
    return "\n".join(env)


# ═══════════════════════════════════════════════════
# §9 主引擎
# ═══════════════════════════════════════════════════
def run_egs(backtest_mode=False, output_root=None):
    if output_root:
        CONF["output_root"] = output_root
        # Isolate intermediate egs_tier1/egs_full CSV+XLSX artifacts to the backtest
        # tree so they cannot overwrite the official A-EGS/Result/ files when a
        # backtest as_of collides with a production run date.
        project_root = os.path.dirname(SCRIPT_DIR)
        base_root = output_root if os.path.isabs(output_root) else os.path.join(project_root, output_root)
        CONF["result_dir"] = os.path.join(base_root, "_intermediate")
        CONF["xlsx_dir"] = CONF["result_dir"]
        os.makedirs(CONF["result_dir"], exist_ok=True)
    log.info("═" * 60)
    log.info(f"  EGS {EGS_VERSION} 量化选股框架 — 周频增强版")
    log.info("═" * 60)

    trade_dates = get_trade_dates(65)   # v7.5: 需要60日数据支持pct_60d
    latest_td   = trade_dates[0]

    df_stocks  = get_stock_list()
    sw_map     = get_sw_industry_map()
    csi300_ret = get_csi300_return(trade_dates)
    all_daily  = get_daily_all(trade_dates)

    # ── 上期候选股追踪（v7.5扩展：周内高低点、最大收益/回撤、是否仍在池）──────────
    import json as _json
    # The former tracker persisted raw closes.  Do not compare those legacy
    # records with the qfq production series after this price-basis migration.
    _LAST_SEL_FILE = _rp("egs_last_selection_qfq_v1.json")
    _last_sel_raw   = []
    _last_sel_report = []
    if (not backtest_mode) and os.path.exists(_LAST_SEL_FILE):
        try:
            with open(_LAST_SEL_FILE, "r", encoding="utf-8") as _f:
                _last_sel_raw = _json.load(_f)
            if _last_sel_raw:
                _latest_day = pd.DataFrame()
                if not all_daily.empty and "trade_date" in all_daily.columns:
                    _latest_day = all_daily[all_daily["trade_date"] == latest_td]
                    if _latest_day.empty and len(trade_dates) > 1:
                        _latest_day = all_daily[all_daily["trade_date"] == trade_dates[1]]
                _cur_prices = {}
                if not _latest_day.empty and "qfq_close" in _latest_day.columns:
                    _cur_prices = dict(zip(
                        _latest_day["ts_code"],
                        pd.to_numeric(_latest_day["qfq_close"], errors="coerce")
                    ))
                for _item in _last_sel_raw:
                    _code      = _item.get("ts_code", "")
                    _ref_price = _item.get("close")
                    _run_date  = _item.get("run_date", "")
                    _cur       = _cur_prices.get(_code)
                    if _ref_price and _cur is not None and not pd.isna(_cur):
                        _ret_pct = (_cur / _ref_price - 1) * 100
                        _ret_str = f"{_ret_pct:+.2f}%"
                    else:
                        _ret_pct, _ret_str = None, "N/A"

                    # 周内最高/最低（从run_date到今天的all_daily数据）
                    _week_data = pd.DataFrame()
                    if not all_daily.empty and "trade_date" in all_daily.columns:
                        _week_data = all_daily[
                            (all_daily["ts_code"] == _code) &
                            (all_daily["trade_date"] > _run_date) &
                            (all_daily["trade_date"] <= latest_td)
                        ]
                    _week_high_p = float(_week_data["qfq_high"].max()) if not _week_data.empty and "qfq_high" in _week_data.columns else None
                    _week_low_p  = float(_week_data["qfq_low"].min())  if not _week_data.empty and "qfq_low"  in _week_data.columns else None
                    if _ref_price and _week_high_p:
                        _max_gain = f"{(_week_high_p/_ref_price-1)*100:+.2f}%"
                    else:
                        _max_gain = "N/A"
                    if _ref_price and _week_low_p:
                        _max_draw = f"{(_week_low_p/_ref_price-1)*100:+.2f}%"
                    else:
                        _max_draw = "N/A"

                    _last_sel_report.append({
                        "代码":       _code,
                        "名称":       _item.get("name", ""),
                        "状态":       "仍在池" if _item.get("still_in_pool", True) else "已落选",
                        "记录日期":   _run_date or "N/A",
                        "上期评分":   f"{_item.get('final_score', 0):.2f}",
                        "入池参考价": f"{_ref_price:.2f}" if _ref_price else "N/A",
                        "当前价":     f"{_cur:.2f}" if _cur is not None and not pd.isna(_cur) else "N/A",
                        "当前收益":   _ret_str,
                        "周内最大涨": _max_gain,
                        "周内最大回": _max_draw,
                    })
        except Exception as _e:
            log.warning(f"上期候选股记录读取失败: {_e}")
    # ─────────────────────────────────────────────────────────────────────────

    # [崩溃修复①] 传入 trade_dates 供回退使用
    df_db = get_daily_basic(latest_td, trade_dates)

    suspended_set = get_suspend_info(trade_dates)
    relisted_set  = get_relisted_stocks(trade_dates)

    # [崩溃修复②] get_unlock_future 内置前置防御
    unlock_set = get_unlock_future(df_stocks, df_db)
    red_dict   = get_holder_reductions()

    all_codes = set(df_stocks["ts_code"])
    stats_df  = precompute_stock_stats(all_codes, all_daily)

    df_l0 = filter_l0(df_stocks, stats_df, unlock_set, red_dict, suspended_set, relisted_set)

    # 行业 ESP 基准(global_ind_med)**有意用全行业样本**(df_stocks 全 universe,非主板过滤后的 df_l0)——
    # 用户 2026-06-20 拍板「含全行业」:ChiNext/STAR 等是同 SW 行业的合法成员,纳入使行业中位数更稳健
    # (主板-only 会让某些 SW L2 主板票<5 而 null 掉、丢 ESP)。**这是有意设计、非主板边界 bug**(Codex S2#2 据此驳回)。
    # B 股(200/900)无 SW 映射 → 落「未知」桶,不污染候选所在真实行业的中位数;候选打分本身仍只限主板(filter_l0 strict)。
    full_codes = df_stocks["ts_code"].tolist()
    df_raw_fin = get_financial_data(full_codes)
    if not df_raw_fin.empty and "q0_dt_yoy" in df_raw_fin.columns:
        df_raw_fin["q0_dt_yoy_n"] = pd.to_numeric(df_raw_fin["q0_dt_yoy"], errors="coerce")
        df_raw_fin["l2_name"]     = df_raw_fin["ts_code"].map(
            lambda c: sw_map.get(c, {}).get("l2_name", "未知"))
        valid = df_raw_fin[df_raw_fin["q0_dt_yoy_n"].notna()]
        meds  = valid.groupby("l2_name")["q0_dt_yoy_n"].median()
        cnts  = valid.groupby("l2_name")["q0_dt_yoy_n"].count()
        global_ind_med = {
            n: float(meds[n]) if n in meds and cnts.get(n, 0) >= 5 else None
            for n in meds.index
        }
    else:
        global_ind_med = {}

    mf_df = get_moneyflow(trade_dates)
    mg_df = get_margin(trade_dates)

    df_master = build_master(
        df_l0, stats_df, df_db, df_fin=df_raw_fin, sw_map=sw_map, red_dict=red_dict
    )
    l1_exclusion_reasons = {}
    df_l1 = score_l1(
        df_master, csi300_ret, exclusion_reasons=l1_exclusion_reasons
    )
    l2_exclusion_reasons = {}
    df_l2 = score_l2(
        df_l1, mg_df, trade_dates, global_ind_med,
        exclusion_reasons=l2_exclusion_reasons,
    )
    df_l3 = score_l3(df_l2, trade_dates, all_daily)
    df_l4 = score_l4(df_l3, mf_df)
    df_full, top50 = score_l5(df_l4, sw_map)

    rank_reconciliation, rank_reconciliation_detail = build_rank_universe_reconciliation(
        df_l0=df_l0,
        stages=[
            ("master_join", df_master, False, "master_join_loss"),
            ("l1_industry_leader", df_l1, True, l1_exclusion_reasons),
            ("l2_quality_risk", df_l2, True, l2_exclusion_reasons),
            ("l3_scoring", df_l3, False, "l3_unexpected_row_loss"),
            ("l4_scoring", df_l4, False, "l4_unexpected_row_loss"),
            ("l5_rank", df_full, False, "l5_unexpected_row_loss"),
        ],
        sources={
            "daily_stats_l0": (df_l0, stats_df, 1.0),
            "daily_basic_l0": (df_l0, df_db, 1.0),
            "financial_l0": (df_l0, df_raw_fin, 1.0),
            "financial_full_universe": (df_stocks, df_raw_fin, 0.95),
        },
    )
    if rank_reconciliation["status"] != "pass":
        raise RuntimeError(
            "rank universe reconciliation failed: "
            f"source_coverage_failures={rank_reconciliation['source_coverage_failure_count']}, "
            f"unexpected_stage_changes={rank_reconciliation['unexpected_stage_change_count']}, "
            f"unaccounted={rank_reconciliation['unaccounted_count']}, "
            f"duplicate_codes={rank_reconciliation['duplicate_code_count']}"
        )

    # The comparison sidecar is written later inside the final official-output
    # transaction.  A failed sidecar must not block EGS, but it must also never
    # be accidentally bound into a fresh official publish marker as stale bytes.
    comparison_sidecar_warnings = []
    _weight_comparison_published = False
    _p4_overlay_score_path = None
    _p4_overlay_inputs = None

    tier1_final, cninfo_checked = stage3_ai_clearing(top50, red_dict, unlock_set, backtest_mode=backtest_mode)
    env_report  = market_environment(trade_dates, stats_df)

    # ── 计算 entry_flag ────────────────────────────────────────────────────────
    def _entry_flag(row):
        if row.get("overheat_flag", False) or row.get("chasing_high", False):
            return "追高风险，周一确认"
        br = row.get("big_ratio", np.nan)
        if not pd.isna(br) and br < -0.05:
            return "资金流背离"
        cs = row.get("cat_score", 50)
        p5 = row.get("pct_5d_n", np.nan)
        if cs > 85 and not pd.isna(p5) and p5 > 5:
            return "题材过热"
        if "LOCK" in str(row.get("l4_flag", "")):
            return "需周一确认"
        return "可直接观察"
    # ─────────────────────────────────────────────────────────────────────────

    if _last_sel_report:
        print("\n" + "═" * 60)
        print("  上期候选股追踪（v7.5 扩展）")
        print("═" * 60)
        print(pd.DataFrame(_last_sel_report).to_string(index=False))

    print("\n" + "═" * 60)
    print("  市场环境")
    print("═" * 60)
    print(env_report)

    # ── Top 15 候选观察池 ─────────────────────────────────────────────────────
    watch_n   = CONF["watch_n"]
    watch_df  = select_profile_watch_pool(df_full, top_n=watch_n)
    watch_eligible_count = int(len(top50))

    # backtest 模式下，若 Tier1 候选不足 watch_n（常见于 esp 准入硬条件激活时），
    # 从 Tier2 按 final_score 补足，让回测样本量稳定在 watch_n。
    # 正式运行不动，避免污染周末观察池的 Tier1 语义。
    if backtest_mode and len(watch_df) < watch_n:
        existing_codes = set(watch_df["ts_code"].tolist()) if "ts_code" in watch_df.columns else set()
        tier2_fill = df_full[(df_full["tier"] == "Tier2") &
                             (df_full["l2_name"] != "未知") &
                             (~df_full["ts_code"].isin(existing_codes))] \
            .sort_values(["final_score", "l4_score", "pct_20d_n"],
                         ascending=[False, False, False]) \
            .head(watch_n - len(watch_df))
        if not tier2_fill.empty:
            log.info(f"[BACKTEST] watch_df Tier1 仅 {len(watch_df)} 只，从 Tier2 补 {len(tier2_fill)} 只达 watch_n={watch_n}")
            watch_df = pd.concat([watch_df, tier2_fill[watch_df.columns]], ignore_index=True)
            watch_eligible_count = max(watch_eligible_count, int(len(watch_df)))

    # 赛道热度 overlay(Slice A,comparison-track 非生产):只覆盖 Stage3 后最终周报候选 watch_df，
    # 与 analysis_input.candidates 保持同一批，避免 M6.7 因多行/缺行而 fail-closed。
    # 生产行业热度评分已在 score_l5 完成；本 overlay 仅供后续分析/对照，不反向改变候选排序。
    watch_df["theme_taxonomy"] = [
        unavailable_theme_taxonomy(
            TODAY,
            "l3_taxonomy_not_available_for_this_run",
            l3_provider=CONF.get("l3_provider"),
            l3_snapshot_date=CONF.get("l3_snapshot_date"),
            l3_coverage=CONF.get("l3_coverage"),
        )
        for _ in range(len(watch_df))
    ]
    try:
        from runners.a_short_theme_overlay_comparison import emit_overlay, overlay_emit_allowed
        from engine.a_short_run_paths import overlay_path
        _l3 = (_load_l3_snapshot(TODAY, include_metadata=True)
               if overlay_emit_allowed(CONF.get("l3_mode")) else None)
        if _l3 is not None:
            _taxonomy_by_code = taxonomy_by_code(
                watch_df, stock_concepts=_l3[1], concept_members=_l3[2],
                concepts_df=_l3[0], as_of=TODAY,
                l3_provider=_l3[4], l3_snapshot_date=_l3[3], l3_coverage=_l3[5],
            )
            watch_df["theme_taxonomy"] = watch_df["ts_code"].astype(str).map(_taxonomy_by_code)
            _ov_pool = watch_df[["ts_code", "esp_score", "l4_score", "overheat_flag", "chasing_high"]].copy()
            _ov_pool["baseline_rank"] = range(1, len(_ov_pool) + 1)
            _ov_gen = datetime.now().astimezone().isoformat(timespec="seconds")
            _ov_written = emit_overlay(CONF.get("l3_mode"), _ov_pool, all_daily, _l3, sw_map,
                                       TODAY, _ov_gen, overlay_path(TODAY, output_root=output_root))
            # P4a needs scores for the immutable Stage3-eligible pool, not the
            # Top15 watch pool consumed by M6.7.  Prepare the comparison-only
            # inputs now, but write no P4 file until the official-output
            # transaction has started below.
            try:
                _p4_eligible = top50[(~top50["ts_code"].isin(set((red_dict or {}).get("veto_10d", set())))) &
                                      (~top50["ts_code"].isin(set(unlock_set or set())))].copy()
                _p4_pool = _p4_eligible[["ts_code", "esp_score", "l4_score", "overheat_flag", "chasing_high"]].copy()
                _p4_pool["baseline_rank"] = range(1, len(_p4_pool) + 1)
                _p4_overlay_inputs = (_p4_pool, _l3, _ov_gen)
            except Exception as _p4_overlay_exc:
                log.warning("P4a Stage3 overlay inputs unavailable; formal EGS output unchanged: %s",
                            safe_exception_summary(_p4_overlay_exc))
            # Preserve comparison metrics with the governed taxonomy for the
            # later forward evaluator. This never feeds ranking or M6.7.
            with open(_ov_written, encoding="utf-8") as _ov_handle:
                _ov_summary = _json.load(_ov_handle)
            _ov_metrics = {str(item.get("ts_code")): item for item in (_ov_summary.get("candidates") or [])}

            def _taxonomy_with_metrics(row):
                _base = dict(row.get("theme_taxonomy") or unavailable_theme_taxonomy(
                    TODAY,
                    "taxonomy_missing_after_overlay",
                    l3_provider=CONF.get("l3_provider"),
                    l3_snapshot_date=CONF.get("l3_snapshot_date"),
                    l3_coverage=CONF.get("l3_coverage"),
                ))
                _metric = _ov_metrics.get(str(row.get("ts_code"))) or {}
                _base["comparison_metrics"] = {
                    "theme_heat_score": _metric.get("theme_heat_score"),
                    "breadth_pass": bool(_metric.get("breadth_pass")),
                    "persistence_mult": _metric.get("persistence_mult"),
                    "fit_score": _metric.get("fit_score"),
                    "fit_pass": bool(_metric.get("fit_pass")),
                    "comparison_status": "available",
                }
                return _base

            watch_df["theme_taxonomy"] = watch_df.apply(_taxonomy_with_metrics, axis=1)
            log.info(f"赛道热度 overlay 已写(非生产,comparison-track）：{_ov_written}")
        else:
            log.info("赛道热度 overlay 跳过:无 L3 快照(l3_mode=neutralize 或无快照),不编造概念")
    except Exception as _ov_exc:  # noqa: BLE001 (non-production side output must never break the run)
        _ov_warning = _comparison_sidecar_warning("theme_overlay", _ov_exc)
        comparison_sidecar_warnings.append(_ov_warning)
        log.warning(_ov_warning["message"])

    watch_df["entry_flag"]  = watch_df.apply(_entry_flag, axis=1)
    if "ts_code" in watch_df.columns:
        watch_df["cninfo_flag"] = watch_df["ts_code"].map(cninfo_checked).fillna("未检查")
    else:
        watch_df["cninfo_flag"] = "未检查"
    rule6_evaluations_by_code = _collect_rule6_evaluations(
        watch_df, all_daily, mg_df, trade_dates, red_dict, df_raw_fin,
    )

    watch_cols = ["ts_code","name","l2_name","final_score","tier",
                  "pct_20d_n","pct_5d_n","pct_60d","drawdown_20d",
                  "cat_score","l4_score","esp_score",
                  "l4_flag","cninfo_flag","entry_flag",
                  "downgrade_reasons","score_penalty_reasons"]
    watch_cols = [c for c in watch_cols if c in watch_df.columns]

    print("\n" + "═" * 60)
    print(
        f"  候选观察池 (EGS {EGS_VERSION})  {len(watch_df)}/{watch_n} "
        "← 合格池不足时不拿 Tier2 凑数；周末分析用，周一开盘前确认"
    )
    print("═" * 60)
    if watch_df.empty:
        print("  [!!] 本次无候选标的。")
    else:
        print(watch_df[watch_cols].to_string(index=False))

    # ── Top 5 最终推荐 ────────────────────────────────────────────────────────
    tier1_final["entry_flag"] = tier1_final.apply(_entry_flag, axis=1)

    out_cols = ["ts_code","name","l2_name","final_score","tier",
                "esp_score","cat_score","l4_score","industry_heat_score",
                "pct_20d_n","pct_5d_n","pct_60d","drawdown_20d",
                "big_ratio","ind_mom_cnt",
                "l2_flags","l4_flag","cninfo_flag","entry_flag",
                "downgrade_reasons","score_penalty_reasons",
                "reduce_penalty","val_bonus","val_penalty"]
    out_cols = [c for c in out_cols if c in tier1_final.columns]

    print("\n" + "═" * 60)
    print(f"  最终推荐 (EGS {EGS_VERSION})  Top {min(len(tier1_final), CONF['final_n'])}  ← 结合分析框架深度分析")
    print("═" * 60)
    if tier1_final.empty:
        print("  [!!] 本次未筛选出满足条件的 Tier 1 标的。")
    else:
        print(tier1_final[out_cols].head(CONF["final_n"]).to_string(index=False))

    tier1_csv_path = _rp(f"egs_tier1_{TODAY}.csv")
    full_csv_path  = _rp(f"egs_full_{TODAY}.csv")
    tier1_xlsx_path = os.path.join(CONF.get("xlsx_dir", SCRIPT_DIR), f"egs_tier1_{TODAY}.xlsx")

    project_root = os.path.dirname(SCRIPT_DIR)
    configured_root = CONF.get("output_root")
    base_root = (
        configured_root if configured_root and os.path.isabs(configured_root)
        else os.path.join(project_root, configured_root) if configured_root
        else os.path.join(project_root, "result", "a_short")
    )
    official_dir = os.path.join(base_root, latest_td)
    rank_reconciliation_path = os.path.join(
        official_dir, "rank_universe_reconciliation.csv"
    )
    transaction_paths = [
        tier1_csv_path,
        tier1_xlsx_path,
        full_csv_path,
        os.path.join(official_dir, "analysis_input.json"),
        os.path.join(official_dir, "snapshot.json"),
        os.path.join(official_dir, "stage3_selection_snapshot.json"),
        os.path.join(official_dir, "stage3_overlay_score.json"),
        os.path.join(official_dir, "candidates.csv"),
        os.path.join(official_dir, "data_health.json"),
        weight_comparison_path(TODAY, output_root=CONF.get("output_root")),
        rank_reconciliation_path,
        os.path.join(official_dir, "official_publish.json"),
    ]
    publish_context = nullcontext() if backtest_mode else official_output_transaction(transaction_paths)
    with publish_context:
        write_csv_atomic(watch_df[watch_cols], tier1_csv_path, index=False, encoding="utf-8-sig")
        save_ranked_xlsx(watch_df[watch_cols], tier1_xlsx_path, group_size=5)
        write_csv_atomic(df_full, full_csv_path, index=False, encoding="utf-8-sig")
        write_csv_atomic(
            rank_reconciliation_detail,
            rank_reconciliation_path,
            index=False,
            encoding="utf-8-sig",
        )
        analysis_path, snapshot_path, candidates_path, analysis_input = export_analysis_input(
            df_full=df_full,
            watch_df=watch_df,
            tier1_final=tier1_final,
            latest_td=latest_td,
            trade_dates=trade_dates,
            unlock_set=unlock_set,
            suspended_set=suspended_set,
            relisted_set=relisted_set,
            red_dict=red_dict,
            tier1_csv_path=tier1_csv_path,
            full_csv_path=full_csv_path,
            output_root=CONF.get("output_root"),
            rank_reconciliation=rank_reconciliation,
            rule6_evaluations_by_code=rule6_evaluations_by_code,
        )
        log.info(f"[OK] analysis_input saved to {analysis_path}")
        log.info(f"[OK] snapshot saved to {snapshot_path}")
        log.info(f"[OK] candidates saved to {candidates_path}")
        _p4_stage3_snapshot_path = None
        try:
            _p4_stage3_snapshot_path = export_stage3_selection_snapshot(
                top50, tier1_final, latest_td,
                (analysis_input.get("source") or {}).get("run_identity") or {},
                red_dict, unlock_set, output_root=CONF.get("output_root"),
            )
            log.info(f"[OK] P4a Stage3 snapshot saved to {_p4_stage3_snapshot_path}")
        except Exception as _p4_snapshot_exc:  # comparison evidence cannot mutate EGS output
            _p4_warning = _comparison_sidecar_warning("p4_stage3_snapshot", _p4_snapshot_exc)
            log.warning(_p4_warning["message"])
        # P4's marker may bind its scorer only together with the matching
        # Stage3 selection receipt.  A failed snapshot therefore suppresses
        # the scorer sidecar rather than publishing a one-sided bundle.
        if _p4_stage3_snapshot_path is not None and _p4_overlay_inputs is not None:
            _p4_pool, _p4_l3, _p4_generated_at = _p4_overlay_inputs
            _p4_overlay_target = os.path.join(official_dir, "stage3_overlay_score.json")
            try:
                # The nested transaction makes a failed P4 sidecar restore its
                # own prior bytes; a rollback failure propagates to the outer
                # formal transaction rather than leaving an unbound artifact.
                with official_output_transaction([_p4_overlay_target]):
                    _p4_overlay_score_path = emit_overlay(
                        CONF.get("l3_mode"), _p4_pool, all_daily, _p4_l3, sw_map,
                        TODAY, _p4_generated_at, _p4_overlay_target,
                    )
            except Exception as _p4_overlay_exc:  # P4a must not alter official output bytes on failure.
                _p4_overlay_score_path = None
                if "rollback was incomplete" in str(_p4_overlay_exc):
                    raise
                log.warning("P4a Stage3 overlay sidecar unavailable; formal EGS output unchanged: %s",
                            safe_exception_summary(_p4_overlay_exc))
        # Same output transaction as analysis_input and its final marker: P5 may
        # later consume this file only when the marker binds these exact bytes.
        try:
            _wc_path = weight_comparison_path(TODAY, output_root=CONF.get("output_root"))
            write_weight_comparison(df_full, _wc_path, as_of=TODAY)
            _weight_comparison_published = True
            log.info(f"egs 权重 variant 对比 diff 已写(非生产）：{_wc_path}")
        except Exception as _wc_exc:  # noqa: BLE001 (sidecar remains non-blocking for EGS/M6.7)
            _wc_warning = _comparison_sidecar_warning("weight_variant", _wc_exc)
            comparison_sidecar_warnings.append(_wc_warning)
            log.warning(_wc_warning["message"])
        if backtest_mode:
            log.info("[BACKTEST] skip egs_last_selection tracking state")
            return tier1_final
        health_path, health = export_data_health(
            df_full=df_full,
            watch_df=watch_df,
            tier1_final=tier1_final,
            analysis_input=analysis_input,
            latest_td=latest_td,
            analysis_path=analysis_path,
            snapshot_path=snapshot_path,
            candidates_path=candidates_path,
            tier1_csv_path=tier1_csv_path,
            full_csv_path=full_csv_path,
            rank_reconciliation=rank_reconciliation,
            rank_reconciliation_path=rank_reconciliation_path,
            watch_eligible_count=watch_eligible_count,
            sidecar_warnings=comparison_sidecar_warnings,
        )
        log_data_health_summary(health_path, health)
        _published_files = {
            "analysis_input": analysis_path,
            "snapshot": snapshot_path,
            "candidates": candidates_path,
            "data_health": health_path,
            "tier1_csv": tier1_csv_path,
            "full_rank": full_csv_path,
            "rank_universe_reconciliation": rank_reconciliation_path,
        }
        if _p4_stage3_snapshot_path:
            _published_files["p4_stage3_selection_snapshot"] = _p4_stage3_snapshot_path
        if _p4_overlay_score_path:
            _published_files["p4_stage3_overlay_score"] = _p4_overlay_score_path
        if _weight_comparison_published:
            _published_files["egs_weight_comparison"] = weight_comparison_path(
                TODAY, output_root=CONF.get("output_root")
            )
        marker_path, _manifest = publish_egs_run_manifest(
            analysis_input,
            health,
            _published_files,
        )
    log.info(f"[OK] official EGS publish marker saved to {marker_path}")
    log.info(f"[OK] 结果已保存至 {tier1_csv_path} / {tier1_xlsx_path} / {full_csv_path}")

    # ── 保存本次候选池记录供下次追踪（v7.5扩展至Top15）──────────────────────────
    try:
        _save_day = all_daily[all_daily["trade_date"] == latest_td]
        if _save_day.empty and len(trade_dates) > 1:
            _save_day = all_daily[all_daily["trade_date"] == trade_dates[1]]
        _save_prices = {}
        if not _save_day.empty and "qfq_close" in _save_day.columns:
            _save_prices = dict(zip(
                _save_day["ts_code"],
                pd.to_numeric(_save_day["qfq_close"], errors="coerce")
            ))
        # 本次Top15的ts_code集合，用于标注"是否仍在池"
        _this_pool = set(watch_df["ts_code"].tolist()) if "ts_code" in watch_df.columns else set()

        # 本次新记录
        _records = []
        for _, _row in watch_df.iterrows():
            _code  = _row["ts_code"]
            _close = _save_prices.get(_code)
            _records.append({
                "ts_code":       _code,
                "name":          _row.get("name", ""),
                "final_score":   float(_row["final_score"]) if "final_score" in _row and not pd.isna(_row["final_score"]) else None,
                "tier":          str(_row.get("tier", "")),
                "entry_flag":    str(_row.get("entry_flag", "")),
                "cninfo_flag":   str(_row.get("cninfo_flag", "未检查")),
                "close":         float(_close) if _close is not None and not pd.isna(_close) else None,
                "price_basis":   "qfq_anchored_as_of",
                "run_date":      TODAY,
                "still_in_pool": True,
            })

        # 保留上期落选记录（still_in_pool=False）一轮，让追踪报告能显示"本周落选"
        _prior_leavers = [
            r for r in _last_sel_raw
            if not r.get("still_in_pool", True)
               and r.get("ts_code","") not in _this_pool
        ]
        # 标注上上期已落选的记录，避免无限堆积（run_date与本次不同且already out则不再保留）
        _current_leavers = [
            {**r, "still_in_pool": False}
            for r in _last_sel_raw
            if r.get("ts_code","") not in _this_pool
               and r.get("run_date","") != TODAY   # 同日重跑不重复
        ]
        # 仅保留直接上一轮的落选（run_date最新的那批）
        if _current_leavers:
            _last_run_date = max(r.get("run_date","") for r in _current_leavers)
            _current_leavers = [r for r in _current_leavers if r.get("run_date","") == _last_run_date]

        _records_to_save = _records + _current_leavers
        write_json_atomic(_LAST_SEL_FILE, _records_to_save)
        log.info(f"[OK] 本次候选池记录（{len(watch_df)}/{watch_n}）已保存至 {_LAST_SEL_FILE}")
    except Exception as _e:
        log.warning(f"候选池记录保存失败: {_e}")

    return tier1_final


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="EGS A-share short-term screener")
    parser.add_argument("--as-of", dest="as_of", help="Run as of an A-share trading date, format YYYYMMDD")
    parser.add_argument("--backtest-mode", action="store_true", help="Skip mutable tracking state for historical batch runs")
    parser.add_argument(
        "--cache-policy", choices=["enabled", "disabled"], default="enabled",
        help="EGS request cache policy. disabled performs no cache reads or writes (required by runtest capsules).",
    )
    parser.add_argument(
        "--reuse-l3-cache", action="store_true",
        help="Testing only: reuse an existing complete HiThink main-board L3 snapshot; never calls the provider",
    )
    parser.add_argument(
        "--allow-stale-l3-cache", action="store_true",
        help="Testing only: allow --reuse-l3-cache to use a snapshot over 14 days old; never calls the provider",
    )
    parser.add_argument("--l3-mode", dest="l3_mode", choices=["pit", "today", "neutralize"],
                        default="today",
                        help="L3 cat_score source. today (default): fresh complete HiThink main-board snapshot (writes "
                             "state/l3_snapshots/*_{today}.pkl). pit: load latest snapshot <= as_of "
                             "(no API calls). neutralize: cat_score=50.0, skip L3 entirely.")
    parser.add_argument("--l3-pit-strict", action="store_true",
                        help="With --l3-mode=pit, raise SystemExit if no snapshot <= as_of (default: warn + fallback to cat_score=50).")
    parser.add_argument("--allow-historical-live-l3", action="store_true",
                        help="Non-evidence/testing only: allow historical --as-of with --l3-mode=today. "
                             "Default blocks this because today-mode L3 uses live concept data.")
    parser.add_argument("--output-root", dest="output_root", default=None,
                        help="Override base output directory for analysis_input/snapshot/candidates "
                             "(default: <project_root>/result/a_short). Used by backtest to isolate generated artifacts.")
    args = parser.parse_args()

    CONF["cache_policy"] = args.cache_policy

    _guard_historical_asof_l3_mode(
        args.as_of,
        args.l3_mode,
        allow_historical_live_l3=args.allow_historical_live_l3,
    )

    if args.reuse_l3_cache:
        CONF["l3_cache_mode"] = "reuse"
        if args.l3_mode != "today":
            parser.error(
                f"--reuse-l3-cache is only valid with --l3-mode=today, not {args.l3_mode}"
            )
    elif args.allow_stale_l3_cache:
        parser.error("--allow-stale-l3-cache requires --reuse-l3-cache")
    CONF["l3_allow_stale_cache"] = bool(args.allow_stale_l3_cache)
    CONF["l3_mode"] = args.l3_mode
    CONF["l3_pit_strict"] = bool(args.l3_pit_strict)
    if args.l3_pit_strict and args.l3_mode != "pit":
        log.warning(
            f"--l3-pit-strict has no effect under --l3-mode={args.l3_mode}; "
            "this flag only gates the pit-mode missing-snapshot behavior"
        )

    if args.as_of:
        set_asof(args.as_of)
    else:
        CONF["cache_ttl"] = REALTIME_CACHE_TTL

    run_egs(backtest_mode=args.backtest_mode, output_root=args.output_root)
